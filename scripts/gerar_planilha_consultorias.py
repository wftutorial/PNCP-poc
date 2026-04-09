#!/usr/bin/env python3
"""
Gera planilha Excel de consultorias de licitacao para parceria.
Output: docs/reports/consultorias-parceria-2026-03-11.xlsx
"""

import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def gerar_msg_wpp(nome_curto, empresa, gancho, potencial="Alto"):
    """WhatsApp. Alto ~75 palavras, Médio ~50 palavras. Conversão máxima."""
    if potencial == "Alto":
        return (
            f"Oi {nome_curto}, tudo bem?\n\n"
            f"Vi que a {empresa} {gancho}.\n\n"
            f"Sou servidor público efetivo há 7 anos na área de licitações. "
            f"Faço um trabalho de mapeamento de editais: analiso os PDFs "
            f"completos, cruzo habilitação por CNPJ e entrego quais editais "
            f"disputar, quem são os concorrentes e onde estão os gaps na "
            f"documentação.\n\n"
            f"Meu custo é R$97 por análise. Vocês cobram o preço que "
            f"quiserem do cliente.\n\n"
            f"Posso te mandar um exemplo pra avaliar?"
        )
    else:
        return (
            f"Oi {nome_curto}, tudo bem?\n\n"
            f"Vi que a {empresa} {gancho}.\n\n"
            f"Sou servidor público efetivo há 7 anos em licitações. "
            f"Faço mapeamento de editais: analiso PDFs, cruzo habilitação "
            f"por CNPJ e entrego o cenário completo.\n\n"
            f"Meu custo é R$97. Vocês cobram o que quiserem.\n\n"
            f"Mando um exemplo?"
        )


def gerar_msg_email(nome_curto, empresa, gancho):
    """Email. Assunto personalizado, corpo ~90 palavras. Conversão máxima."""
    return (
        f"Assunto: {nome_curto}, ideia pra aumentar margem "
        f"com seus clientes de licitação\n\n"
        f"Oi {nome_curto},\n\n"
        f"Vi que a {empresa} {gancho}.\n\n"
        f"Sou servidor público efetivo há 7 anos na área de licitações. "
        f"Faço um trabalho de mapeamento de editais: analiso os PDFs "
        f"reais, cruzo habilitação com o CNPJ do cliente e entrego "
        f"editais disputáveis, concorrentes e gaps na documentação.\n\n"
        f"Meu custo é R$97 por análise. Vocês embalam como quiserem "
        f"e cobram o preço que fizer sentido pro cliente de vocês. "
        f"Eu faço, vocês entregam.\n\n"
        f"Posso te mandar um exemplo pra avaliar?\n\n"
        f"Abraço,\n"
        f"Tiago Sasaki\n"
        f"(48) 9 8834-4559"
    )


def gerar_msg_followup(nome_curto):
    """Follow-up WhatsApp D+3. Único, 2 linhas."""
    return (
        f"Oi {nome_curto}, mandei um exemplo por email. "
        f"Conseguiu ver?"
    )


def create_workbook():
    wb = Workbook()

    ws = wb.active
    ws.title = "Consultorias"

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    alto_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medio_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Empresa", "Tipo", "Cidade/UF", "Potencial",
        "Quem abordar", "WhatsApp", "Email", "Site",
        "Redes sociais", "Sobre a consultoria", "Por que e bom parceiro",
        "Msg WhatsApp", "Msg Email", "Follow-up D+3", "Status", "Notas",
    ]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda

    # Dados — 20 existentes enriquecidos + 18 novas = 38 consultorias
    consultorias = [
        # ===== EXISTENTES (enriquecidas) =====
        {"empresa": "Licitando Consultoria", "tipo": "Consultoria + Cursos", "local": "Porto Alegre/RS", "potencial": "Alto",
         "quem": "Marcela Moneta (proprietária)", "wpp": "(51) 98101-3520", "email": "marcelamoneta@licitandoconsultoria.com.br",
         "site": "licitandoconsultoria.com.br", "redes": "FB: /licitandooficial",
         "sobre": "20+ anos. Consultoria 360 + cursos de pregão. Assessoria jurídica com advogada 20+ anos. Buscador de editais por email.",
         "parceiro": "Decisora com WhatsApp público. Base de alunos que já querem licitar. Relatório complementa o trabalho consultivo.",
         "nome_curto": "Marcela", "gancho": "oferece consultoria personalizada e cursos de pregão há mais de 20 anos. Seria um serviço prático pros seus alunos"},
        {"empresa": "Route Assessoria", "tipo": "Consultoria + Cursos", "local": "Botucatu/SP", "potencial": "Alto",
         "quem": "Ricardo Ribas da Costa Berloffa (proprietário, advogado)", "wpp": "(11) 98445-9878", "email": "ricardo@routeassessoria.com.br",
         "site": "routeassessoria.com.br", "redes": "IG: @prof.ricardo.ribas / LI: prof-ricardo-ribas / YT: Prof. Ricardo Ribas / Hotmart: 7+ cursos",
         "sobre": "22+ anos ensino. 15.000+ alunos formados. 6 livros sobre licitações. 100+ aulas grátis no YT. Cursos na Hotmart (Nova Lei, ETP com IA). Ex-ARSESP.",
         "parceiro": "Maior educador do nicho. 15K alunos = público exato. Pode revender como serviço pós-curso.",
         "nome_curto": "Ricardo", "gancho": "já formou mais de 15 mil licitantes e tem 6 livros publicados. Esse mapeamento complementa o curso como serviço prático"},
        {"empresa": "Brasil Licitar", "tipo": "Consultoria full-service", "local": "São Paulo/SP", "potencial": "Alto",
         "quem": "L. Gustavo Oliveira (CEO, MBA FIA, Dir. Adm. FGV)", "wpp": "(11) 95894-4465", "email": "comercial@brasillicitar.com.br",
         "site": "brasillicitar.com.br", "redes": "IG: @brasillicitar / LI: gustaoliver / LI empresa: brasil-licitar",
         "sobre": "10+ anos, 500+ licitações ganhas. Jessica Monteiro como consultora empresarial (ex-Itaú 6 anos).",
         "parceiro": "Perfil consultivo forte. Gustavo é decisor acessível com LinkedIn ativo. Atua em todas as etapas.",
         "nome_curto": "Gustavo", "gancho": "já apoiou mais de 500 licitações e atua em todas as etapas. Com esse volume, a triagem dos editais deve consumir bastante tempo"},
        {"empresa": "Concreta Licitações", "tipo": "Consultoria + Franquia", "local": "Rio do Sul/SC + franquias", "potencial": "Alto",
         "quem": "Lucas de Medeiros (CEO/Admin) - decisor principal", "wpp": "(47) 3300-1524", "email": "contato@concretalicitacoes.com.br",
         "site": "concretalicitacoes.com.br", "redes": "IG: @concretalicitacoes (2.285 seg) / LI: lucas-medeiros-7babb1131",
         "sobre": "Fundada 2019. 350+ empresas, R$1B+ em licitações. ~20 func. Acelerada InovAtiva Brasil (1 de 160). Monique Junkes é sócia.",
         "parceiro": "Modelo de franquia = múltiplos pontos de indicação. Lucas é CEO com LinkedIn ativo. Escalável.",
         "nome_curto": "Lucas", "gancho": "já apoiou mais de 350 empresas com R$1 bilhão em licitações e foi acelerada pelo InovAtiva Brasil"},
        {"empresa": "E3 Licitações", "tipo": "Consultoria (obras)", "local": "Porto Alegre/RS", "potencial": "Alto",
         "quem": "Fabricio Frizzo Pagnossin (sócio-admin, advogado)", "wpp": "(51) 3307-8049", "email": "contato@e3licitacoes.com.br",
         "site": "e3licitacoes.com.br", "redes": "FB: @E3.licitacoes / LI empresa: e3-licitacoes",
         "sobre": "Fundada 2011. Fabricio + Diego Magoga Conde (sócio). Especialista em obras, reequilíbrio econômico e defesas. Equipe ex-setor público.",
         "parceiro": "Foco em obras = editais mais complexos, onde esse mapeamento gera mais valor. Construtoras precisam decidir rápido.",
         "nome_curto": "Fabricio", "gancho": "é referência em licitações de obras há mais de 15 anos. Construtoras que precisam decidir rápido em quais editais entrar são o público ideal pra esse tipo de análise"},
        {"empresa": "Inovação Assessoria", "tipo": "Franquia nacional", "local": "Londrina/PR + 22 unidades", "potencial": "Alto",
         "quem": "Marco Antonio Cito (fundador/CEO, 26 anos exp.)", "wpp": "(43) 98858-3264", "email": "marco@inovesempre.com.br",
         "site": "inovesempre.com.br", "redes": "IG: @inovacaolicitacoes (1.704 seg) / LI: marco-antonio-cito-55b95859 / LI empresa: inovacaobrasil",
         "sobre": "1a franquia de assessoria em licitações do BR. 22 unidades em 10 estados. Filtra 5K+ editais/dia, 1K pregões/mes. Investimento R$39K. Marco é contador + MBA Gestão Pública.",
         "parceiro": "Escala via franquias. Cada unidade pode indicar. Marco é decisor e membro BNI Legendários.",
         "nome_curto": "Marco", "gancho": "criou a primeira franquia de assessoria em licitações do Brasil com 22 unidades. Cada franqueado poderia oferecer esse serviço pros clientes"},
        {"empresa": "Pró Licitante", "tipo": "Lawtech", "local": "Lages/SC", "potencial": "Alto",
         "quem": "Tiago Griebeler Sandi (admin/fundador, consultor jurídico)", "wpp": "(49) 99144-2670", "email": "contato@prolicitante.com.br",
         "site": "prolicitante.com.br", "redes": "IG: @prolicitante / LI: tiago-sandi-770531166 / LI empresa: prolicitante",
         "sobre": "Lawtech desde 2014. 339K+ casos. R$4,5B+ homologados. 11-50 func. Tiago também é Gov Director na Multilaser. Carlos e Marcos são funcionários, NÃO decisores.",
         "parceiro": "Mentalidade tech. Volume enorme. Eles automatizam execução, nós geramos inteligência. Complementar.",
         "nome_curto": "Tiago", "gancho": "automatiza processos licitatórios e já processou mais de 339 mil casos. Vocês cuidam da execução, eu gero a inteligência de quais editais valem a pena"},
        {"empresa": "Eagle Consultoria", "tipo": "Consultoria full-service", "local": "Sorocaba/SP e São Paulo/SP", "potencial": "Alto",
         "quem": "Pedro Gali e Rafael Montoro (ambos administradores)", "wpp": "(11) 97386-8755", "email": "lgpd@eaglelicitacoes.com.br",
         "site": "consultoriaemlicitacao.com.br", "redes": "IG: @eaglelicitacoes / LI empresa: ecprojetos",
         "sobre": "Fundada 2008. CNPJ 09.543.462. Dois escritórios: Sorocaba (Ed. Illimite) + SP Vila Mariana. Prospecção, documentação, representação.",
         "parceiro": "Base consolidada há 18 anos. Mapeamento complementa sem competir. Sem email comercial público (usar WhatsApp).",
         "nome_curto": "Pedro", "gancho": "é referência em consultoria de licitações desde 2008. Com dois escritórios em SP, seus clientes devem precisar decidir rápido em quais editais entrar"},
        {"empresa": "Licijur Consultoria", "tipo": "Consultoria jurídica (infra/obras)", "local": "Porto Alegre/RS", "potencial": "Alto",
         "quem": "Marcelo Gazen (diretor, advogado) / Mauricio Gazen (admin)", "wpp": "(51) 99527-2261", "email": "contato@licijur.com.br",
         "site": "licijur.com.br", "redes": "LI: marcelo-gazen-8bb0909a / FB: Licijur",
         "sobre": "Fundada 2002. 24 anos de mercado. Equipe 30+ anos exp. Diagnóstico, captação, jurídico. Priscila Jardim (sócia). Entrevista Programa Poder RS.",
         "parceiro": "Foco em infra/obras = editais mais densos. Clientes de ticket alto. Marcelo tem LinkedIn ativo.",
         "nome_curto": "Marcelo", "gancho": "é referência em licitações de infraestrutura há mais de 24 anos. Clientes de obras grandes são os que mais precisam desse tipo de mapeamento"},
        {"empresa": "Jus Licitações", "tipo": "Consultoria jurídica", "local": "Curitiba/PR", "potencial": "Médio-Alto",
         "quem": "Anilda Patricia de Sousa Santos e Glaucia Rosane de Souza (sócias-admin)", "wpp": "(41) 99758-5425", "email": "contato@juslicitacoes.com.br",
         "site": "juslicitacoes.com.br", "redes": "IG: @juslicitacoes / FB: juslicitacoes / LI: jus-licitacoes",
         "sobre": "Fundada 2019, 17+ anos exp. pessoal. 1.990+ clientes, 2.500+ processos, 740+ projetos. Ed. New York Building, 15o andar, Centro Curitiba.",
         "parceiro": "Dizem ser 'a maior do Sul'. Treinamentos in company = acesso direto a empresas B2G.",
         "nome_curto": "pessoal da Jus", "gancho": "atende quase 2 mil clientes no Paraná e diz ser a maior consultoria do Sul"},
        {"empresa": "Consultoria Licitações (Brasília)", "tipo": "Consultoria estratégica", "local": "Brasília/DF", "potencial": "Médio-Alto",
         "quem": "Responsável comercial (presença digital limitada)", "wpp": "(61) 99869-6144", "email": "comercial@consultorialicitacoes.com.br",
         "site": "consultorialicitacoes.com.br", "redes": "Presença digital limitada",
         "sobre": "Base em Brasília. Foco na Lei 14.133. Também atende Salvador, Manaus e Aracaju. Sem equipe identificada publicamente.",
         "parceiro": "Brasília = epicentro licitações federais. Clientes de alto valor. Contato apenas por telefone.",
         "nome_curto": "pessoal", "gancho": "está em Brasília atuando de perto com licitações federais e Lei 14.133"},
        {"empresa": "LicitaMinas", "tipo": "Consultoria regional", "local": "Belo Horizonte/MG", "potencial": "Médio-Alto",
         "quem": "Responsável comercial (proprietário não identificado)", "wpp": "(31) 8802-5445", "email": "contato@licitaminas.com.br",
         "site": "licitaminas.com.br", "redes": "Blog: licitaminas.blogspot.com (desde 2007)",
         "sobre": "Foco em MG. Assessoria documental, jurídica, representação. Rede de colaboradores em MG e outros estados. Endereço: R. Lindolfo de Azevedo 1883, BH.",
         "parceiro": "MG = um dos maiores mercados B2G. Conhecimento profundo do mercado local. Blog ativo desde 2007.",
         "nome_curto": "pessoal da LicitaMinas", "gancho": "atende empresas em Minas com assessoria documental e jurídica em licitações"},
        {"empresa": "Liciticon", "tipo": "Consultoria regional", "local": "Campo Grande/RJ + Angra + Maricá", "potencial": "Médio",
         "quem": "Luiz Gabriel Reis dos Santos Silva (proprietário/admin)", "wpp": "(21) 97514-3001", "email": "contato@liciticon.com.br",
         "site": "liciticon.com", "redes": "IG: @liciticon_assessoria / FB: /liciticon",
         "sobre": "Fundada 2019. CNPJ 35.400.780. 3 unidades no RJ. Capital R$30K. 2o WhatsApp: (21) 96741-0183.",
         "parceiro": "Cobertura regional RJ. Luiz Gabriel é decisor identificado. Clientes precisam priorizar editais.",
         "nome_curto": "Luiz Gabriel", "gancho": "atende empresas no RJ com 3 unidades. Seus clientes precisam decidir rápido em quais editais entrar"},
        {"empresa": "LICIT Consultoria", "tipo": "Consultoria + Cursos", "local": "Recife/PE", "potencial": "Médio",
         "quem": "Lizziane Ferreira Fernandes (proprietária)", "wpp": "(81) 98883-0037", "email": "contato@licit.net.br",
         "site": "licit.net.br", "redes": "FB: Licit Consultoria e Assessoria",
         "sobre": "CNPJ 23.549.856. Experiência como órgão contratante E licitante (visão dos dois lados). Cursos, palestras, workshops.",
         "parceiro": "Nordeste sub-atendido. Lizziane tem visão de ambos os lados (público + privado).",
         "nome_curto": "Lizziane", "gancho": "tem experiência dos dois lados da licitação, tanto como órgão quanto como licitante. Essa visão dupla é rara"},
        {"empresa": "Êxito Licitações", "tipo": "Consultoria", "local": "Porto Alegre/RS", "potencial": "Médio",
         "quem": "Marcelo Castiglia (sócio-admin, consultor comercial)", "wpp": "(51) 98420-2429", "email": "comercial@exitolicitacoes.com.br",
         "site": "exitolicitacoes.com.br", "redes": "IG: @exitolicitacoes (2.065 seg) / LI empresa: exitolicitacoes",
         "sobre": "Fundada 2018. 3 sócios: Marcelo Castiglia, Jamille Medeiros, Simone Hengist. Recursos, impugnações, compliance.",
         "parceiro": "Complementar: eles fazem recursos e impugnações, nós fazemos a inteligência prévia.",
         "nome_curto": "Marcelo", "gancho": "é especialista em análise de editais e recursos administrativos. Nós fazemos a inteligência prévia, vocês fazem a defesa"},
        {"empresa": "Licitar Engenharia", "tipo": "Consultoria de engenharia", "local": "Salvador/BA", "potencial": "Médio-Alto",
         "quem": "Responsável comercial (proprietário não identificado)", "wpp": "(71) 99914-9221", "email": "formulário no site",
         "site": "licitarengenharia.com.br", "redes": "IG: @licitarengenharia / LI: licitar-orcamentos-e-engenharia",
         "sobre": "12+ anos. Foco em planilhas orçamentárias e engenharia de custos para construção civil. Contato apenas via WhatsApp/formulário.",
         "parceiro": "Engenharia = editais mais complexos. Construtoras são público ideal pra esse mapeamento.",
         "nome_curto": "pessoal da Licitar", "gancho": "atua com orçamentos e engenharia de custos para licitações de obras"},
        {"empresa": "Otimiza Licitações", "tipo": "Franquia nacional (ABF)", "local": "Guarulhos/SP (sede) + 40 unidades", "potencial": "Médio-Alto",
         "quem": "Geovana Nina (sócia Floripa) / Bergson Costa (matriz)", "wpp": "widget no site", "email": "bergson.costa@otimizalicitacoes.com",
         "site": "licitacoes.app / otimizalicitacoes.com", "redes": "IG: @otimiza_licitacoes / LI empresa: otimiza-licitacoes",
         "sobre": "Maior franquia ABF de licitações. 40+ unidades. ~20 anos. Investimento a partir R$20K, home office. Sede real: Guarulhos/SP, não Floripa.",
         "parceiro": "40 franqueados = 40 canais de indicação. Selo ABF traz credibilidade. Escalável.",
         "nome_curto": "Bergson", "gancho": "construiu a maior franquia de licitações do Brasil com mais de 40 unidades. Cada franqueado poderia revender esse serviço"},
        {"empresa": "LicitaBR", "tipo": "Consultoria + Cursos", "local": "Guarulhos/SP", "potencial": "Médio",
         "quem": "Thiago Rocha Benedito (CEO/fundador)", "wpp": "(11) 4386-1386", "email": "contato@licitabr.com.br",
         "site": "licitabr.com", "redes": "IG: @licitabr / LI: thirocha / FB: /licitabr",
         "sobre": "Fundada 2015. 20K+ licitações participadas, R$8B+ em negócios, 500+ empresas. Thiago também fundou BR4 Licitações.",
         "parceiro": "500+ empresas atendidas = base grande. Modelo de depto externo traz clientes recorrentes.",
         "nome_curto": "Thiago", "gancho": "já participou de mais de 20 mil licitações e atende 500+ empresas. Esse mapeamento complementa o trabalho de vocês"},
        {"empresa": "Licit Mais Brasil", "tipo": "Plataforma de captação", "local": "Goiânia/GO", "potencial": "Médio",
         "quem": "Guilherme Pereira Gomes (diretor) / Aurora dos Anjos (presidente)", "wpp": "(62) 9 9842-2521", "email": "faleconosco@licitmaisbrasil.com.br",
         "site": "licitmaisbrasil.com.br", "redes": "IG: @licitmais / LI: licitmaisbrasil / FB: licitmaisbrasil",
         "sobre": "Fundada 2013 em Goiânia. S.A. Capital R$190K. Captação diária de editais. Curso 'Método Reshape'. Fixo: (62) 3989-9903.",
         "parceiro": "Eles captam, nós analisamos em profundidade. Complementar. Clientes já pagam por informação.",
         "nome_curto": "Guilherme", "gancho": "já distribui editais segmentados pra uma base grande de clientes. Eu adiciono a análise profunda que falta"},
        {"empresa": "ConLicitação", "tipo": "Plataforma SaaS + Assessoria", "local": "São Paulo/SP", "potencial": "Médio",
         "quem": "Sonia Moura (fundadora)", "wpp": "(11) 3783-8666", "email": "contato@conlicitacao.com.br",
         "site": "conlicitacao.com.br", "redes": "LI: conlicitacao / FB: @conlicitacao",
         "sobre": "Fundada 1999. 27 anos. 20K+ clientes. ~100 func. App próprio. Sede: Estr. do Jaguaré 422, Butantã, SP. Também tem consultoria jurídica.",
         "parceiro": "Base enorme mas modelo SaaS. Mapeamento avulso = complemento premium. Cuidado: concorrente adjacente.",
         "nome_curto": "Sonia", "gancho": "construiu a maior base do segmento com mais de 20 mil empresas. Esse mapeamento funciona como complemento premium pro que vocês já entregam"},
        # ===== NOVAS (pesquisa intel-b2g) =====
        {"empresa": "MEP Licitações", "tipo": "Assessoria regional", "local": "Cuiabá/MT", "potencial": "Alto",
         "quem": "Responsável comercial", "wpp": "widget no site", "email": "contato via site",
         "site": "meplicitacoes.com.br", "redes": "LI: mep-licitacoes",
         "sobre": "Maior assessoria de licitações do MT. Av. Miguel Sutil 8388, 10o andar. Filtra editais e envia só os relevantes. Atendimento nacional.",
         "parceiro": "Dominante no MT, sem concorrente mapeado. Centro-Oeste é gap. Modelo de filtragem = complementar.",
         "nome_curto": "pessoal da MEP", "gancho": "é a maior assessoria de licitações do Mato Grosso. Eu complemento o filtro de vocês com análise profunda do PDF"},
        {"empresa": "LBR Licitar", "tipo": "Assessoria + Consultoria", "local": "Fortaleza/CE", "potencial": "Alto",
         "quem": "Responsável comercial", "wpp": "verificar site", "email": "verificar site",
         "site": "licitarbr.com.br", "redes": "IG: @licitarbr (8.554 seg)",
         "sobre": "Fundada 2021. Analistas, advogados e contadores. Clientes faturaram R$300M+ em contratos públicos. Scopa Platinum Corporate, Fortaleza.",
         "parceiro": "8.5K seguidores = audiência forte no NE. R$300M em contratos = credibilidade. Empresa jovem e digital.",
         "nome_curto": "pessoal da LBR", "gancho": "já ajudou clientes a faturar mais de R$300 milhões em contratos públicos. Esse mapeamento fortalece o resultado"},
        {"empresa": "EXPER Consultoria", "tipo": "Assessoria", "local": "Guarapari/ES", "potencial": "Médio-Alto",
         "quem": "Robson Amorim Mendes (proprietário)", "wpp": "(27) 3361-1338", "email": "exper@experconsultores.com.br",
         "site": "experconsultores.com.br", "redes": "IG: @experconsultoria (1.753 seg)",
         "sobre": "5 anos. R$21M+ em contratos públicos em 2025. Transforma empresas em fornecedores gov. ES sem concorrente mapeado.",
         "parceiro": "ES é gap total na lista. Robson é decisor identificado. Instagram ativo com resultados mensuráveis.",
         "nome_curto": "Robson", "gancho": "gerou R$21 milhões em contratos públicos só em 2025. Eu complemento isso com inteligência de quais editais valem a pena"},
        {"empresa": "LicitaMundo", "tipo": "Assessoria jurídica", "local": "Betim/MG", "potencial": "Médio-Alto",
         "quem": "Eduardo Araujo (CEO, advogado desde 2003, PUC Minas)", "wpp": "verificar site", "email": "verificar site",
         "site": "verificar", "redes": "LI: licitamundo (1.932 seg)",
         "sobre": "Fundada 2019. CNPJ 34.049.529. Ciclo completo: seleção de edital até entrega do serviço. Background jurídico forte.",
         "parceiro": "MG complementa LicitaMinas. Eduardo tem LinkedIn ativo. Background jurídico = clientes maiores.",
         "nome_curto": "Eduardo", "gancho": "atua no ciclo completo de licitações com background jurídico desde 2003. Esse mapeamento agiliza a seleção de editais"},
        {"empresa": "LCT Assessoria", "tipo": "Assessoria jurídica premium", "local": "Brasília/DF", "potencial": "Médio-Alto",
         "quem": "Responsável comercial", "wpp": "verificar site", "email": "verificar site",
         "site": "lctassessoria.com.br", "redes": "LI: lct-assessoria / FB: lct.assessoria",
         "sobre": "Representação em Tribunais de Contas e MP. Programas de integridade. Reequilíbrio econômico-financeiro. Posicionamento premium.",
         "parceiro": "Brasília + Tribunais de Contas = clientes de alto ticket. Complementa a consultoria de Brasília já mapeada.",
         "nome_curto": "pessoal da LCT", "gancho": "atua com representação em Tribunais de Contas e programas de integridade. Clientes desse perfil são os que mais precisam de mapeamento de editais"},
        {"empresa": "Licitatio Consultores", "tipo": "Consultoria", "local": "Curitiba/PR", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "(41) 99997-6360", "email": "verificar site",
         "site": "licitatio.com.br", "redes": "LI: licitatio-consultores-associados / FB: licitatio",
         "sobre": "Insere empresas no setor público a custo fixo baixo. R. Emiliano Perneta 390, 11o andar, Curitiba.",
         "parceiro": "Curitiba complementa Jus Licitações. Modelo custo fixo baixo = clientes PME que precisam de mapeamento.",
         "nome_curto": "pessoal da Licitatio", "gancho": "insere empresas no setor público a custo fixo baixo. Esse mapeamento ajuda a priorizar os editais certos"},
        {"empresa": "i9 Licitações", "tipo": "Consultoria full-service", "local": "São Paulo/SP", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "(11) 3297-3319", "email": "comercial@i9licitacoes.com.br",
         "site": "i9licitacoes.com.br", "redes": "verificar",
         "sobre": "Fundada 2009. 15+ anos. Ciclo completo: busca, cadastro, representação em pregão, departamento jurídico. Consultoria online via Skype/email.",
         "parceiro": "Madura (desde 2009). Consultoria online = pode integrar esse serviço no fluxo remoto.",
         "nome_curto": "pessoal da i9", "gancho": "atua há mais de 15 anos com ciclo completo de licitações. Esse mapeamento complementa a busca de editais"},
        {"empresa": "Grupo Licita", "tipo": "Consultoria (gestão pública)", "local": "Fortaleza/CE", "potencial": "Médio-Alto",
         "quem": "Responsável comercial", "wpp": "(85) 99618-2727", "email": "contato@grupolicita.com.br",
         "site": "grupolicita.com.br", "redes": "IG: @grupolicita (5.848 seg)",
         "sobre": "15+ anos. Atende órgãos públicos E empresas. Auditoria interna (TCM/TCU/CGU). Presença semanal presencial.",
         "parceiro": "Visão dos dois lados (público + privado). 5.8K seguidores = audiência forte no CE.",
         "nome_curto": "pessoal do Grupo Licita", "gancho": "atende tanto órgãos públicos quanto empresas há mais de 15 anos no Ceará. Essa visão dos dois lados é rara"},
        {"empresa": "Solicita Consultoria", "tipo": "Consultoria (órgãos públicos)", "local": "São Luís/MA", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "(98) 3303-8252", "email": "verificar site",
         "site": "solicitaconsultoria.com.br", "redes": "verificar",
         "sobre": "10+ anos. Advogados, administradores, engenheiros e técnicos. Guia órgãos públicos em procedimentos licitatórios. Única no Maranhão.",
         "parceiro": "MA = gap total. Trabalha com órgãos = pode indicar fornecedores que precisam de mapeamento de editais.",
         "nome_curto": "pessoal da Solicita", "gancho": "orienta órgãos públicos em licitações no Maranhão. Vocês conhecem os dois lados e podem indicar esse mapeamento pros fornecedores"},
        {"empresa": "Tech Licitações", "tipo": "Assessoria (pregão eletrônico)", "local": "Recife/PE", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "(81) 98998-0487", "email": "contato@techlicitacoes.com",
         "site": "techlicitacoes.com", "redes": "verificar",
         "sobre": "Equipe técnica para pregão eletrônico. ETP, Termo de Referência, Mapa de Riscos, editais. Atende órgãos e empresas.",
         "parceiro": "PE complementa LICIT. Foco técnico (ETP, TR) = clientes que precisam de inteligência prévia.",
         "nome_curto": "pessoal da Tech", "gancho": "atua com a parte técnica de pregões eletrônicos em Pernambuco. Eu complemento com a inteligência de quais editais valem a pena"},
        {"empresa": "Lince Licitações", "tipo": "Consultoria jurídica", "local": "Belém/PA", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "verificar site", "email": "verificar site",
         "site": "lincelicitacoes.com.br", "redes": "LI: lincelct",
         "sobre": "Advogados, engenheiros e contadores. Foco em obras públicas no Pará. Única consultoria mapeada no Norte.",
         "parceiro": "PA = gateway pro Norte inteiro. Foco em obras = editais complexos onde esse mapeamento gera mais valor.",
         "nome_curto": "pessoal da Lince", "gancho": "atua com licitações de obras no Pará. Editais complexos da região Norte são onde esse mapeamento gera mais valor"},
        {"empresa": "3S Licitações", "tipo": "Assessoria + Consultoria", "local": "Cuiabá/MT", "potencial": "Médio",
         "quem": "Paulo Gomes (Consultor Master, 10+ anos servidor público)", "wpp": "verificar site", "email": "verificar site",
         "site": "3slicitacoes.com.br", "redes": "verificar",
         "sobre": "Solução completa da participação à execução do contrato. Paulo Gomes tem 10+ anos como agente público.",
         "parceiro": "MT = mercado ativo (3 consultorias mapeadas). Paulo é ex-servidor = credibilidade.",
         "nome_curto": "Paulo", "gancho": "tem mais de 10 anos como agente público e agora assessora empresas em licitações. Essa visão de dentro do governo é um diferencial"},
        {"empresa": "Inovasus", "tipo": "Consultoria (saúde pública)", "local": "Frutal/MG", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "(34) 9 9896-5641", "email": "contato@inovasus.com.br",
         "site": "inovasus.com.br", "redes": "verificar",
         "sobre": "Fundada 2021. Gestão de saúde pública municipal. Planejamento, sistemas de informação, treinamento. Atende Secretarias de Saúde.",
         "parceiro": "Nicho saúde pública. Clientes (prefeituras) são os COMPRADORES em licitações de saúde. Referência reversa.",
         "nome_curto": "pessoal da Inovasus", "gancho": "atende Secretarias de Saúde municipais. Os fornecedores de saúde que participam das licitações dessas prefeituras são o público ideal pra esse tipo de análise"},
        {"empresa": "MedAssist", "tipo": "Contabilidade + Assessoria médica", "local": "São Paulo/SP", "potencial": "Médio-Alto",
         "quem": "Responsável comercial", "wpp": "(11) 3802-3004", "email": "contato@medassistservicos.com.br",
         "site": "medassistservicos.com.br", "redes": "verificar",
         "sobre": "6.000+ médicos atendidos. Abertura de empresa, otimização fiscal e assessoria em licitações médicas (PS, UTI, anestesia, obstetrícia).",
         "parceiro": "6K médicos = base enorme de potenciais compradores desse mapeamento. Nicho saúde com ticket alto.",
         "nome_curto": "pessoal da MedAssist", "gancho": "atende mais de 6 mil médicos e já assessora em licitações de saúde. Eu mostro quais editais cada médico pode disputar"},
        {"empresa": "Magna Licitações", "tipo": "Assessoria + Cursos", "local": "Hortolândia/SP", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "(19) 2518-0157", "email": "verificar site",
         "site": "magnalicitacoes.com.br", "redes": "IG: @magnalicitacoes / LI: magna-licitacoes / Telegram: magnata",
         "sobre": "Fundada 2016 por ex-pregoeiros do Exército (8 anos servidores). R$200M+ em contratos. Treinamento online e presencial.",
         "parceiro": "Background militar = credibilidade. Interior SP = mercado sub-atendido. Cursos = base de alunos.",
         "nome_curto": "pessoal da Magna", "gancho": "foi fundada por ex-pregoeiros do Exército e já apoiou R$200 milhões em contratos. Esse mapeamento é um serviço prático pros alunos"},
        {"empresa": "Solicita Licitações", "tipo": "Consultoria regional", "local": "Cuiabá/MT", "potencial": "Médio",
         "quem": "Responsável comercial", "wpp": "verificar site", "email": "verificar site",
         "site": "solicitalicitacoes.com.br", "redes": "verificar",
         "sobre": "16+ anos. R$250M+ em contratos. Atendimento nacional a partir de Cuiabá.",
         "parceiro": "R$250M em contratos = base consolidada. MT ativo. Complementa MEP e 3S.",
         "nome_curto": "pessoal da Solicita", "gancho": "tem mais de 16 anos e R$250 milhões em contratos. Esse mapeamento ajuda a priorizar os melhores editais"},
    ]

    for row_idx, c in enumerate(consultorias, 2):
        msg_wpp = gerar_msg_wpp(
            c["nome_curto"], c["empresa"], c["gancho"], c["potencial"],
        )
        msg_email = gerar_msg_email(
            c["nome_curto"], c["empresa"], c["gancho"],
        )
        msg_followup = gerar_msg_followup(c["nome_curto"])

        row_data = [
            c["empresa"], c["tipo"], c["local"], c["potencial"],
            c["quem"], c["wpp"], c["email"], c["site"],
            c["redes"], c["sobre"], c["parceiro"],
            msg_wpp, msg_email, msg_followup, "Pendente", "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            cell.border = borda
            cell.font = Font(name="Calibri", size=10)

            if col_idx == 4:
                if value == "Alto":
                    cell.fill = alto_fill
                elif value in ("Médio-Alto", "Médio"):
                    cell.fill = medio_fill

    col_widths = [22, 18, 28, 10, 28, 20, 30, 24, 28, 35, 35, 55, 70, 40, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{len(consultorias) + 1}"

    # --- Sheet 2: Resumo ---
    ws2 = wb.create_sheet("Resumo")

    titulo_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    secao_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    normal_font = Font(name="Calibri", size=11)

    linhas = [
        ("t", "Consultorias mapeadas para parceria", ""),
        ("", "", ""),
        ("n", "Data:", "11/03/2026"),
        ("n", "Total mapeado:", str(len(consultorias))),
        ("n", "Potencial Alto:", str(sum(1 for c in consultorias if c["potencial"] == "Alto"))),
        ("n", "Médio-Alto:", str(sum(1 for c in consultorias if c["potencial"] == "Médio-Alto"))),
        ("n", "Médio:", str(sum(1 for c in consultorias if c["potencial"] == "Médio"))),
        ("", "", ""),
        ("s", "Cobertura geográfica", ""),
        ("n", "Sul (RS, SC, PR):", "E3, Licitando, Êxito, Concreta, Pró Licitante, Jus, Inovação, Licijur, Licitatio"),
        ("n", "Sudeste (SP, RJ, MG, ES):", "Eagle, Brasil Licitar, LicitaMinas, Liciticon, Route, EXPER, LicitaMundo, i9, Magna, Inovasus, MedAssist"),
        ("n", "Nordeste (PE, BA, CE, MA):", "LICIT, Licitar Engenharia, LBR Licitar, Grupo Licita, Solicita, Tech Licitações"),
        ("n", "Centro-Oeste (DF, MT):", "Consultoria Licitações Brasília, LCT, MEP, 3S, Solicita Licitações"),
        ("n", "Norte (PA):", "Lince Licitações"),
        ("n", "Nacional:", "ConLicitação, LicitaBR, Licit Mais, Otimiza, Inovação, Licijur"),
        ("", "", ""),
        ("s", "Prioridade de abordagem (Top 10)", ""),
        ("n", "1.", "Marcela Moneta (Licitando). Decisora, WhatsApp direto, 20+ anos"),
        ("n", "2.", "Ricardo Ribas (Route). Educador, YouTube, alunos são público ideal"),
        ("n", "3.", "Gustavo Oliveira (Brasil Licitar). Diretor, MBA FIA, 500+ licitações"),
        ("n", "4.", "Lucas de Medeiros (Concreta). CEO, franquia 110+ clientes, escalável"),
        ("n", "5.", "Fabricio Pagnossin (E3). Sócio-admin, especialista obras, 15+ anos"),
        ("n", "6.", "Marco Antonio Cito (Inovação). Fundador/CEO, 22 unidades franqueadas"),
        ("n", "7.", "Tiago Griebeler Sandi (Pró Licitante). Admin real, lawtech, 339K+ casos"),
        ("n", "8.", "Marcelo Gazen (Licijur). Diretor, atuação nacional, WhatsApp direto"),
        ("n", "9.", "Eagle Assessoria. Engenharia civil SP, volume alto, potencial escala"),
        ("n", "10.", "EXPER (Guarapari/ES). ES sem concorrência, nicho obras/engenharia"),
        ("", "", ""),
        ("s", "Estrategia de mensagens (Copymasters v2)", ""),
        ("n", "Framework:", "PAS (Problem-Agitation-Solution). Destinatário já conhece o problema"),
        ("n", "WhatsApp Alto:", "Versão completa ~79 palavras. Abre com o destinatário, 3 deliverables"),
        ("n", "WhatsApp Médio:", "Versão curta ~52 palavras. Direto ao ponto, menos contexto pessoal"),
        ("n", "Email:", "Assunto personalizado com nome + 'receita extra'. Corpo ~90 palavras"),
        ("n", "Follow-up D+3:", "WhatsApp único, 3 linhas, sem pressão. Máximo 1 follow-up."),
        ("n", "Regras:", "Sem áudio/anexo no 1o contato. Sem 'plano/assinatura/tier'. Enviar Ter-Qui 8-10h."),
        ("", "", ""),
        ("s", "O que destacar na abordagem", ""),
        ("n", "1.", "Abre com o DESTINATÁRIO, não com o remetente (2s pra decidir se lê)"),
        ("n", "2.", "Leitura do PDF real do edital (não só título ou resumo)"),
        ("n", "3.", "3 deliverables concretos: editais disputáveis, concorrentes, gaps de habilitação"),
        ("n", "4.", "Custo R$97, margem livre pro parceiro (modelo white-label)"),
        ("n", "5.", "Posicionar como trabalho pessoal, não ferramenta/software"),
        ("n", "6.", "CTA: 'Posso te mandar um exemplo pra avaliar?'"),
        ("", "", ""),
        ("s", "Modelo de parceria", ""),
        ("n", "Custo:", "R$97 por análise (preço fixo pro parceiro)"),
        ("n", "Revenda:", "Parceiro cobra o preço que quiser do cliente final"),
        ("n", "Entrega:", "Tiago faz, parceiro entrega ao cliente como serviço próprio"),
        ("n", "Custo pro parceiro:", "Zero upfront. Paga R$97 só quando vender"),
        ("", "", ""),
        ("s", "Cadência de abordagem", ""),
        ("n", "D+0:", "WhatsApp (Ter-Qui, 8-10h)"),
        ("n", "D+2:", "Email (se sem resposta no WhatsApp)"),
        ("n", "D+5:", "Follow-up WhatsApp (único, 3 linhas)"),
        ("n", "Depois:", "Não insistir. Tom colegial > conversão forçada."),
        ("", "", ""),
        ("s", "Próximos passos", ""),
        ("n", "1.", "Criar exemplo anonimizado (construção civil). PREREQUISITO"),
        ("n", "2.", "WhatsApp para Top 5 (Marcela, Ricardo, Gustavo, Lucas, Fabricio). Ter-Qui 8-10h"),
        ("n", "3.", "Criar cupons Stripe por parceiro (LICITANDO30, ROUTE30, etc)"),
        ("n", "4.", "Email para os 10 prioritários. D+2"),
        ("n", "5.", "Follow-up WhatsApp. D+5, apenas quem não respondeu"),
        ("n", "6.", "Segunda onda: WhatsApp para leads 11-20. Semana seguinte"),
        ("n", "7.", "Terceira onda: leads 21-38 (novos, menos validados). Após primeiros resultados"),
    ]

    for row_idx, (tipo, col1, col2) in enumerate(linhas, 1):
        c1 = ws2.cell(row=row_idx, column=1, value=col1)
        c2 = ws2.cell(row=row_idx, column=2, value=col2)
        c1.alignment = wrap
        c2.alignment = wrap

        if tipo == "t":
            c1.font = titulo_font
        elif tipo == "s":
            c1.font = secao_font
        else:
            c1.font = normal_font
            c2.font = normal_font

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 65

    # Save
    output = os.path.normpath(os.path.join(
        os.path.dirname(__file__), "..", "docs", "reports",
        "consultorias-parceria-2026-03-11.xlsx",
    ))
    wb.save(output)
    print(f"Planilha gerada: {output}")
    print(f"  {len(consultorias)} consultorias, 2 abas")
    return output


if __name__ == "__main__":
    create_workbook()
