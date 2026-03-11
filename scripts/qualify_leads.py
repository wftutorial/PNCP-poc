#!/usr/bin/env python3
"""
/qualify-b2g — Qualificação Inteligente de Leads B2G

Lê diretamente das planilhas originais do /intel-b2g, aplica scoring 7D,
seleciona top 20% (Pareto), e consolida em uma única planilha preservando
a estrutura original (5 abas) + colunas de scoring adicionais.

Uso:
    python scripts/qualify_leads.py
    python scripts/qualify_leads.py --top-pct 30   # top 30% ao invés de 20%
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: pip install openpyxl")
    sys.exit(1)


# ── Config ──────────────────────────────────────────────────────────────────

INPUT_FILES = [
    ("docs/intel-b2g/leads-engenharia-sul-2026-03-06.xlsx", "Sul"),
    ("docs/intel-b2g/leads-engenharia-sudeste-2026-03-09.xlsx", "Sudeste"),
    ("docs/intel-b2g/leads-engenharia-nordeste-2026-03-10.xlsx", "Nordeste"),
]

OUTPUT_FILE = "docs/intel-b2g/qualified-engenharia-consolidado-2026-03-10.xlsx"
MARKDOWN_FILE = "docs/intel-b2g/qualified-engenharia-consolidado-2026-03-10.md"

# Original column indices (0-based) in Leads tab
COL = {
    "num": 0,
    "relevancia": 1,
    "empresa": 2,
    "nome_fantasia": 3,
    "cnpj": 4,
    "cidade_sede": 5,
    "uf_sede": 6,
    "cnae_principal": 7,
    "porte": 8,
    "capital_social": 9,
    "ufs_atuacao": 10,
    "municipios": 11,
    "fat_gov_mensal": 12,
    "contratos_6m": 13,
    "valor_total": 14,
    "orgaos": 15,
    "decisor": 16,
    "cargo_decisor": 17,
    "telefone1": 18,
    "whatsapp_flag": 19,
    "telefone2": 20,
    "email": 21,
    "sancoes": 22,
    "objetos": 23,
    "status_contato": 24,
    "whatsapp_numero": 36,
    "link_wame": 37,
    "mensagem_whatsapp": 38,
}

# Scoring thresholds
TIER_THRESHOLDS = {1: 65, 2: 48, 3: 30}  # >= score

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TIER1_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
TIER2_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
TIER3_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TIER_FILLS = {1: TIER1_FILL, 2: TIER2_FILL, 3: TIER3_FILL}
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── Scoring Functions ───────────────────────────────────────────────────────

def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_str(val, default=""):
    if val is None:
        return default
    return str(val).strip()


def score_d1_volume(fat_gov_mensal: float) -> float:
    """D1: Volume Governamental (peso 25%)"""
    if fat_gov_mensal > 500_000:
        return 100
    elif fat_gov_mensal > 200_000:
        return 85
    elif fat_gov_mensal > 100_000:
        return 70
    elif fat_gov_mensal > 50_000:
        return 55
    elif fat_gov_mensal > 20_000:
        return 40
    elif fat_gov_mensal > 5_000:
        return 25
    else:
        return 10


def score_d2_frequencia(contratos_6m: int) -> float:
    """D2: Frequência de Participação (peso 20%) — usando contratos em 6m, extrapola para 12m"""
    contratos_ano = contratos_6m * 2  # extrapolação
    if contratos_ano >= 20:
        return 100
    elif contratos_ano >= 10:
        return 85
    elif contratos_ano >= 5:
        return 70
    elif contratos_ano >= 3:
        return 55
    elif contratos_ano >= 1:
        return 40
    else:
        return 15


def score_d3_porte(capital_social: float, porte: str, fat_gov_mensal: float) -> float:
    """D3: Porte e Capacidade (peso 15%)"""
    # Se capital_social indisponível, usar fat_gov_mensal como proxy
    if capital_social <= 1:
        if fat_gov_mensal > 1_000_000:
            return 90
        elif fat_gov_mensal > 500_000:
            return 75
        elif fat_gov_mensal > 100_000:
            return 60
        elif fat_gov_mensal > 20_000:
            return 40
        else:
            return 25

    porte_upper = porte.upper() if porte else ""
    if capital_social > 1_000_000 or "GRANDE" in porte_upper or "MEDIO" in porte_upper:
        return 100
    elif capital_social > 500_000:
        return 80
    elif capital_social > 100_000:
        return 60
    else:
        return 30


def score_d4_acessibilidade(decisor: str, telefone1: str, whatsapp_flag: str,
                             telefone2: str, email: str) -> float:
    """D4: Acessibilidade do Decisor (peso 15%)"""
    has_decisor = decisor and decisor not in ("N/D", "None", "", "—")
    has_whatsapp = whatsapp_flag and str(whatsapp_flag).lower() in ("sim", "yes", "true", "s")
    has_celular = False
    for tel in [telefone1, telefone2]:
        if tel and ("9" in str(tel)[5:7] or "9" in str(tel)[4:6]):
            has_celular = True
            break
    has_email = email and email not in ("N/D", "None", "", "—") and "@" in str(email)
    has_fone = telefone1 and telefone1 not in ("N/D", "None", "", "—")

    if (has_whatsapp or has_celular) and has_email and has_decisor:
        return 100
    elif (has_whatsapp or has_celular) and has_decisor:
        return 85
    elif has_email and has_decisor:
        return 70
    elif has_celular:
        return 50
    elif has_fone:
        return 30
    else:
        return 10


def score_d5_saude(sancoes: str) -> float:
    """D5: Saúde Jurídica (peso 10%)"""
    s = safe_str(sancoes).lower()
    if not s or s in ("limpo", "nenhuma", "zero", "n/d"):
        return 100
    elif "ativa" in s or "ceis" in s or "cnep" in s:
        return 10
    elif "resolvida" in s or "historico" in s:
        return 70
    else:
        return 100  # assume clean if unclear


def score_d6_geo(ufs_atuacao: str) -> float:
    """D6: Diversificação Geográfica (peso 10%)"""
    if not ufs_atuacao or ufs_atuacao in ("N/D", "None", ""):
        return 30
    ufs = [u.strip() for u in str(ufs_atuacao).replace(";", ",").split(",") if u.strip()]
    n = len(ufs)
    if n >= 5:
        return 100
    elif n >= 3:
        return 75
    elif n >= 2:
        return 50
    else:
        return 30


def score_d7_upsell(capital_social: float, cnae_principal: str, fat_gov_mensal: float) -> float:
    """D7: Potencial de Upsell (peso 5%)"""
    bonus = 0
    if capital_social > 500_000:
        bonus += 25
    if fat_gov_mensal > 200_000:
        bonus += 25
    # Multiple sectors signal = higher upsell
    if cnae_principal and len(str(cnae_principal)) >= 7:
        bonus += 25
    # Cap at 100
    return min(bonus + 25, 100)


def compute_score(row: tuple) -> dict:
    """Compute 7D score for a lead row (0-indexed tuple from Leads tab)."""
    fat = safe_float(row[COL["fat_gov_mensal"]])
    contratos = safe_int(row[COL["contratos_6m"]])
    capital = safe_float(row[COL["capital_social"]])
    porte = safe_str(row[COL["porte"]])
    decisor = safe_str(row[COL["decisor"]])
    tel1 = safe_str(row[COL["telefone1"]])
    wa_flag = safe_str(row[COL["whatsapp_flag"]])
    tel2 = safe_str(row[COL["telefone2"]])
    email = safe_str(row[COL["email"]])
    sancoes = safe_str(row[COL["sancoes"]])
    ufs = safe_str(row[COL["ufs_atuacao"]])
    cnae = safe_str(row[COL["cnae_principal"]])

    d1 = score_d1_volume(fat)
    d2 = score_d2_frequencia(contratos)
    d3 = score_d3_porte(capital, porte, fat)
    d4 = score_d4_acessibilidade(decisor, tel1, wa_flag, tel2, email)
    d5 = score_d5_saude(sancoes)
    d6 = score_d6_geo(ufs)
    d7 = score_d7_upsell(capital, cnae, fat)

    final = (d1 * 0.25) + (d2 * 0.20) + (d3 * 0.15) + (d4 * 0.15) + (d5 * 0.10) + (d6 * 0.10) + (d7 * 0.05)

    tier = 4
    for t in [1, 2, 3]:
        if final >= TIER_THRESHOLDS[t]:
            tier = t
            break

    # Low-hanging fruit detection
    flags = []
    if contratos >= 5 and fat < 100_000:
        flags.append("Ganha sem otimizar")
    if fat > 500_000 and contratos <= 2:
        flags.append("Grande subutilizado")
    if ufs and len(str(ufs).split(",")) == 1 and cnae and len(cnae) >= 7:
        flags.append("Multi-setor inexplorado")

    return {
        "d1": round(d1, 1), "d2": round(d2, 1), "d3": round(d3, 1),
        "d4": round(d4, 1), "d5": round(d5, 1), "d6": round(d6, 1),
        "d7": round(d7, 1), "final": round(final, 1),
        "tier": tier, "flags": " | ".join(flags) if flags else "",
    }


# ── Tier label helper ───────────────────────────────────────────────────────

def tier_label(tier: int) -> str:
    return {1: "Tier 1 — Hot", 2: "Tier 2 — Warm", 3: "Tier 3 — Cold", 4: "Tier 4 — Descartado"}[tier]


def tier_action(tier: int) -> str:
    return {
        1: "WhatsApp direto",
        2: "WhatsApp + Email follow-up",
        3: "Email cadência automática",
        4: "Não abordar",
    }[tier]


# ── Main Pipeline ───────────────────────────────────────────────────────────

def load_and_score_leads(base_dir: Path) -> list:
    """Load leads from all 3 original Excels, score them, return sorted list."""
    all_leads = []  # list of (score_dict, row_tuple, region, wa_row_tuple_or_None)

    for rel_path, region in INPUT_FILES:
        fpath = base_dir / rel_path
        if not fpath.exists():
            print(f"  WARN: {fpath} nao encontrado, pulando.")
            continue

        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        ws_leads = wb["Leads"]
        ws_wa = wb["WhatsApp Outreach"]

        # Build CNPJ -> WhatsApp row mapping
        wa_headers = [c.value for c in next(ws_wa.iter_rows(min_row=1, max_row=1))]
        wa_by_empresa = {}
        for wa_row in ws_wa.iter_rows(min_row=2, values_only=True):
            empresa_key = safe_str(wa_row[1]).upper().strip()
            if empresa_key:
                wa_by_empresa[empresa_key] = wa_row

        # Read leads
        lead_headers = [c.value for c in next(ws_leads.iter_rows(min_row=1, max_row=1))]
        count = 0
        for row in ws_leads.iter_rows(min_row=2, values_only=True):
            empresa = safe_str(row[COL["empresa"]])
            if not empresa:
                continue

            scores = compute_score(row)

            # Match WhatsApp row
            empresa_key = empresa.upper().strip()
            wa_row = wa_by_empresa.get(empresa_key)

            all_leads.append({
                "scores": scores,
                "lead_row": row,
                "region": region,
                "wa_row": wa_row,
                "lead_headers": lead_headers,
                "wa_headers": wa_headers,
            })
            count += 1

        wb.close()
        print(f"  {region}: {count} leads carregados")

    # Sort by score descending
    all_leads.sort(key=lambda x: x["scores"]["final"], reverse=True)
    return all_leads


def select_top_pareto(leads: list, top_pct: float = 0.20) -> list:
    """Select top N% leads (Pareto principle)."""
    n = max(1, int(len(leads) * top_pct))
    selected = leads[:n]
    print(f"  Pareto {int(top_pct*100)}%: {n} leads selecionados de {len(leads)} total")
    return selected


def build_consolidated_excel(selected: list, all_leads: list, output_path: Path):
    """Build consolidated Excel preserving original structure + scoring columns."""
    wb = Workbook()

    # ── Tab 1: Resumo ───────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    tier_counts = {1: 0, 2: 0, 3: 0, 4: 0}
    for lead in selected:
        tier_counts[lead["scores"]["tier"]] += 1

    total_all = len(all_leads)
    total_selected = len(selected)
    avg_score = sum(l["scores"]["final"] for l in selected) / max(len(selected), 1)
    lhf_count = sum(1 for l in selected if l["scores"]["flags"])

    resumo_data = [
        ["Qualificação B2G — Engenharia/Construção (Consolidado 3 Regiões)"],
        [f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')} | Scoring 7 dimensões | Track A1 Consultoria"],
        [],
        ["Métricas", "Valor"],
        ["Leads brutos (3 regiões)", total_all],
        ["Leads selecionados (top 20%)", total_selected],
        ["Tier 1 — Hot (≥65)", tier_counts[1]],
        ["Tier 2 — Warm (≥48)", tier_counts[2]],
        ["Tier 3 — Cold (≥30)", tier_counts[3]],
        ["Tier 4 — Descartado (<30)", tier_counts[4]],
        ["Low-hanging fruit", lhf_count],
        ["Score médio", round(avg_score, 1)],
        [],
        ["Regiões consolidadas: Sul, Sudeste, Nordeste"],
        ["Oferta: Assessoria de Licitações R$1.500/mês (Track A1 MAYDAY)"],
    ]

    for row_data in resumo_data:
        ws_resumo.append(row_data)

    # Style resumo
    ws_resumo["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws_resumo["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    for cell in ws_resumo[4]:
        if cell.value:
            cell.font = Font(name="Calibri", bold=True, size=11)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
    ws_resumo.column_dimensions["A"].width = 35
    ws_resumo.column_dimensions["B"].width = 20

    # ── Tab 2: Leads (original 41 cols + 12 scoring cols) ───────────────
    ws_leads = wb.create_sheet("Leads")

    # Headers: original 41 + scoring
    if selected:
        orig_headers = list(selected[0]["lead_headers"])
    else:
        orig_headers = ["#", "Empresa"]  # fallback

    scoring_headers = [
        "Região", "Score D1 Volume", "Score D2 Freq", "Score D3 Porte",
        "Score D4 Acesso", "Score D5 Saúde", "Score D6 Geo", "Score D7 Upsell",
        "Score Final", "Tier", "Canal Sugerido", "Low-Hanging Fruit",
    ]

    all_headers = orig_headers + scoring_headers
    ws_leads.append(all_headers)

    # Style headers
    for col_idx, _ in enumerate(all_headers, 1):
        cell = ws_leads.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for i, lead in enumerate(selected, 1):
        row_data = list(lead["lead_row"])

        # Override # with new sequential number
        row_data[0] = i

        # Replace "Olá" with "Bom dia" in WhatsApp message (col 38 = Mensagem WhatsApp)
        if len(row_data) > 38 and row_data[38] and isinstance(row_data[38], str):
            row_data[38] = row_data[38].replace("Olá", "Bom dia").replace("Ol\xe1", "Bom dia")

        # Append scoring columns
        s = lead["scores"]
        row_data.extend([
            lead["region"],
            s["d1"], s["d2"], s["d3"], s["d4"], s["d5"], s["d6"], s["d7"],
            s["final"],
            tier_label(s["tier"]),
            tier_action(s["tier"]),
            s["flags"],
        ])

        ws_leads.append(row_data)

        # Color tier rows
        tier = s["tier"]
        if tier in TIER_FILLS:
            fill = TIER_FILLS[tier]
            for col_idx in range(1, len(all_headers) + 1):
                cell = ws_leads.cell(row=i + 1, column=col_idx)
                cell.border = THIN_BORDER
                # Only color the scoring columns area
                if col_idx > len(orig_headers):
                    cell.fill = fill

    # Column widths
    important_widths = {
        1: 5, 2: 12, 3: 35, 4: 25, 5: 20, 6: 25, 7: 6,
        13: 18, 14: 12, 15: 18, 17: 25, 18: 20, 19: 18,
    }
    for col, w in important_widths.items():
        ws_leads.column_dimensions[get_column_letter(col)].width = w

    # Scoring cols width
    score_start = len(orig_headers) + 1
    for j in range(len(scoring_headers)):
        ws_leads.column_dimensions[get_column_letter(score_start + j)].width = 15

    # Format Fat. Gov Mensal as currency
    for row_idx in range(2, len(selected) + 2):
        cell = ws_leads.cell(row=row_idx, column=COL["fat_gov_mensal"] + 1)
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

    # Auto-filter
    ws_leads.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(selected) + 1}"

    # Freeze panes
    ws_leads.freeze_panes = "D2"

    # ── Tab 3: WhatsApp Outreach (original 20 cols + scoring) ───────────
    ws_wa = wb.create_sheet("WhatsApp Outreach")

    if selected and selected[0]["wa_headers"]:
        wa_orig_headers = list(selected[0]["wa_headers"])
    else:
        wa_orig_headers = [
            "#", "Empresa", "Nome Fantasia", "Decisor", "Cargo", "WhatsApp",
            "Link wa.me", "Email", "Cidade", "Fat. Gov Mensal", "Contratos",
            "UFs", "Mensagem WhatsApp", "Enviada?", "Data Envio", "Resposta",
            "Data Resposta", "Follow-up", "Status", "Notas",
        ]

    wa_extra = ["Região", "Score Final", "Tier", "Low-Hanging Fruit"]
    wa_all_headers = wa_orig_headers + wa_extra
    ws_wa.append(wa_all_headers)

    # Style WA headers
    for col_idx in range(1, len(wa_all_headers) + 1):
        cell = ws_wa.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    wa_count = 0
    for i, lead in enumerate(selected, 1):
        wa_row = lead.get("wa_row")
        if not wa_row:
            continue

        row_data = list(wa_row)
        # Override # with new sequential
        row_data[0] = wa_count + 1

        # Replace "Olá" with "Bom dia" in WhatsApp message (col 12 = Mensagem WhatsApp)
        if row_data[12] and isinstance(row_data[12], str):
            row_data[12] = row_data[12].replace("Olá", "Bom dia").replace("Ol\xe1", "Bom dia")

        s = lead["scores"]
        row_data.extend([
            lead["region"],
            s["final"],
            tier_label(s["tier"]),
            s["flags"],
        ])

        ws_wa.append(row_data)
        wa_count += 1

        # Color tier
        tier = s["tier"]
        if tier in TIER_FILLS:
            for col_idx in range(len(wa_orig_headers) + 1, len(wa_all_headers) + 1):
                ws_wa.cell(row=wa_count + 1, column=col_idx).fill = TIER_FILLS[tier]

    # WA column widths
    wa_widths = {
        1: 5, 2: 35, 3: 25, 4: 25, 5: 20, 6: 18, 7: 30,
        8: 30, 9: 25, 10: 18, 11: 10, 12: 10, 13: 60,
        14: 10, 15: 12, 16: 10, 17: 12, 18: 12, 19: 12, 20: 25,
    }
    for col, w in wa_widths.items():
        ws_wa.column_dimensions[get_column_letter(col)].width = w

    # Mensagem column wrap text
    for row_idx in range(2, wa_count + 2):
        cell = ws_wa.cell(row=row_idx, column=13)  # Mensagem WhatsApp
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_wa.freeze_panes = "D2"

    # ── Tab 4: Ref_CRM ─────────────────────────────────────────────────
    ws_crm = wb.create_sheet("Ref_CRM")
    crm_data = [
        ["Campo", "Valores Possíveis", "Descrição"],
        ["Status Contato", "Novo / Contactado / Em Negociação / Fechado / Perdido", "Status atual do lead no funil"],
        ["Canal Contato", "WhatsApp / Email / LinkedIn / Telefone / Presencial", "Canal do primeiro contato"],
        ["Interesse (1-5)", "1=Muito Baixo / 2=Baixo / 3=Médio / 4=Alto / 5=Muito Alto", "Nível de interesse demonstrado"],
        ["Tier", "Tier 1 Hot / Tier 2 Warm / Tier 3 Cold / Tier 4 Descartado", "Classificação por scoring 7D"],
        ["Low-Hanging Fruit", "Ganha sem otimizar / Grande subutilizado / Multi-setor inexplorado", "Padrão especial detectado"],
    ]
    for row_data in crm_data:
        ws_crm.append(row_data)
    for cell in ws_crm[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws_crm.column_dimensions["A"].width = 20
    ws_crm.column_dimensions["B"].width = 55
    ws_crm.column_dimensions["C"].width = 40

    # ── Tab 5: Metodologia ──────────────────────────────────────────────
    ws_met = wb.create_sheet("Metodologia")
    met_data = [
        ["METODOLOGIA DE QUALIFICAÇÃO"],
        [],
        ["Scoring Multi-Dimensional (7 dimensões)"],
        ["Dimensão", "Peso", "Descrição"],
        ["D1 — Volume Governamental", "25%", "Fat. Gov Mensal: >500k=100, 200-500k=85, 100-200k=70, 50-100k=55, 20-50k=40, 5-20k=25, <5k=10"],
        ["D2 — Frequência de Participação", "20%", "Contratos/ano: 20+=100, 10-19=85, 5-9=70, 3-4=55, 1-2=40, 0=15"],
        ["D3 — Porte e Capacidade", "15%", "Capital social + porte (proxy: fat gov mensal quando capital indisponível)"],
        ["D4 — Acessibilidade do Decisor", "15%", "Celular+Email+Decisor=100, Celular+Decisor=85, Email+Decisor=70, Celular=50, Fixo=30, Nenhum=10"],
        ["D5 — Saúde Jurídica", "10%", "Limpo=100, Sanção resolvida=70, Sanção ativa=10, Irregular=0"],
        ["D6 — Diversificação Geográfica", "10%", "5+ UFs=100, 3-4=75, 2=50, 1=30"],
        ["D7 — Potencial de Upsell", "5%", "Capital>500k, Fat>200k, CNAE diverso = bônus cumulativo"],
        [],
        ["Classificação em Tiers"],
        ["Tier", "Score", "Ação"],
        ["Tier 1 — Hot", "≥65", "WhatsApp direto em 48h"],
        ["Tier 2 — Warm", "≥48", "WhatsApp + Email em 1 semana"],
        ["Tier 3 — Cold", "≥30", "Email cadência automática"],
        ["Tier 4 — Descartado", "<30", "Não abordar"],
        [],
        ["Low-Hanging Fruit (padrões especiais)"],
        ["Padrão", "Critério"],
        ["Ganha sem otimizar", "5+ contratos mas fat <100k — já participa, precisa escalar"],
        ["Grande subutilizado", "Fat >500k mas ≤2 contratos — potencial não explorado"],
        ["Multi-setor inexplorado", "1 UF + CNAE diverso — cross-sell possível"],
        [],
        ["Fontes de dados: PNCP, OpenCNPJ, Portal da Transparência"],
        [f"Data de qualificação: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        ["Regiões: Sul, Sudeste, Nordeste"],
        ["Oferta padrão: Assessoria de Licitações R$1.500/mês (Track A1 MAYDAY Strategy)"],
    ]
    for row_data in met_data:
        ws_met.append(row_data)
    ws_met["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws_met["A3"].font = Font(name="Calibri", bold=True, size=12)
    ws_met["A13"].font = Font(name="Calibri", bold=True, size=12)
    ws_met["A20"].font = Font(name="Calibri", bold=True, size=12)
    for cell in ws_met[4]:
        if cell.value:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
    for cell in ws_met[14]:
        if cell.value:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
    for cell in ws_met[21]:
        if cell.value:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
    ws_met.column_dimensions["A"].width = 35
    ws_met.column_dimensions["B"].width = 15
    ws_met.column_dimensions["C"].width = 80

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\n  Excel salvo: {output_path}")
    return tier_counts, wa_count


def build_markdown_report(selected: list, all_leads: list, output_path: Path):
    """Generate markdown summary report."""
    tier_counts = {1: 0, 2: 0, 3: 0, 4: 0}
    for lead in selected:
        tier_counts[lead["scores"]["tier"]] += 1

    avg_score = sum(l["scores"]["final"] for l in selected) / max(len(selected), 1)
    lhf_count = sum(1 for l in selected if l["scores"]["flags"])

    lines = []
    lines.append(f"# Qualificação B2G — Engenharia/Construção (Consolidado Sul+Sudeste+Nordeste)")
    lines.append(f"> Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Scoring 7 dimensões")
    lines.append("")
    lines.append("## Resumo")
    lines.append(f"- **Leads brutos:** {len(all_leads)} (3 regiões)")
    lines.append(f"- **Leads qualificados:** {len(selected)}")
    lines.append(f"- **Tier 1 (Hot):** {tier_counts[1]}")
    lines.append(f"- **Tier 2 (Warm):** {tier_counts[2]}")
    lines.append(f"- **Tier 3 (Cold):** {tier_counts[3]}")
    lines.append(f"- **Tier 4 (Descartado):** {tier_counts[4]}")
    lines.append(f"- **Low-hanging fruit:** {lhf_count}")
    lines.append(f"- **Score médio:** {avg_score:.1f}")
    lines.append("")

    # Top 20 leads
    lines.append("## Top 20 Leads — Prioridade Máxima")
    lines.append("")
    lines.append("| # | Empresa | Região | Score | Fat. Gov/Mês | Tier | Decisor | Canal |")
    lines.append("|---|---------|--------|-------|-------------|------|---------|-------|")

    for i, lead in enumerate(selected[:20], 1):
        s = lead["scores"]
        row = lead["lead_row"]
        empresa = safe_str(row[COL["empresa"]])[:40]
        decisor = safe_str(row[COL["decisor"]])[:25]
        fat = safe_float(row[COL["fat_gov_mensal"]])
        fat_str = f"R$ {fat:,.0f}" if fat else "N/D"
        lines.append(
            f"| {i} | {empresa} | {lead['region']} | {s['final']} | "
            f"{fat_str} | {tier_label(s['tier'])} | {decisor} | {tier_action(s['tier'])} |"
        )

    lines.append("")

    # Low-hanging fruit
    lhf_leads = [l for l in selected if l["scores"]["flags"]]
    if lhf_leads:
        lines.append("## Low-Hanging Fruit")
        lines.append("")
        lines.append("| Empresa | Flag | Score | Fat. Gov/Mês |")
        lines.append("|---------|------|-------|-------------|")
        for lead in lhf_leads[:10]:
            s = lead["scores"]
            row = lead["lead_row"]
            empresa = safe_str(row[COL["empresa"]])[:40]
            fat = safe_float(row[COL["fat_gov_mensal"]])
            fat_str = f"R$ {fat:,.0f}" if fat else "N/D"
            lines.append(f"| {empresa} | {s['flags']} | {s['final']} | {fat_str} |")
        lines.append("")

    # Next steps
    lines.append("## Próximos Passos")
    lines.append("1. `/cadencia-b2g engenharia --tier 1` — cadência WhatsApp para os hot leads")
    lines.append("2. `/proposta-b2g {CNPJ}` — proposta personalizada para os top 5")
    lines.append("3. Cruzar com editais abertos via SmartLic para timing de abordagem")
    lines.append("4. Expandir para novos setores (medicamentos, TI, segurança)")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Markdown salvo: {output_path}")


# ── Entry Point ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Qualify B2G leads — 7D scoring + Pareto consolidation")
    parser.add_argument("--top-pct", type=float, default=0.20, help="Top percentage to select (default 0.20)")
    args = parser.parse_args()

    base_dir = Path(__file__).parent.parent
    output_xlsx = base_dir / OUTPUT_FILE
    output_md = base_dir / MARKDOWN_FILE

    print("=" * 60)
    print("  /qualify-b2g — Qualificação B2G Consolidada")
    print("  Track A1: Assessoria de Licitações R$1.500/mês")
    print("=" * 60)
    print()

    print("[1/4] Carregando leads das 3 regiões...")
    all_leads = load_and_score_leads(base_dir)
    print(f"  Total: {len(all_leads)} leads carregados e scorados")
    print()

    print("[2/4] Selecionando top {:.0f}% (Pareto)...".format(args.top_pct * 100))
    selected = select_top_pareto(all_leads, args.top_pct)
    print()

    print("[3/4] Gerando Excel consolidado...")
    tier_counts, wa_count = build_consolidated_excel(selected, all_leads, output_xlsx)
    print(f"  {wa_count} leads com WhatsApp Outreach (mensagens originais preservadas)")
    print()

    print("[4/4] Gerando relatório Markdown...")
    build_markdown_report(selected, all_leads, output_md)
    print()

    # Summary
    print("=" * 60)
    print("  RESULTADO FINAL")
    print("=" * 60)
    print(f"  Leads brutos: {len(all_leads)}")
    print(f"  Selecionados (top {int(args.top_pct*100)}%): {len(selected)}")
    print(f"  Tier 1 Hot: {tier_counts[1]}")
    print(f"  Tier 2 Warm: {tier_counts[2]}")
    print(f"  Tier 3 Cold: {tier_counts[3]}")
    print(f"  Tier 4 Desc: {tier_counts[4]}")
    print(f"  WhatsApp ready: {wa_count}")
    print(f"  Excel: {output_xlsx}")
    print(f"  Markdown: {output_md}")
    print()


if __name__ == "__main__":
    main()
