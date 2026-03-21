#!/usr/bin/env python3
"""
Gerador de planilha Excel para Intel-Busca.

Gera planilha profissional com TODAS as oportunidades encontradas,
classificadas por compatibilidade CNAE e capacidade financeira.

Usage:
    python scripts/intel-excel.py --input docs/intel/intel-12345678000190-2026-03-18.json
    python scripts/intel-excel.py --input data.json --output custom.xlsx
"""

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# Windows console encoding fix
if sys.platform == "win32":
    import io
    try:
        if hasattr(sys.stdout, "buffer"):
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "buffer"):
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except (ValueError, AttributeError):
        pass

# Ensure scripts/ is on sys.path for lib imports
_scripts_dir = str(Path(__file__).resolve().parent)
if _scripts_dir not in sys.path:
    sys.path.insert(0, _scripts_dir)

from lib.intel_logging import setup_intel_logging

logger = setup_intel_logging("intel-excel")

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Design tokens (from report-b2g)
INK = "1B2A3D"
BG_SUBTLE = "F5F6F8"
WHITE = "FFFFFF"
GREEN_TEXT = "1B7A3D"
RED_TEXT = "B5342A"
AMBER_TEXT = "B8860B"
BLUE_TEXT = "1565C0"
GRAY_TEXT = "808080"
LINK_BLUE = "0563C1"

CURRENCY_FMT = '[$R$-416] #,##0.00'
DATE_FMT = "DD/MM/YYYY"
DATETIME_FMT = "DD/MM/YYYY HH:MM"
PCT_FMT = "0.0%"

# Regex to strip illegal XML 1.0 control characters that openpyxl rejects.
# Keeps tab (\x09), newline (\x0a), and carriage return (\x0d).
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")

# Column definitions: (header, width, align)
COLUMNS = [
    ("Recomendação", 16, "center"),           # 1  - analise.recomendacao_acao (PARTICIPAR/NÃO PARTICIPAR)
    ("N\u00ba", 5, "center"),                # 2
    ("Objeto", 55, "left"),                  # 3
    ("\u00d3rg\u00e3o", 35, "left"),         # 4
    ("UF", 5, "center"),                     # 5
    ("Munic\u00edpio", 20, "left"),          # 6
    ("Valor Estimado", 18, "right"),         # 7
    ("Modalidade", 20, "left"),              # 8
    ("Publica\u00e7\u00e3o", 13, "center"),  # 9
    ("Abertura Propostas", 18, "center"),    # 10
    ("Encerramento", 18, "center"),          # 11
    ("Urgência", 12, "center"),             # 12 - status_temporal (URGENTE/IMINENTE/PLANEJÁVEL)
    ("Distância (km)", 12, "right"),         # 13
    ("Custo Proposta", 14, "right"),         # 14
    ("Retorno Estimado", 14, "center"),      # 15
    ("Aderência Perfil", 14, "center"),      # 16 - _victory_fit_label
    ("Concorrência", 14, "center"),          # 17
    ("Fornecedores no Órgão", 14, "right"),  # 18
    ("Principal Fornecedor", 30, "left"),    # 19
    ("Desconto Mediano do Órgão", 16, "center"),  # 20
    ("Lance Sugerido", 18, "right"),         # 21
    ("P(Vitória)", 10, "center"),            # 22 - win probability
    ("População", 12, "right"),              # 23
    ("Compatível", 12, "center"),            # 24
    ("Confiança CNAE", 12, "center"),        # 25 - cnae_confidence %
    ("Dentro Capacidade", 14, "center"),     # 26
    ("Relevância", 10, "center"),            # 27
    ("Setor", 25, "left"),                   # 28
    ("Link PNCP", 12, "center"),             # 29
    # v2.0 columns
    ("Bid Score", 10, "center"),             # 30 - _bid_score._composite
    ("Compliance", 16, "center"),            # 31 - analise._compliance_summary
    ("Urgência Calc.", 14, "center"),        # 32 - analise._urgency.nivel + dias
    ("Risco Decomposto", 25, "left"),        # 33 - nivel_dificuldade decomposed
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _sanitize(value) -> str:
    """Remove illegal XML control characters from a string."""
    if value is None:
        return ""
    s = str(value)
    return _ILLEGAL_XML_RE.sub("", s)


def _parse_dt(value) -> datetime | None:
    """Parse datetime flexibly (ISO, BR, with/without time). Returns naive."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    s = str(value).strip()
    # Try ISO with timezone
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except (ValueError, AttributeError):
        pass
    # Try ISO without tz
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _safe_float(value) -> float | None:
    """Convert value to float, returning None on failure."""
    if value is None:
        return None
    try:
        if isinstance(value, str):
            # Handle BR format "1.234,56" -> "1234.56"
            value = value.replace(".", "").replace(",", ".")
        return float(value)
    except (ValueError, TypeError):
        return None


def _cnae_label(item: dict) -> str:
    """Build CNAE label from cnae_compatible field."""
    val = item.get("cnae_compatible")
    if val is True or str(val).upper() in ("TRUE", "SIM", "YES", "1"):
        return "SIM"
    if val is False or str(val).upper() in ("FALSE", "NAO", "N\u00c3O", "NO", "0"):
        return "N\u00c3O"
    if val is None:
        return "AVALIAR"
    # String values like "AVALIAR", "PARCIAL", etc.
    s = str(val).upper().strip()
    if s in ("SIM", "YES", "TRUE"):
        return "SIM"
    if s in ("NAO", "N\u00c3O", "NO", "FALSE"):
        return "N\u00c3O"
    return "AVALIAR"


def _capacity_label(valor: float | None, capacity_10x: float | None) -> str:
    """Determine if value is within financial capacity (10x capital)."""
    if valor is None or capacity_10x is None:
        return "N/D"
    return "SIM" if valor <= capacity_10x else "N\u00c3O"


def _format_cnpj(cnpj: str) -> str:
    """Format CNPJ: 12.345.678/0001-90."""
    d = re.sub(r"\D", "", str(cnpj))
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return str(cnpj)


def _format_brl(value) -> str:
    """Format value as R$ string."""
    v = _safe_float(value)
    if v is None:
        return "N/D"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ---------------------------------------------------------------------------
# Styles factory
# ---------------------------------------------------------------------------


def _make_styles():
    """Create all reusable styles."""
    header_font = Font(bold=True, color=WHITE, size=11)
    header_fill = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_fill = PatternFill(start_color=BG_SUBTLE, end_color=BG_SUBTLE, fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    green_font = Font(color=GREEN_TEXT, bold=True)
    red_font = Font(color=RED_TEXT, bold=True)
    amber_font = Font(color=AMBER_TEXT, bold=True)
    blue_font = Font(color=BLUE_TEXT, bold=True)
    gray_font = Font(color=GRAY_TEXT)
    link_font = Font(color=LINK_BLUE, underline="single")
    bold_font = Font(bold=True)

    return {
        "header_font": header_font,
        "header_fill": header_fill,
        "header_align": header_align,
        "alt_fill": alt_fill,
        "no_fill": no_fill,
        "thin_border": thin_border,
        "green_font": green_font,
        "red_font": red_font,
        "amber_font": amber_font,
        "blue_font": blue_font,
        "gray_font": gray_font,
        "link_font": link_font,
        "bold_font": bold_font,
    }


# ---------------------------------------------------------------------------
# WriteOnlyCell helpers
# ---------------------------------------------------------------------------


def _woc(ws, value=None, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Create a WriteOnlyCell with optional style attributes."""
    c = WriteOnlyCell(ws, value=value)
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if alignment is not None:
        c.alignment = alignment
    if border is not None:
        c.border = border
    if number_format is not None:
        c.number_format = number_format
    return c


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_oportunidades(wb: Workbook, items: list[dict], capacity_10x: float | None):
    """Build Sheet 1: Oportunidades (write-only mode)."""
    ws = wb.create_sheet("Oportunidades")
    st = _make_styles()

    # Column widths MUST be set before appending rows in write-only mode
    for col_idx, (_, width, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # --- Header row ---
    header_row = []
    for header, _, _ in COLUMNS:
        c = _woc(
            ws,
            value=header,
            font=st["header_font"],
            fill=st["header_fill"],
            alignment=st["header_align"],
            border=st["thin_border"],
        )
        header_row.append(c)
    ws.append(header_row)

    # --- Sort by valor desc (None values last) ---
    def sort_key(it):
        v = _safe_float(it.get("valor_estimado"))
        return v if v is not None else -1

    sorted_items = sorted(items, key=sort_key, reverse=True)

    # --- Data rows ---
    for row_num, item in enumerate(sorted_items, start=2):
        data_idx = row_num - 1  # 1-based row number
        is_alt = (row_num % 2) == 0
        row_fill = st["alt_fill"] if is_alt else st["no_fill"]

        valor = _safe_float(item.get("valor_estimado"))
        cnae_label = _cnae_label(item)
        cap_label = _capacity_label(valor, capacity_10x)
        kw_density = _safe_float(item.get("keyword_density"))

        # Sector name (human-readable)
        sector_name = item.get("sector_name", item.get("setor", ""))
        if not sector_name:
            # Fallback: use match_keywords as sector hint
            match_kw = item.get("match_keywords", item.get("keywords_matched", ""))
            if isinstance(match_kw, list):
                sector_name = ", ".join(str(k) for k in match_kw[:3])
            elif isinstance(match_kw, str):
                sector_name = match_kw[:50]

        # Link
        link_pncp = item.get("link_pncp", item.get("link", ""))

        # Helper: standard data cell with row fill and alignment
        def _dc(col_idx, value=None, font=None, fill=None, number_format=None):
            _, _, h_align = COLUMNS[col_idx - 1]
            eff_fill = fill if fill is not None else row_fill
            return _woc(
                ws,
                value=value,
                font=font,
                fill=eff_fill,
                alignment=Alignment(horizontal=h_align, vertical="top", wrap_text=True),
                border=st["thin_border"],
                number_format=number_format,
            )

        row = []

        # Col 1: Recomendação (analise.recomendacao_acao) — replaces old _delta_status
        analise = item.get("analise", {}) or {}
        if isinstance(analise, str):
            try:
                import json as _json
                analise = _json.loads(analise)
            except Exception:
                analise = {}
        rec_acao = analise.get("recomendacao_acao", "")
        # Normalize: extract just PARTICIPAR or NÃO PARTICIPAR from free text
        rec_label = ""
        if isinstance(rec_acao, str):
            upper = rec_acao.upper()
            if "NÃO PARTICIPAR" in upper or "NAO PARTICIPAR" in upper or "NÃO" in upper.split(".")[0]:
                rec_label = "NÃO PARTICIPAR"
            elif "PARTICIPAR" in upper:
                rec_label = "PARTICIPAR"
        if rec_label == "PARTICIPAR":
            rec_font = st["green_font"]
        elif rec_label == "NÃO PARTICIPAR":
            rec_font = st["red_font"]
        else:
            rec_font = st.get("gray_font")
            rec_label = rec_label or ""
        row.append(_dc(1, value=rec_label, font=rec_font))

        # Col 2: Row number
        row.append(_dc(2, value=data_idx))

        # Col 3: Objeto
        row.append(_dc(3, value=_sanitize(item.get("objeto", item.get("objetoCompra", "")))))

        # Col 4: Orgao
        row.append(_dc(4, value=_sanitize(item.get("orgao", item.get("nomeOrgao", "")))))

        # Col 5: UF
        row.append(_dc(5, value=item.get("uf", "")))

        # Col 6: Municipio
        row.append(_dc(6, value=_sanitize(item.get("municipio", ""))))

        # Col 7: Valor Estimado
        if valor is not None:
            row.append(_dc(7, value=valor, number_format=CURRENCY_FMT))
        else:
            row.append(_dc(7, value="Sigiloso"))

        # Col 8: Modalidade
        row.append(_dc(8, value=_sanitize(
            item.get("modalidade_nome", item.get("modalidadeNome", ""))
        )))

        # Col 9: Data Publicacao
        dt_pub = _parse_dt(item.get("data_publicacao", item.get("dataPublicacaoPncp")))
        row.append(_dc(9, value=dt_pub, number_format=DATE_FMT if dt_pub else None))

        # Col 10: Abertura Propostas
        dt_ab = _parse_dt(item.get("data_abertura_proposta", item.get("dataAberturaProposta")))
        row.append(_dc(10, value=dt_ab, number_format=DATETIME_FMT if dt_ab else None))

        # Col 11: Encerramento
        dt_enc = _parse_dt(item.get("data_encerramento_proposta", item.get("dataEncerramentoProposta")))
        row.append(_dc(11, value=dt_enc, number_format=DATETIME_FMT if dt_enc else None))

        # Col 12: Urgência (status_temporal)
        urg_val = item.get("status_temporal", "")
        urg_label = {
            "URGENTE": "URGENTE",
            "IMINENTE": "IMINENTE",
            "PLANEJAVEL": "PLANEJÁVEL",
            "SEM_DATA": "SEM DATA",
        }.get(str(urg_val).upper(), str(urg_val) if urg_val else "")
        if str(urg_val).upper() == "URGENTE":
            urg_font = st["red_font"]
        elif str(urg_val).upper() == "IMINENTE":
            urg_font = Font(name="Calibri", size=10, color="FF8C00")  # orange
        else:
            urg_font = st.get("green_font")
        row.append(_dc(12, value=urg_label, font=urg_font))

        # Col 13: Distancia (km)
        dist_data = item.get("distancia", {})
        dist_km = _safe_float(dist_data.get("km")) if isinstance(dist_data, dict) else None
        if dist_km is not None:
            row.append(_dc(13, value=dist_km, number_format='#,##0'))
        else:
            row.append(_dc(13, value=""))

        # Col 14: Custo Proposta
        custo_data = item.get("custo_proposta", {})
        custo_total = _safe_float(custo_data.get("total")) if isinstance(custo_data, dict) else None
        if custo_total is not None:
            row.append(_dc(14, value=custo_total, number_format=CURRENCY_FMT))
        else:
            row.append(_dc(14, value=""))

        # Col 15: Retorno Estimado (ROI)
        roi_data = item.get("roi_proposta", {})
        roi_class = roi_data.get("classificacao", "") if isinstance(roi_data, dict) else ""
        if roi_class in ("EXCELENTE", "BOM"):
            roi_font = st["green_font"]
        elif roi_class in ("MARGINAL", "DESFAVORAVEL"):
            roi_font = st["red_font"]
        elif roi_class == "MODERADO":
            roi_font = st["amber_font"]
        else:
            roi_font = None
        row.append(_dc(15, value=roi_class, font=roi_font))

        # Col 16: Aderencia Perfil (_victory_fit_label)
        fit_label = item.get("_victory_fit_label", "")
        if fit_label == "Excelente":
            fit_font = st["green_font"]
        elif fit_label == "Bom":
            fit_font = st["blue_font"]
        elif fit_label == "Moderado":
            fit_font = st["amber_font"]
        elif fit_label == "Baixo":
            fit_font = st["red_font"]
        else:
            fit_font = None
        row.append(_dc(16, value=fit_label, font=fit_font))

        # Col 17: Competicao (competition_level)
        comp = item.get("competitive_intel", {})
        comp_level = comp.get("competition_level", "") if isinstance(comp, dict) else ""
        if comp_level in ("ALTA", "MUITO_ALTA"):
            comp_font = st["green_font"]
        elif comp_level == "MEDIA":
            comp_font = st["amber_font"]
        elif comp_level == "BAIXA":
            comp_font = st["red_font"]
        else:
            comp_font = None
        row.append(_dc(17, value=comp_level, font=comp_font))

        # Col 18: Fornecedores (unique_suppliers)
        unique_sup = comp.get("unique_suppliers", "") if isinstance(comp, dict) else ""
        row.append(_dc(18, value=unique_sup))

        # Col 19: Fornecedor Recorrente (top supplier name)
        top_sups = comp.get("top_suppliers", []) if isinstance(comp, dict) else []
        fornecedor_recorrente = top_sups[0].get("nome", "")[:30] if top_sups else ""
        row.append(_dc(19, value=_sanitize(fornecedor_recorrente)))

        # Col 20: Desc. Mediano Orgao
        bench = item.get("price_benchmark", {})
        desc_med = bench.get("desconto_mediano_orgao") if isinstance(bench, dict) else None
        if desc_med is not None:
            row.append(_dc(20, value=desc_med, number_format=PCT_FMT))
        else:
            row.append(_dc(20, value=""))

        # Col 21: Lance Sugerido (single column from _bid_simulation or price_benchmark)
        bid_sim = item.get("_bid_simulation", {})
        lance_sugerido = None
        if isinstance(bid_sim, dict):
            lance_sugerido = _safe_float(bid_sim.get("lance_sugerido"))
        if lance_sugerido is None:
            # Fallback: use price_benchmark min as single value
            lance_sugerido = _safe_float(bench.get("valor_sugerido_min")) if isinstance(bench, dict) else None
        if lance_sugerido is not None:
            row.append(_dc(21, value=lance_sugerido, number_format=CURRENCY_FMT))
        else:
            row.append(_dc(21, value=""))

        # Col 22: P(Vitoria) — win probability
        win_prob = None
        if isinstance(bid_sim, dict):
            win_prob = _safe_float(bid_sim.get("win_probability"))
        if win_prob is not None:
            if win_prob >= 0.50:
                prob_font = st["green_font"]
            elif win_prob >= 0.30:
                prob_font = st["amber_font"]
            else:
                prob_font = st["red_font"]
            row.append(_dc(22, value=win_prob, font=prob_font, number_format="0%"))
        else:
            row.append(_dc(22, value=""))

        # Col 23: Populacao
        ibge_data = item.get("ibge", {})
        pop = ibge_data.get("populacao") if isinstance(ibge_data, dict) else None
        if pop is not None:
            row.append(_dc(23, value=pop, number_format='#,##0'))
        else:
            row.append(_dc(23, value=""))

        # Col 24: Compativel CNAE
        if cnae_label == "SIM":
            cnae_font = st["green_font"]
        elif cnae_label == "N\u00c3O":
            cnae_font = st["red_font"]
        else:
            cnae_font = st["amber_font"]
        row.append(_dc(24, value=cnae_label, font=cnae_font))

        # Col 25: Confianca CNAE (cnae_confidence as percentage)
        cnae_conf = _safe_float(item.get("cnae_confidence"))
        if cnae_conf is not None:
            # Normalize: if value is 0-1, convert to 0-100 for display
            if cnae_conf <= 1.0:
                cnae_conf_display = cnae_conf
            else:
                cnae_conf_display = cnae_conf / 100.0
            if cnae_conf_display >= 0.70:
                conf_font = st["green_font"]
            elif cnae_conf_display >= 0.35:
                conf_font = st["amber_font"]
            else:
                conf_font = st["red_font"]
            row.append(_dc(25, value=cnae_conf_display, font=conf_font, number_format="0%"))
        else:
            row.append(_dc(25, value=""))

        # Col 26: Dentro Capacidade
        if cap_label == "SIM":
            cap_font = st["green_font"]
        elif cap_label == "N\u00c3O":
            cap_font = st["red_font"]
        else:
            cap_font = st["gray_font"]
        row.append(_dc(26, value=cap_label, font=cap_font))

        # Col 27: Relevancia — based on cnae_confidence + capacity
        # Alta: confidence >= 0.8 AND within capacity
        # Media: confidence >= 0.5 (or no confidence data)
        # Baixa: confidence < 0.5
        _cnae_conf_raw = _safe_float(item.get("cnae_confidence"))
        if _cnae_conf_raw is not None:
            # Normalize to 0-1 range (same logic as col 24)
            _cnae_conf_norm = _cnae_conf_raw if _cnae_conf_raw <= 1.0 else _cnae_conf_raw / 100.0
            _within_cap = cap_label == "SIM" or cap_label == "N/D"
            if _cnae_conf_norm >= 0.8 and _within_cap:
                rel_label, rel_font = "Alta", st["green_font"]
            elif _cnae_conf_norm >= 0.5:
                rel_label, rel_font = "M\u00e9dia", st["amber_font"]
            else:
                rel_label, rel_font = "Baixa", st["red_font"]
        else:
            rel_label, rel_font = "M\u00e9dia", st["amber_font"]
        row.append(_dc(27, value=rel_label, font=rel_font))

        # Col 28: Setor
        row.append(_dc(28, value=_sanitize(sector_name)))

        # Col 29: Link PNCP — use HYPERLINK formula (write-only mode doesn't support cell.hyperlink)
        if link_pncp:
            safe_url = str(link_pncp).replace('"', '%22')
            link_val = f'=HYPERLINK("{safe_url}","Abrir")'
            row.append(_dc(29, value=link_val, font=st["link_font"]))
        else:
            row.append(_dc(29, value=""))

        # Col 30: Bid Score (_bid_score._composite as percentage)
        bid_score_raw = (item.get("_bid_score") or {}).get("_composite")
        bid_score = _safe_float(bid_score_raw)
        if bid_score is not None:
            bid_score_str = f"{bid_score:.0%}"
            if bid_score >= 0.70:
                bs_font = st["green_font"]
            elif bid_score >= 0.45:
                bs_font = st["amber_font"]
            else:
                bs_font = st["red_font"]
            row.append(_dc(30, value=bid_score_str, font=bs_font))
        else:
            row.append(_dc(30, value=""))

        # Col 31: Compliance (analise._compliance_summary)
        compliance = (analise.get("_compliance_summary") or "") if isinstance(analise, dict) else ""
        row.append(_dc(31, value=_sanitize(compliance)))

        # Col 32: Urgência Calc. (analise._urgency.nivel + dias_restantes)
        urgency = (analise.get("_urgency") or {}) if isinstance(analise, dict) else {}
        if isinstance(urgency, dict):
            nivel_u = urgency.get("nivel", "")
            dias = urgency.get("dias_restantes")
            urgency_str = f"{nivel_u} ({dias}d)" if dias is not None else str(nivel_u) if nivel_u else ""
        else:
            urgency_str = ""
        if urgency_str:
            nivel_upper = str(urgency.get("nivel", "")).upper() if isinstance(urgency, dict) else ""
            if nivel_upper in ("CRITICO", "URGENTE"):
                uc_font = st["red_font"]
            elif nivel_upper in ("ALTO", "IMINENTE"):
                uc_font = Font(name="Calibri", size=10, color="FF8C00")  # orange
            elif nivel_upper in ("MEDIO", "MODERADO"):
                uc_font = st["amber_font"]
            else:
                uc_font = st["green_font"]
        else:
            uc_font = None
        row.append(_dc(32, value=urgency_str, font=uc_font))

        # Col 33: Risco Decomposto (analise.nivel_dificuldade broken down by dimension)
        dif = (analise.get("nivel_dificuldade") or {}) if isinstance(analise, dict) else {}
        if isinstance(dif, dict):
            parts = []
            for key in ("tecnico", "prazo", "regulatorio", "logistico", "financeiro"):
                val = dif.get(key)
                if isinstance(val, (int, float)):
                    parts.append(f"{key[:3].upper()}:{val:.0f}")
            risco_str = " ".join(parts) if parts else str(dif.get("geral", ""))
        else:
            risco_str = str(dif) if dif else ""
        row.append(_dc(33, value=_sanitize(risco_str)))

        ws.append(row)

    # --- Total row ---
    if sorted_items:
        total_row_idx = len(sorted_items) + 2
        total_row = []
        for col_idx in range(1, len(COLUMNS) + 1):
            _, _, h_align = COLUMNS[col_idx - 1]
            if col_idx == 2:
                # Nº column: show count
                c = _woc(
                    ws, value=len(sorted_items),
                    font=st["bold_font"],
                    border=st["thin_border"],
                    alignment=Alignment(horizontal=h_align, vertical="top"),
                )
            elif col_idx == 6:
                # Municipio column: "TOTAL:" label
                c = _woc(
                    ws, value="TOTAL:",
                    font=st["bold_font"],
                    border=st["thin_border"],
                    alignment=Alignment(horizontal=h_align, vertical="top"),
                )
            elif col_idx == 7:
                # Valor Estimado column: SUM formula (column G)
                c = _woc(
                    ws, value=f"=SUM(G2:G{total_row_idx - 1})",
                    font=st["bold_font"],
                    border=st["thin_border"],
                    number_format=CURRENCY_FMT,
                    alignment=Alignment(horizontal=h_align, vertical="top"),
                )
            else:
                c = _woc(
                    ws, value=None,
                    border=st["thin_border"],
                    alignment=Alignment(horizontal=h_align, vertical="top"),
                )
            total_row.append(c)
        ws.append(total_row)


def _build_resumo_uf(wb: Workbook, items: list[dict]):
    """Build Sheet 2: Resumo por UF (write-only mode)."""
    ws = wb.create_sheet("Resumo por UF")
    st = _make_styles()

    # Aggregate
    uf_data: dict[str, dict] = defaultdict(lambda: {
        "qtd_total": 0, "qtd_compat": 0, "valor_total": 0.0, "valor_compat": 0.0,
    })
    for item in items:
        uf = item.get("uf", "N/D") or "N/D"
        d = uf_data[uf]
        d["qtd_total"] += 1
        valor = _safe_float(item.get("valor_estimado")) or 0.0
        d["valor_total"] += valor
        if _cnae_label(item) == "SIM":
            d["qtd_compat"] += 1
            d["valor_compat"] += valor

    headers = ["UF", "Qtd Total", "Qtd Compat\u00edvel", "Valor Total", "Valor Compat\u00edvel"]
    widths = [8, 12, 16, 20, 20]

    # Set column widths before appending
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    # Header row
    header_row = []
    for h in headers:
        c = _woc(
            ws, value=h,
            font=st["header_font"],
            fill=st["header_fill"],
            alignment=st["header_align"],
            border=st["thin_border"],
        )
        header_row.append(c)
    ws.append(header_row)

    sorted_ufs = sorted(uf_data.items())
    for uf, d in sorted_ufs:
        row = [
            _woc(ws, value=uf, border=st["thin_border"]),
            _woc(ws, value=d["qtd_total"], border=st["thin_border"]),
            _woc(ws, value=d["qtd_compat"], border=st["thin_border"]),
            _woc(ws, value=d["valor_total"], border=st["thin_border"], number_format=CURRENCY_FMT),
            _woc(ws, value=d["valor_compat"], border=st["thin_border"], number_format=CURRENCY_FMT),
        ]
        ws.append(row)

    # Totals
    if uf_data:
        tr = len(uf_data) + 2
        total_row = [
            _woc(ws, value="TOTAL", font=st["bold_font"], border=st["thin_border"]),
            _woc(ws, value=f"=SUM(B2:B{tr-1})", font=st["bold_font"], border=st["thin_border"]),
            _woc(ws, value=f"=SUM(C2:C{tr-1})", font=st["bold_font"], border=st["thin_border"]),
            _woc(ws, value=f"=SUM(D2:D{tr-1})", font=st["bold_font"], border=st["thin_border"], number_format=CURRENCY_FMT),
            _woc(ws, value=f"=SUM(E2:E{tr-1})", font=st["bold_font"], border=st["thin_border"], number_format=CURRENCY_FMT),
        ]
        ws.append(total_row)


def _build_resumo_modalidade(wb: Workbook, items: list[dict]):
    """Build Sheet 3: Resumo por Modalidade (write-only mode)."""
    ws = wb.create_sheet("Resumo por Modalidade")
    st = _make_styles()

    mod_data: dict[str, dict] = defaultdict(lambda: {"qtd": 0, "valor_total": 0.0})
    for item in items:
        mod = _sanitize(
            item.get("modalidade_nome", item.get("modalidadeNome", ""))
        ) or "N/D"
        d = mod_data[mod]
        d["qtd"] += 1
        d["valor_total"] += _safe_float(item.get("valor_estimado")) or 0.0

    headers = ["Modalidade", "Qtd", "Valor Total"]
    widths = [35, 10, 22]

    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    header_row = []
    for h in headers:
        c = _woc(
            ws, value=h,
            font=st["header_font"],
            fill=st["header_fill"],
            alignment=st["header_align"],
            border=st["thin_border"],
        )
        header_row.append(c)
    ws.append(header_row)

    for mod, d in sorted(mod_data.items()):
        row = [
            _woc(ws, value=mod, border=st["thin_border"]),
            _woc(ws, value=d["qtd"], border=st["thin_border"]),
            _woc(ws, value=d["valor_total"], border=st["thin_border"], number_format=CURRENCY_FMT),
        ]
        ws.append(row)

    if mod_data:
        tr = len(mod_data) + 2
        total_row = [
            _woc(ws, value="TOTAL", font=st["bold_font"], border=st["thin_border"]),
            _woc(ws, value=f"=SUM(B2:B{tr-1})", font=st["bold_font"], border=st["thin_border"]),
            _woc(ws, value=f"=SUM(C2:C{tr-1})", font=st["bold_font"], border=st["thin_border"], number_format=CURRENCY_FMT),
        ]
        ws.append(total_row)


def _build_metadata(wb: Workbook, data: dict, items: list[dict]):
    """Build Sheet 4: Metadata (write-only mode)."""
    ws = wb.create_sheet("Metadata")
    st = _make_styles()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    empresa = data.get("empresa", {})
    busca = data.get("busca", {})
    stats = data.get("stats", data.get("estatisticas", {}))

    # Parse capital social
    capital_raw = empresa.get("capital_social")
    capital = _safe_float(capital_raw)
    capacity_10x = capital * 10 if capital else None

    # Count compatible
    total = len(items)
    compat_count = sum(1 for i in items if _cnae_label(i) == "SIM")
    incompat_count = total - compat_count
    valor_compat = sum(
        _safe_float(i.get("valor_estimado")) or 0
        for i in items
        if _cnae_label(i) == "SIM"
    )

    # CNAE info
    cnae_principal = empresa.get("cnae_principal", {})
    if isinstance(cnae_principal, dict):
        cnae_str = f"{cnae_principal.get('codigo', '')} - {cnae_principal.get('descricao', '')}"
    else:
        cnae_str = str(cnae_principal) if cnae_principal else "N/D"

    # UFs
    ufs = busca.get("ufs", [])
    if isinstance(ufs, list):
        ufs_str = ", ".join(str(u) for u in ufs)
    else:
        ufs_str = str(ufs) if ufs else "N/D"

    # Period
    data_ini = busca.get("data_inicio", busca.get("dataInicial", ""))
    data_fim = busca.get("data_fim", busca.get("dataFinal", ""))
    dt_ini = _parse_dt(data_ini)
    dt_fim = _parse_dt(data_fim)
    if dt_ini and dt_fim:
        periodo_str = f"{dt_ini.strftime('%d/%m/%Y')} a {dt_fim.strftime('%d/%m/%Y')}"
        dias = (dt_fim - dt_ini).days
    else:
        periodo_str = f"{data_ini} a {data_fim}" if data_ini else "N/D"
        dias = busca.get("dias", "N/D")

    setor = busca.get("setor", busca.get("setor_mapeado", "N/D"))
    total_bruto = stats.get("total_bruto", stats.get("total_pncp", total))

    # SICAF / Sanctions enrichment data
    sicaf = empresa.get("sicaf", {})
    sicaf_status = sicaf.get("status", "N/D") if isinstance(sicaf, dict) else "N/D"
    sancoes = empresa.get("sancoes", {})
    sancionada = empresa.get("sancionada", False)
    restricao_sicaf = empresa.get("restricao_sicaf")

    sancoes_str = "Nenhuma"
    if isinstance(sancoes, dict) and sancionada:
        ativas = [k.upper() for k, v in sancoes.items() if v and k not in ("sancionada", "inconclusive")]
        sancoes_str = ", ".join(ativas) if ativas else "SIM (detalhes indispon\u00edveis)"

    restricao_str = "SIM" if restricao_sicaf else ("N\u00c3O" if restricao_sicaf is not None else "N/D")

    label_font = Font(bold=True)
    label_fill = PatternFill(start_color="E8EBF0", end_color="E8EBF0", fill_type="solid")

    rows = [
        ("CNPJ", _format_cnpj(empresa.get("cnpj", "N/D"))),
        ("Raz\u00e3o Social", empresa.get("razao_social", "N/D")),
        ("Atividade Principal", cnae_str),
        ("Capital Social", _format_brl(capital) if capital else "N/D"),
        ("Capacidade (10\u00d7)", _format_brl(capacity_10x) if capacity_10x else "N/D"),
        ("", ""),  # separator
        ("Cadastro Federal (SICAF)", sicaf_status),
        ("Restri\u00e7\u00e3o Cadastral", restricao_str),
        ("San\u00e7\u00f5es Ativas", sancoes_str),
        ("Empresa Sancionada", "SIM \u26d4" if sancionada else "N\u00c3O \u2705"),
        ("", ""),  # separator
        ("UFs Buscadas", ufs_str),
        ("Per\u00edodo", periodo_str),
        ("Dias", dias),
        ("Setor Mapeado", setor),
        ("Publicações PNCP Consultadas", total_bruto),
        ("Oportunidades Identificadas", compat_count),
        ("Valor Total das Oportunidades", _format_brl(valor_compat)),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Script", "intel-collect.py + intel-enrich.py + intel-excel.py"),
        ("", ""),  # separator
        ("NOTA - Aba 'Oportunidades'", f"{compat_count} oportunidades abertas identificadas em {ufs_str}. CNAE compatível + prazo vigente."),
    ]

    for label, value in rows:
        label_cell = _woc(
            ws, value=label,
            font=label_font,
            fill=label_fill,
            border=st["thin_border"],
        )
        val_cell = _woc(
            ws, value=value,
            border=st["thin_border"],
            alignment=Alignment(wrap_text=True),
        )
        ws.append([label_cell, val_cell])


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------


def generate_excel(data: dict, output_path: str) -> str:
    """
    Generate Excel workbook from intel-collect JSON data.

    Args:
        data: Parsed JSON dict with keys: empresa, busca, oportunidades/items, stats.
        output_path: Output .xlsx file path.

    Returns:
        Absolute path to generated file.
    """
    # Extract items — support multiple key names
    items = (
        data.get("oportunidades")
        or data.get("items")
        or data.get("editais")
        or data.get("resultados")
        or []
    )

    if not isinstance(items, list):
        logger.warning("campo de oportunidades nao e lista, tipo=%s", type(items).__name__)
        items = []

    # Compute capacity
    empresa = data.get("empresa", {})
    capital = _safe_float(empresa.get("capital_social"))
    capacity_10x = capital * 10 if capital else None

    # write_only=True: streams rows directly to disk — no in-memory cell graph.
    # Reduces RAM from ~500MB to ~10MB for 8000+ rows.
    wb = Workbook(write_only=True)

    # Split: main sheet = compatible + not expired; reference = all
    compat_items = [
        it for it in items
        if _cnae_label(it) == "SIM"
        and str(it.get("status_temporal", "")).upper() != "EXPIRADO"
    ]

    # Sheet 1: Oportunidades (only CNAE-compatible items)
    _build_oportunidades(wb, compat_items, capacity_10x)

    # Sheet 2: Resumo por UF
    _build_resumo_uf(wb, items)

    # Sheet 3: Resumo por Modalidade
    _build_resumo_modalidade(wb, items)

    # Sheet 4: Metadata
    _build_metadata(wb, data, items)

    # Save
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    return os.path.abspath(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    """Entry point for intel-excel CLI."""
    from lib.constants import INTEL_VERSION
    from lib.cli_validation import validate_input_file

    parser = argparse.ArgumentParser(
        description="Gera planilha Excel profissional a partir do JSON do intel-collect.py.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Exemplos:
  python scripts/intel-excel.py --input docs/intel/intel-12345678000190-slug-2026-03-18.json
  python scripts/intel-excel.py --input data.json --output planilha.xlsx""",
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Caminho do JSON de entrada (output do intel-collect.py). Deve existir.",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Caminho do .xlsx de saida (default: mesmo basename do input com extensao .xlsx)",
    )
    parser.add_argument("--version", action="version",
                        version=f"%(prog)s {INTEL_VERSION}")
    args = parser.parse_args()

    # ── Validate arguments ──
    validate_input_file(args.input)

    input_path = args.input

    # Default output: same name with .xlsx
    if args.output:
        output_path = args.output
    else:
        base = os.path.splitext(input_path)[0]
        output_path = f"{base}.xlsx"

    # Load JSON
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Generate
    abs_path = generate_excel(data, output_path)

    # Summary
    items = (
        data.get("oportunidades")
        or data.get("items")
        or data.get("editais")
        or data.get("resultados")
        or []
    )
    size_kb = os.path.getsize(abs_path) / 1024
    n = len(items) if isinstance(items, list) else 0

    compat = sum(
        1 for i in items
        if _cnae_label(i) == "SIM" and str(i.get("status_temporal", "")).upper() != "EXPIRADO"
    ) if isinstance(items, list) else 0

    print(f"Excel gerado: {abs_path} ({compat} oportunidades, {size_kb:.0f}KB)")
    print(f"  Abas: Oportunidades, Resumo por UF, Resumo por Modalidade, Metadata")


if __name__ == "__main__":
    main()
