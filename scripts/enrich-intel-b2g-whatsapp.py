#!/usr/bin/env python3
"""
Enrich Intel B2G spreadsheets with WhatsApp outreach columns.

Reads existing .xlsx files from docs/intel-b2g/, adds:
- WhatsApp Numero (cleaned 55+DDD+number)
- Link wa.me (clickable)
- Mensagem WhatsApp (personalized per lead)
- Tracking columns (Mensagem Enviada?, Resposta WhatsApp)
- New "WhatsApp Outreach" sheet with only contactable leads

Usage:
    python scripts/enrich-intel-b2g-whatsapp.py                    # All xlsx in docs/intel-b2g/
    python scripts/enrich-intel-b2g-whatsapp.py --file leads-engenharia-sul-2026-03-06.xlsx
"""
import argparse
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ============================================================
# CONFIG
# ============================================================

INTEL_DIR = Path(__file__).parent.parent / "docs" / "intel-b2g"

SETOR_DEFAULT = "engenharia e construcao civil"

OFERTA = {
    "valor": "R$1.500/mês",
    "remetente": "Tiago Sasaki",
}

# Column indices (0-based) in existing Leads sheet
COL_MAP = {
    "num": 0,
    "relevancia": 1,
    "empresa": 2,
    "nome_fantasia": 3,
    "cnpj": 4,
    "cidade_sede": 5,
    "uf_sede": 6,
    "cnae": 7,
    "porte": 8,
    "capital_social": 9,
    "ufs_atuacao": 10,
    "municipios": 11,
    "fat_gov_mensal": 12,
    "contratos": 13,
    "valor_total": 14,
    "orgaos": 15,
    "decisor": 16,
    "cargo_decisor": 17,
    "telefone1": 18,
    "whatsapp_flag": 19,
    "telefone2": 20,
    "email": 21,
    "sancoes": 22,
    "objetos": 23,
    # Tracking columns 24-35
    "status_contato": 24,
    "observacoes": 35,
}

# New columns to add after Observacoes
NEW_COLS = [
    "WhatsApp Numero",
    "Link wa.me",
    "Mensagem WhatsApp",
    "Mensagem Enviada?",
    "Resposta WhatsApp",
]

# Styles
HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WA_GREEN_FILL = PatternFill(start_color="25D366", end_color="25D366", fill_type="solid")
WA_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(color="0563C1", underline="single", size=10)
MSG_FONT = Font(size=9, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def extract_first_name(full_name: str) -> str:
    """Extract first name from full name, title-cased."""
    if not full_name or full_name == "N/D":
        return ""
    parts = full_name.strip().split()
    if not parts:
        return ""
    name = parts[0].strip().title()
    # Skip common prefixes
    if name.lower() in ("dr", "dr.", "dra", "dra.", "sr", "sr.", "sra", "sra."):
        return parts[1].strip().title() if len(parts) > 1 else name
    return name


def clean_phone_for_whatsapp(phone: str) -> str:
    """
    Clean phone number for wa.me link.
    Returns 55+DDD+number (digits only) or empty string.
    """
    if not phone:
        return ""
    # Remove all non-digits
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) < 10:
        return ""
    # If already starts with 55, return as-is
    if digits.startswith("55") and len(digits) >= 12:
        return digits
    # Assume Brazilian: DDD + number
    return f"55{digits}"


def is_likely_cellphone(phone: str) -> bool:
    """
    Check if phone is likely a cellphone (can receive WhatsApp).
    Brazilian cellphones: (DD) 9XXXX-XXXX (9 digits starting with 9)
    Old format: (DD) [6-9]XXX-XXXX (8 digits starting with 6-9)
    """
    if not phone:
        return False
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) < 10:
        return False
    # Remove country code if present
    if digits.startswith("55"):
        digits = digits[2:]
    # Remove DDD (2 digits)
    number = digits[2:]
    # 9 digits starting with 9 = confirmed cellphone
    if len(number) == 9 and number.startswith("9"):
        return True
    # 8 digits starting with 6-9 = likely old cellphone format
    if len(number) == 8 and number[0] in "6789":
        return True
    return False


def detect_setor_from_filename(filename: str) -> str:
    """Try to extract sector from filename."""
    name = filename.lower()
    if "engenharia" in name or "construcao" in name:
        return "engenharia e construcao civil"
    if "saude" in name or "medicamento" in name:
        return "saude e medicamentos"
    if "tecnologia" in name or "software" in name or "ti" in name:
        return "tecnologia da informacao"
    if "limpeza" in name or "facilities" in name:
        return "facilities e limpeza"
    if "aliment" in name:
        return "alimentacao e nutricao"
    return SETOR_DEFAULT


def build_whatsapp_message(
    decisor: str,
    nome_fantasia: str,
    empresa: str,
    setor: str,
    ufs: str,
    contratos: int,
) -> str:
    """Build personalized WhatsApp message for a lead."""
    first_name = extract_first_name(decisor)
    company = nome_fantasia if nome_fantasia and nome_fantasia != "N/D" else empresa
    if not company:
        company = "sua empresa"

    # Greeting
    if first_name:
        greeting = f"Olá, {first_name}! Tudo bem?"
    else:
        greeting = f"Olá! Tudo bem? Falo com o responsável por licitações da {company}?"

    # UFs formatting
    ufs_str = ufs if ufs else "diversas regiões"

    msg = f"""{greeting}

Me chamo Tiago Sasaki, sou engenheiro civil com 7 anos de experiência como servidor público efetivo em SC. Hoje atendo empresas de engenharia do Brasil todo, ajudando a não perderem editais relevantes.

Vi que a {company} atua com licitações em {ufs_str}. Faço um trabalho de consolidação de editais: monitoro PNCP, Portal de Compras e ComprasGov diariamente, filtro o que é relevante pro perfil de vocês e entrego um relatório semanal organizado.

O valor é {OFERTA['valor']}.

Posso te mandar um exemplo com editais reais da sua região?

Abraço,
Tiago Sasaki"""

    return msg.strip()


def enrich_workbook(filepath: Path) -> dict:
    """
    Add WhatsApp columns to existing Leads sheet and create WhatsApp Outreach sheet.
    Returns stats dict.
    """
    print(f"\n{'='*60}")
    print(f"Processing: {filepath.name}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath)
    ws_leads = wb["Leads"]

    setor = detect_setor_from_filename(filepath.name)
    print(f"  Setor detected: {setor}")

    # Check if already enriched
    existing_headers = [cell.value for cell in ws_leads[1]]
    already_enriched = "Mensagem WhatsApp" in existing_headers
    if already_enriched:
        print("  Already enriched — updating messages + rebuilding WhatsApp Outreach")
        wa_num_col = existing_headers.index("WhatsApp Numero") + 1
        wa_link_col = existing_headers.index("Link wa.me") + 1
        wa_msg_col = existing_headers.index("Mensagem WhatsApp") + 1
        wa_sent_col = existing_headers.index("Mensagem Enviada?") + 1
        wa_resp_col = existing_headers.index("Resposta WhatsApp") + 1
    else:
        # Add new column headers
        start_col = ws_leads.max_column + 1
        for i, col_name in enumerate(NEW_COLS):
            col_idx = start_col + i
            cell = ws_leads.cell(row=1, column=col_idx, value=col_name)
            cell.fill = WA_GREEN_FILL
            cell.font = WA_HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

        wa_num_col = start_col
        wa_link_col = start_col + 1
        wa_msg_col = start_col + 2
        wa_sent_col = start_col + 3
        wa_resp_col = start_col + 4

        # Set column widths
        ws_leads.column_dimensions[get_column_letter(wa_num_col)].width = 16
        ws_leads.column_dimensions[get_column_letter(wa_link_col)].width = 35
        ws_leads.column_dimensions[get_column_letter(wa_msg_col)].width = 60
        ws_leads.column_dimensions[get_column_letter(wa_sent_col)].width = 16
        ws_leads.column_dimensions[get_column_letter(wa_resp_col)].width = 25

    # Process each lead row
    stats = {"total": 0, "with_cellphone": 0, "with_message": 0, "no_phone": 0}
    wa_leads = []  # For WhatsApp Outreach sheet

    for row_idx in range(2, ws_leads.max_row + 1):
        row_data = [ws_leads.cell(row=row_idx, column=c + 1).value for c in range(ws_leads.max_column)]

        # Skip empty rows
        if not row_data[COL_MAP["empresa"]]:
            continue

        stats["total"] += 1

        empresa = str(row_data[COL_MAP["empresa"]] or "")
        nome_fantasia = str(row_data[COL_MAP["nome_fantasia"]] or "")
        decisor = str(row_data[COL_MAP["decisor"]] or "N/D")
        ufs = str(row_data[COL_MAP["ufs_atuacao"]] or "")
        contratos = row_data[COL_MAP["contratos"]] or 0
        telefone1 = str(row_data[COL_MAP["telefone1"]] or "")
        telefone2 = str(row_data[COL_MAP["telefone2"]] or "")
        email = str(row_data[COL_MAP["email"]] or "")

        # Find best WhatsApp number
        wa_number = ""
        wa_phone_display = ""
        if is_likely_cellphone(telefone1):
            wa_number = clean_phone_for_whatsapp(telefone1)
            wa_phone_display = telefone1
        elif is_likely_cellphone(telefone2):
            wa_number = clean_phone_for_whatsapp(telefone2)
            wa_phone_display = telefone2
        elif telefone1:
            # Try anyway — might be cellphone in non-standard format
            wa_number = clean_phone_for_whatsapp(telefone1)
            wa_phone_display = telefone1

        if not wa_number:
            stats["no_phone"] += 1
        else:
            stats["with_cellphone"] += 1

        # Build message for ALL leads (even without confirmed cellphone)
        msg = build_whatsapp_message(
            decisor=decisor,
            nome_fantasia=nome_fantasia,
            empresa=empresa,
            setor=setor,
            ufs=ufs,
            contratos=int(contratos) if contratos else 0,
        )
        stats["with_message"] += 1

        # Write to Leads sheet (always update — messages may have changed)
        if True:
            ws_leads.cell(row=row_idx, column=wa_num_col, value=wa_number or "").border = THIN_BORDER
            link_cell = ws_leads.cell(
                row=row_idx,
                column=wa_link_col,
                value=f"https://wa.me/{wa_number}" if wa_number else "",
            )
            link_cell.font = LINK_FONT if wa_number else Font(size=10)
            link_cell.border = THIN_BORDER
            if wa_number:
                link_cell.hyperlink = f"https://wa.me/{wa_number}"

            msg_cell = ws_leads.cell(row=row_idx, column=wa_msg_col, value=msg)
            msg_cell.font = MSG_FONT
            msg_cell.alignment = Alignment(wrap_text=True, vertical="top")
            msg_cell.border = THIN_BORDER

            ws_leads.cell(row=row_idx, column=wa_sent_col, value="").border = THIN_BORDER
            ws_leads.cell(row=row_idx, column=wa_resp_col, value="").border = THIN_BORDER

        # Collect for WhatsApp Outreach sheet (only leads with some phone)
        if wa_number:
            wa_leads.append({
                "num": row_data[COL_MAP["num"]],
                "empresa": empresa,
                "nome_fantasia": nome_fantasia,
                "cnpj": str(row_data[COL_MAP["cnpj"]] or ""),
                "decisor": decisor,
                "cargo": str(row_data[COL_MAP["cargo_decisor"]] or ""),
                "whatsapp": wa_number,
                "whatsapp_display": wa_phone_display,
                "is_cellphone": is_likely_cellphone(telefone1) or is_likely_cellphone(telefone2),
                "link": f"https://wa.me/{wa_number}",
                "email": email,
                "cidade": str(row_data[COL_MAP["cidade_sede"]] or ""),
                "fat_gov": row_data[COL_MAP["fat_gov_mensal"]] or 0,
                "contratos": contratos,
                "ufs": ufs,
                "message": msg,
            })

    # Sort: confirmed cellphones first, then by faturamento desc
    wa_leads.sort(key=lambda x: (not x["is_cellphone"], -(x["fat_gov"] or 0)))

    # Create/Replace WhatsApp Outreach sheet
    if "WhatsApp Outreach" in wb.sheetnames:
        del wb["WhatsApp Outreach"]

    ws_wa = wb.create_sheet("WhatsApp Outreach", 2)  # Position after Leads

    # Headers
    wa_headers = [
        "#", "Empresa", "Nome Fantasia", "CNPJ", "Decisor", "Cargo",
        "WhatsApp", "Link wa.me", "Email", "Cidade",
        "Fat. Gov Mensal", "Contratos", "UFs",
        "Mensagem WhatsApp",
        "Enviada?", "Data Envio", "Resposta", "Data Resposta",
        "Follow-up", "Status", "Notas",
    ]
    for col_idx, header in enumerate(wa_headers, 1):
        cell = ws_wa.cell(row=1, column=col_idx, value=header)
        cell.fill = WA_GREEN_FILL
        cell.font = WA_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Column widths
    widths = [5, 30, 25, 20, 20, 18, 16, 35, 30, 20, 15, 10, 15, 60, 12, 12, 25, 12, 12, 12, 30]
    for i, w in enumerate(widths, 1):
        ws_wa.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for idx, lead in enumerate(wa_leads, 1):
        row = idx + 1
        ws_wa.cell(row=row, column=1, value=idx).border = THIN_BORDER
        ws_wa.cell(row=row, column=2, value=lead["empresa"]).border = THIN_BORDER
        ws_wa.cell(row=row, column=3, value=lead["nome_fantasia"]).border = THIN_BORDER
        ws_wa.cell(row=row, column=4, value=lead["cnpj"]).border = THIN_BORDER
        ws_wa.cell(row=row, column=5, value=lead["decisor"]).border = THIN_BORDER
        ws_wa.cell(row=row, column=6, value=lead["cargo"]).border = THIN_BORDER
        ws_wa.cell(row=row, column=7, value=lead["whatsapp_display"]).border = THIN_BORDER

        link_cell = ws_wa.cell(row=row, column=8, value=lead["link"])
        link_cell.font = LINK_FONT
        link_cell.hyperlink = lead["link"]
        link_cell.border = THIN_BORDER

        ws_wa.cell(row=row, column=9, value=lead["email"]).border = THIN_BORDER
        ws_wa.cell(row=row, column=10, value=lead["cidade"]).border = THIN_BORDER

        fat_cell = ws_wa.cell(row=row, column=11, value=lead["fat_gov"])
        fat_cell.number_format = '#,##0.00'
        fat_cell.border = THIN_BORDER

        ws_wa.cell(row=row, column=12, value=lead["contratos"]).border = THIN_BORDER
        ws_wa.cell(row=row, column=13, value=lead["ufs"]).border = THIN_BORDER

        msg_cell = ws_wa.cell(row=row, column=14, value=lead["message"])
        msg_cell.font = MSG_FONT
        msg_cell.alignment = Alignment(wrap_text=True, vertical="top")
        msg_cell.border = THIN_BORDER

        # Empty tracking columns
        for col in range(15, 22):
            ws_wa.cell(row=row, column=col, value="").border = THIN_BORDER

        # Color confirmed cellphones green-ish
        if lead["is_cellphone"]:
            ws_wa.cell(row=row, column=7).fill = PatternFill(
                start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
            )

    # Freeze top row + auto-filter
    ws_wa.freeze_panes = "A2"
    ws_wa.auto_filter.ref = f"A1:{get_column_letter(len(wa_headers))}{len(wa_leads) + 1}"

    # Save
    wb.save(filepath)

    print(f"  Total leads: {stats['total']}")
    print(f"  With cellphone: {stats['with_cellphone']}")
    print(f"  Messages generated: {stats['with_message']}")
    print(f"  No phone: {stats['no_phone']}")
    print(f"  WhatsApp Outreach rows: {len(wa_leads)}")
    print(f"  Saved: {filepath}")

    return stats


def main():
    parser = argparse.ArgumentParser(description="Enrich Intel B2G spreadsheets with WhatsApp outreach")
    parser.add_argument("--file", help="Specific .xlsx file to process (in docs/intel-b2g/)")
    parser.add_argument("--dir", default=str(INTEL_DIR), help="Directory with .xlsx files")
    args = parser.parse_args()

    target_dir = Path(args.dir)
    if not target_dir.exists():
        print(f"ERROR: Directory not found: {target_dir}")
        sys.exit(1)

    if args.file:
        files = [target_dir / args.file]
        if not files[0].exists():
            print(f"ERROR: File not found: {files[0]}")
            sys.exit(1)
    else:
        files = sorted(target_dir.glob("leads-*.xlsx"))

    if not files:
        print("No .xlsx files found in", target_dir)
        sys.exit(1)

    print(f"Found {len(files)} file(s) to process")

    total_stats = {"total": 0, "with_cellphone": 0, "with_message": 0, "no_phone": 0}
    for f in files:
        stats = enrich_workbook(f)
        for k in total_stats:
            total_stats[k] += stats[k]

    print(f"\n{'='*60}")
    print("TOTALS")
    print(f"{'='*60}")
    print(f"  Files processed: {len(files)}")
    print(f"  Total leads: {total_stats['total']}")
    print(f"  With cellphone: {total_stats['with_cellphone']}")
    print(f"  Messages generated: {total_stats['with_message']}")
    print(f"  No phone: {total_stats['no_phone']}")


if __name__ == "__main__":
    main()
