#!/usr/bin/env python3
"""
Gerador de PDF executivo para Relatório B2G de Oportunidades.

Recebe JSON com dados coletados pelos agentes e gera PDF institucional
com análise estratégica por edital.

Design: Big Four / Management Consulting aesthetic (McKinsey, BCG, Deloitte).
Typography: Serif headings (Times) + sans-serif data (Helvetica).
Palette: Monochromatic (charcoal navy + bronze accent + neutral grays).

Usage:
    python scripts/generate-report-b2g.py --input data.json --output report.pdf
    python scripts/generate-report-b2g.py --input data.json  # output auto-named

Input JSON schema: see SCHEMA section below.
"""
from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.pdfgen import canvas as pdfgen_canvas
    from reportlab.platypus import (
        KeepTogether,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
except ImportError:
    print("ERROR: reportlab not installed. Run: pip install reportlab")
    sys.exit(1)

# ============================================================
# DESIGN TOKENS — Big Four Aesthetic
# ============================================================
# Principles: restraint, gravitas, whitespace, monocromia, open tables

# Palette — 3 colors + gray scale
INK = colors.HexColor("#1B2A3D")           # Charcoal navy — headings, emphasis
ACCENT = colors.HexColor("#8B7355")        # Warm bronze — subtle accents
SIGNAL_RED = colors.HexColor("#B5342A")    # Muted red — critical alerts only
SIGNAL_GREEN = colors.HexColor("#1B7A3D")  # Muted green — positive recommendations
SIGNAL_AMBER = colors.HexColor("#B8860B")  # Dark goldenrod — cautionary recommendations

TEXT_COLOR = colors.HexColor("#2D3748")    # Body text
TEXT_SECONDARY = colors.HexColor("#5A6577")  # Subtitles, labels
TEXT_MUTED = colors.HexColor("#8896A6")    # Footnotes, metadata
RULE_COLOR = colors.HexColor("#C8CDD3")   # Table inner rules (hairline)
RULE_HEAVY = colors.HexColor("#4A5568")    # Table top rule (heavy)
BG_SUBTLE = colors.HexColor("#F5F6F8")    # Rare subtle background

FOOTER_LINE1 = "Tiago Sasaki — Consultor de Inteligência em Licitações"
FOOTER_LINE2 = "Relatório confidencial preparado exclusivamente para o destinatário"

PAGE_WIDTH, PAGE_HEIGHT = A4
MARGIN = 2.2 * cm

ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Recommendation styling — color-coded for instant visual parsing
REC_STYLES = {
    "PARTICIPAR": {"color": SIGNAL_GREEN, "weight": "bold"},
    "AVALIAR": {"color": SIGNAL_AMBER, "weight": "bold"},
    "AVALIAR COM CAUTELA": {"color": SIGNAL_AMBER, "weight": "bold"},
    "NÃO RECOMENDADO": {"color": SIGNAL_RED, "weight": "bold"},
}

# Source confidence — textual, no emoji
SOURCE_LABELS = {
    "API": ("Confirmado", TEXT_COLOR),
    "CALCULATED": ("Calculado", TEXT_COLOR),
    "API_PARTIAL": ("Parcial", TEXT_SECONDARY),
    "ESTIMATED": ("Estimado", TEXT_SECONDARY),
    "API_FAILED": ("Indisponível", SIGNAL_RED),
    "UNAVAILABLE": ("N/D", TEXT_MUTED),
}


# ============================================================
# ACCENT RESTORATION (API data often lacks PT-BR diacritics)
# ============================================================
_ACCENT_MAP = {
    "construcao": "construção", "construcoes": "construções",
    "licitacao": "licitação", "licitacoes": "licitações",
    "contratacao": "contratação", "contratacoes": "contratações",
    "pavimentacao": "pavimentação", "habitacao": "habitação",
    "edificacao": "edificação", "edificacoes": "edificações",
    "ampliacao": "ampliação", "revitalizacao": "revitalização",
    "execucao": "execução", "manutencao": "manutenção",
    "atencao": "atenção", "educacao": "educação",
    "implantacao": "implantação", "instalacao": "instalação",
    "instalacoes": "instalações", "fundacao": "fundação",
    "recreacao": "recreação", "recuperacao": "recuperação",
    "reabilitacao": "reabilitação", "fiscalizacao": "fiscalização",
    "sinalizacao": "sinalização", "urbanizacao": "urbanização",
    "impermeabilizacao": "impermeabilização",
    "adequacao": "adequação", "adequacoes": "adequações",
    "operacao": "operação", "operacoes": "operações",
    "medicao": "medição", "medicoes": "medições",
    "informacao": "informação", "informacoes": "informações",
    "situacao": "situação", "avaliacao": "avaliação",
    "qualificacao": "qualificação", "habilitacao": "habilitação",
    "documentacao": "documentação", "certificacao": "certificação",
    "certificacoes": "certificações", "restricao": "restrição",
    "restricoes": "restrições", "sancao": "sanção", "sancoes": "sanções",
    "pregao": "pregão", "concorrencia": "concorrência",
    "eletronica": "eletrônica", "eletronicas": "eletrônicas",
    "eletrica": "elétrica", "eletricas": "elétricas",
    "tecnico": "técnico", "tecnica": "técnica",
    "tecnicas": "técnicas", "tecnicos": "técnicos",
    "economico": "econômico", "economica": "econômica",
    "municipio": "município", "municipios": "municípios",
    "orgao": "órgão", "orgaos": "órgãos",
    "publico": "público", "publica": "pública",
    "publicos": "públicos", "publicas": "públicas",
    "indice": "índice", "indices": "índices",
    "area": "área", "areas": "áreas",
    "agua": "água", "aguas": "águas",
    "analise": "análise", "analises": "análises",
    "registro": "registro", "minimo": "mínimo", "minima": "mínima",
    "maximo": "máximo", "maxima": "máxima",
    "necessario": "necessário", "necessaria": "necessária",
    "necessarios": "necessários", "necessarias": "necessárias",
    "especifico": "específico", "especifica": "específica",
    "especificos": "específicos", "historico": "histórico",
    "historica": "histórica", "obrigatorio": "obrigatório",
    "obrigatoria": "obrigatória", "provisorio": "provisório",
    "viavel": "viável", "responsavel": "responsável",
    "compativel": "compatível", "acessivel": "acessível",
    "gerenciavel": "gerenciável", "possivel": "possível",
    "impossivel": "impossível", "provavel": "provável",
    "disponivel": "disponível",
    "nao": "não", "ja": "já", "tambem": "também",
    "ate": "até", "apos": "após", "so": "só",
    "sera": "será", "serao": "serão",
    "voce": "você",
    "recomendacao": "recomendação",
    "preco": "preço", "precos": "preços",
    "orcamento": "orçamento", "orcamentos": "orçamentos",
    "orcado": "orçado", "servico": "serviço", "servicos": "serviços",
    "acervo": "acervo", "sessao": "sessão",
    "competicao": "competição", "competicoes": "competições",
    "reputacao": "reputação", "precificacao": "precificação",
    "associacao": "associação",
    "vigencia": "vigência", "exigencia": "exigência",
    "exigencias": "exigências", "frequencia": "frequência",
    "discrepancia": "discrepância",
    "pratica": "prática", "praticas": "práticas",
    "ceramico": "cerâmico", "ceramica": "cerâmica",
    "metalica": "metálica", "metalico": "metálico",
    "termoacustico": "termoacústico",
    "logistica": "logística", "basica": "básica",
    "estrategia": "estratégia",
    # E9 — termos faltantes (construção, licitação, jurídico)
    "demolicao": "demolição", "demolicoes": "demolições",
    "aprovacao": "aprovação", "aprovacoes": "aprovações",
    "orcamentario": "orçamentário", "orcamentaria": "orçamentária",
    "contribuicao": "contribuição", "contribuicoes": "contribuições",
    "resolucao": "resolução", "resolucoes": "resoluções",
    "suspensao": "suspensão", "suspensoes": "suspensões",
    "impugnacao": "impugnação", "impugnacoes": "impugnações",
    "rescisao": "rescisão", "rescisoes": "rescisões",
    "homologacao": "homologação", "homologacoes": "homologações",
    "adjudicacao": "adjudicação", "adjudicacoes": "adjudicações",
    "revogacao": "revogação", "revogacoes": "revogações",
    "anulacao": "anulação", "anulacoes": "anulações",
    "inspecao": "inspeção", "inspecoes": "inspeções",
    "isencao": "isenção", "isencoes": "isenções",
    "aquisicao": "aquisição", "aquisicoes": "aquisições",
    "consultorio": "consultório", "consultorios": "consultórios",
    "obrigacao": "obrigação", "obrigacoes": "obrigações",
    "proporcao": "proporção", "proporcoes": "proporções",
    "alteracao": "alteração", "alteracoes": "alterações",
    "cotacao": "cotação", "cotacoes": "cotações",
    "publicacao": "publicação", "publicacoes": "publicações",
    "participacao": "participação",
    "classificacao": "classificação", "classificacoes": "classificações",
    "dispensacao": "dispensação",
    "desclassificacao": "desclassificação",
    "improcedencia": "improcedência",
    "juridico": "jurídico", "juridica": "jurídica",
    "juridicos": "jurídicos", "juridicas": "jurídicas",
    "relatorio": "relatório", "relatorios": "relatórios",
    "valido": "válido", "valida": "válida",
    "solido": "sólido", "solida": "sólida",
    "unico": "único", "unica": "única",
    "unicos": "únicos", "unicas": "únicas",
    "inteligencia": "inteligência",
    "referencia": "referência", "referencias": "referências",
    "potencia": "potência", "potencial": "potencial",
    "suficiencia": "suficiência", "insuficiencia": "insuficiência",
    "eficiencia": "eficiência", "deficiencia": "deficiência",
    "presidencia": "presidência", "gerencia": "gerência",
    "transparencia": "transparência",
    "denuncia": "denúncia", "denuncias": "denúncias",
    "comercio": "comércio",
    "patrimonio": "patrimônio",
    "territorio": "território", "territorios": "territórios",
    "beneficio": "benefício", "beneficios": "benefícios",
    "edificio": "edifício", "edificios": "edifícios",
    "exercicio": "exercício", "exercicios": "exercícios",
    "previo": "prévio", "previa": "prévia",
    "obice": "óbice",
    "veiculos": "veículos", "veiculo": "veículo",
    "residuos": "resíduos",
    "conteudo": "conteúdo",
    "periodo": "período", "periodos": "períodos",
    "criterio": "critério", "criterios": "critérios",
    "equilibrio": "equilíbrio",
    "portfolio": "portfólio",
    "diagnostico": "diagnóstico", "diagnosticos": "diagnósticos",
    "cobertura": "cobertura",
    "mobilizacao": "mobilização",
    "regiao": "região", "regioes": "regiões",
}

_ACCENT_PATTERN = re.compile(
    r"\b(" + "|".join(re.escape(k) for k in sorted(_ACCENT_MAP.keys(), key=len, reverse=True)) + r")\b",
    re.IGNORECASE,
)


def _restore_accents(text: str) -> str:
    if not text:
        return text

    def _replace(m: re.Match) -> str:
        word = m.group(0)
        lower = word.lower()
        replacement = _ACCENT_MAP.get(lower, word)
        if word.isupper():
            return replacement.upper()
        if word[0].isupper():
            return replacement[0].upper() + replacement[1:]
        return replacement

    return _ACCENT_PATTERN.sub(_replace, text)


_PNCP_HYPHEN_LINK_RE = re.compile(
    r"https://pncp\.gov\.br/app/editais/(\d{14})-(\d{4})-(\d+)$"
)
_PNCP_SEARCH_LINK_RE = re.compile(
    r"https://pncp\.gov\.br/app/editais\?q="
)


# ============================================================
# CONSTANTS
# ============================================================

# Hard cap: only Top N editais get full detailed rendering; the rest get a condensed row
MAX_DETAILED_EDITAIS = 15
# Hard cap: overview table rows before truncating to Excel companion
MAX_OVERVIEW_ROWS = 30

# Canonical recommendation ordering — single source of truth
REC_ORDER = {"PARTICIPAR": 0, "AVALIAR COM CAUTELA": 1, "AVALIAR": 1, "NÃO RECOMENDADO": 2}

# Footnote for corrupt currency values
CURRENCY_CORRUPT_NOTE = "* Valor original não numérico no dado-fonte"

# ============================================================
# HELPERS
# ============================================================

def _normalize_recommendation(rec: str) -> str:
    rec = rec.strip().upper()
    rec = rec.replace("NAO RECOMENDADO", "NÃO RECOMENDADO")
    rec = rec.replace("NAO ", "NÃO ")
    if "PARTICIPAR" in rec:
        return "PARTICIPAR"
    if "CAUTELA" in rec or "AVALIAR" in rec:
        return "AVALIAR COM CAUTELA"
    if "NÃO" in rec or "RECOMENDADO" in rec:
        return "NÃO RECOMENDADO"
    return rec


def _summarize_discard_reasons(descartados: list[dict]) -> str:
    """Build a human-readable summary of why editais were discarded."""
    if not descartados:
        return ""
    reasons: dict[str, int] = {}
    for e in descartados:
        justif = e.get("justificativa", "").strip()
        if not justif:
            justif = "Sem aderência aos CNAEs da empresa"
        # Group similar justifications by first clause
        key = justif.split(".")[0].strip()
        if len(key) > 80:
            key = key[:77] + "..."
        reasons[key] = reasons.get(key, 0) + 1
    parts = [f"{reason} ({count})" if count > 1 else reason for reason, count in reasons.items()]
    return "; ".join(parts)


def _validate_json(data: dict) -> tuple[list[str], list[str]]:
    warnings = []
    errors = []
    if "empresa" not in data:
        warnings.append("Campo 'empresa' ausente")
    else:
        emp = data["empresa"]
        for field in ["cnpj", "razao_social"]:
            if not emp.get(field):
                warnings.append(f"empresa.{field} ausente")
    if "editais" not in data:
        warnings.append("Campo 'editais' ausente")
    for i, ed in enumerate(data.get("editais", [])):
        if not ed.get("objeto"):
            warnings.append(f"edital[{i}].objeto ausente")
        if not ed.get("orgao"):
            warnings.append(f"edital[{i}].orgao ausente")
        rec = (ed.get("recomendacao") or "").upper()
        status = ed.get("status_edital", "")
        if rec and rec != "DESCARTADO" and status != "ENCERRADO" and not ed.get("justificativa"):
            errors.append(
                f"edital[{i}].justificativa ausente — recomendação '{rec}' "
                f"para \"{(ed.get('objeto') or 'sem título')[:60]}\" não tem fundamentação"
            )
    if warnings:
        print(f"  Validação JSON: {len(warnings)} avisos")
        for w in warnings[:10]:
            print(f"  - {w}")
    if errors:
        print(f"\n  Validação JSON: {len(errors)} ERROS BLOQUEANTES")
        for e in errors:
            print(f"  - {e}")
    return warnings, errors


def _get_source_label(source: dict | str | None) -> tuple[str, Any]:
    if not source:
        return SOURCE_LABELS["UNAVAILABLE"]
    if isinstance(source, str):
        return SOURCE_LABELS.get(source, SOURCE_LABELS["UNAVAILABLE"])
    status = source.get("status", "UNAVAILABLE") if isinstance(source, dict) else "UNAVAILABLE"
    return SOURCE_LABELS.get(status, SOURCE_LABELS["UNAVAILABLE"])


def _fix_pncp_link(link: str | None) -> str:
    if not link:
        return ""
    link = str(link).strip()
    m = _PNCP_HYPHEN_LINK_RE.match(link)
    if m:
        cnpj, ano, seq = m.groups()
        return f"https://pncp.gov.br/app/editais/{cnpj}/{ano}/{seq}"
    if _PNCP_SEARCH_LINK_RE.match(link):
        return ""
    return link


def _s(value: Any, restore_accents: bool = True) -> str:
    """Sanitize text for PDF. restore_accents=False for structural fields (CNPJ, codes, links)."""
    if value is None:
        return ""
    text = ILLEGAL_CHARS_RE.sub(" ", str(value))
    if restore_accents:
        text = _restore_accents(text)
    return text


def _currency(value: Any, default: str = "N/I") -> str:
    if value is None:
        return default
    try:
        v = float(value)
    except (ValueError, TypeError):
        return "N/I*"  # corrupt value marker — see CURRENCY_CORRUPT_NOTE
    if v == 0:
        return "R$ 0,00"
    formatted = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def _currency_short(value: Any) -> str:
    if value is None:
        return "N/I"
    try:
        v = float(value)
    except (ValueError, TypeError):
        return "N/I"
    if v >= 1_000_000:
        return f"R$ {v / 1_000_000:,.1f}M".replace(",", "X").replace(".", ",").replace("X", ".")
    if v >= 1_000:
        return f"R$ {v / 1_000:,.0f}K".replace(",", "X").replace(".", ",").replace("X", ".")
    return _currency(v)


def _pct(value: float, decimals: int = 0) -> str:
    """Format a ratio (0.0-1.0) as Brazilian percentage: 0.125 → '12,5%'."""
    try:
        v = float(value) * 100
    except (ValueError, TypeError):
        return "N/I"
    if decimals == 0:
        return f"{v:.0f}%".replace(".", ",")
    return f"{v:.{decimals}f}%".replace(".", ",")


def _dec(value: float, decimals: int = 1) -> str:
    """Format a float in Brazilian decimal: 12.5 → '12,5'."""
    try:
        v = float(value)
    except (ValueError, TypeError):
        return "N/I"
    return f"{v:.{decimals}f}".replace(".", ",")


def _date(value: str | None) -> str:
    if not value:
        return "N/I"
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y", "%Y%m%d"):
        try:
            dt = datetime.strptime(text[:10], fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return text[:10]


def _today() -> str:
    return datetime.now(timezone.utc).strftime("%d/%m/%Y")


def _trunc(text: str, n: int = 100) -> str:
    text = _s(text)
    return text if len(text) <= n else text[: n - 3].rstrip() + "..."


def _safe_float(v: Any, d: float = 0.0) -> float:
    try:
        return float(v) if v is not None else d
    except (ValueError, TypeError):
        return d


def _safe_int(v: Any, d: int = 0) -> int:
    try:
        return int(v) if v is not None else d
    except (ValueError, TypeError):
        return d


def _format_dias_restantes(dias: Any) -> str:
    if dias is None or dias == "None":
        return "N/D"
    d = _safe_int(dias)
    if d < 0:
        return f"Encerrado ({abs(d)}d atrás)"
    if d == 0:
        return "Hoje"
    if d == 1:
        return "Amanhã"
    return f"{d} dias"


def _format_prazo_short(dias: Any) -> str:
    if dias is None or dias == "None":
        return "—"
    d = _safe_int(dias)
    if d < 0:
        return "Enc."
    if d == 0:
        return "Hoje"
    return f"{d}d"


def _fmt_pop(pop: int | float | None) -> str:
    """Format population with exact thousand-separators for cities <1M."""
    if pop is None:
        return "N/I"
    pop = int(pop)
    if pop >= 1_000_000:
        return f"{pop / 1_000_000:,.1f}M".replace(",", ".")
    else:
        return f"{pop:,}".replace(",", ".")  # exact with thousand separator


def _collapse_cnaes(cnaes: Any, max_show: int = 5) -> str:
    if not cnaes:
        return ""
    if isinstance(cnaes, str):
        parts = [c.strip() for c in cnaes.replace(";", ",").split(",") if c.strip()]
    elif isinstance(cnaes, list):
        parts = [str(c).strip() for c in cnaes if c]
    else:
        return str(cnaes)
    if len(parts) <= max_show:
        return ", ".join(parts)
    shown = ", ".join(parts[:max_show])
    return f"{shown} (e mais {len(parts) - max_show})"


# ============================================================
# THREE-RULE TABLE HELPER
# ============================================================

def _three_rule_table(rows: list, col_widths: list, repeat_rows: int = 1) -> Table:
    """Create a table with Big Four 'three-rule' styling.

    Heavy rule on top, hairline between rows, medium rule on bottom.
    No colored headers, no grid, no zebra striping.
    """
    t = Table(rows, colWidths=col_widths, repeatRows=repeat_rows)
    n = len(rows)
    style_cmds = [
        # Top heavy rule
        ("LINEABOVE", (0, 0), (-1, 0), 1.2, RULE_HEAVY),
        # Bottom of header
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, RULE_HEAVY),
        # Bottom of table
        ("LINEBELOW", (0, n - 1), (-1, n - 1), 0.8, RULE_COLOR),
        # Inner hairlines
        *[("LINEBELOW", (0, i), (-1, i), 0.3, RULE_COLOR) for i in range(1, n - 1)],
        # Alignment & padding
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]
    t.setStyle(TableStyle(style_cmds))
    return t


# ============================================================
# SECTION HEADING HELPER
# ============================================================

def _section_heading(title: str, styles: dict) -> list:
    """Create a minimal section heading: thin bronze rule + serif title."""
    avail = PAGE_WIDTH - 2 * MARGIN
    rule_t = Table([[""]],  colWidths=[avail], rowHeights=[1])
    rule_t.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (0, 0), 0.6, ACCENT),
        ("TOPPADDING", (0, 0), (0, 0), 0),
        ("BOTTOMPADDING", (0, 0), (0, 0), 0),
    ]))
    rule_t.keepWithNext = True
    h = Paragraph(title, styles["h1"])
    h.keepWithNext = True
    return [rule_t, Spacer(1, 2 * mm), h]


# ============================================================
# STYLES
# ============================================================

def _build_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    s: dict[str, ParagraphStyle] = {}

    # Cover
    s["cover_title"] = ParagraphStyle(
        "cover_title", parent=base["Normal"],
        fontName="Times-Bold", fontSize=26, textColor=INK,
        alignment=TA_LEFT, leading=32, spaceAfter=4 * mm,
    )
    s["cover_subtitle"] = ParagraphStyle(
        "cover_subtitle", parent=base["Normal"],
        fontName="Times-Roman", fontSize=14, textColor=TEXT_SECONDARY,
        alignment=TA_LEFT, spaceAfter=3 * mm, leading=18,
    )
    s["cover_info"] = ParagraphStyle(
        "cover_info", parent=base["Normal"],
        fontName="Helvetica", fontSize=9, textColor=TEXT_SECONDARY,
        alignment=TA_LEFT, leading=13, spaceAfter=1.5 * mm,
    )

    # Headings — serif for gravitas
    s["h1"] = ParagraphStyle(
        "h1_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=14, textColor=INK,
        spaceBefore=8 * mm, spaceAfter=4 * mm, leading=18,
    )
    s["h2"] = ParagraphStyle(
        "h2_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=11, textColor=INK,
        spaceBefore=5 * mm, spaceAfter=3 * mm, leading=14,
    )
    s["h3"] = ParagraphStyle(
        "h3_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=10, textColor=TEXT_COLOR,
        spaceBefore=3 * mm, spaceAfter=2 * mm, leading=13,
    )

    # Body — serif, justified
    s["body"] = ParagraphStyle(
        "body_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=TEXT_COLOR,
        alignment=TA_JUSTIFY, leading=14, spaceAfter=2 * mm,
    )
    s["body_small"] = ParagraphStyle(
        "body_small_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=9, textColor=TEXT_SECONDARY,
        leading=12, spaceAfter=1.5 * mm,
    )
    s["bullet"] = ParagraphStyle(
        "bullet_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=TEXT_COLOR,
        leading=14, leftIndent=10, spaceAfter=1.5 * mm,
    )
    s["caption"] = ParagraphStyle(
        "caption_r", parent=base["Normal"],
        fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED,
        leading=9,
    )

    # Metrics
    s["metric_value"] = ParagraphStyle(
        "mv_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=18, textColor=INK,
        alignment=TA_CENTER, leading=22,
    )
    s["metric_label"] = ParagraphStyle(
        "ml_r", parent=base["Normal"],
        fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED,
        alignment=TA_CENTER, leading=9,
    )

    # Table cells — sans-serif for data clarity
    for name, align in [("cell", TA_LEFT), ("cell_center", TA_CENTER), ("cell_right", TA_RIGHT)]:
        s[name] = ParagraphStyle(
            f"{name}_r", parent=base["Normal"],
            fontName="Helvetica", fontSize=8, textColor=TEXT_COLOR,
            leading=10, alignment=align,
        )
    s["cell_header"] = ParagraphStyle(
        "ch_r", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=INK,
        leading=10, alignment=TA_LEFT,
    )
    s["cell_header_center"] = ParagraphStyle(
        "chc_r", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=INK,
        leading=10, alignment=TA_CENTER,
    )
    s["cell_header_right"] = ParagraphStyle(
        "chr_r", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=INK,
        leading=10, alignment=TA_RIGHT,
    )

    return s


# ============================================================
# FOOTER with "Página X de Y"
# ============================================================

class _NumberedCanvas(pdfgen_canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states: list = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_page_number(total)
            pdfgen_canvas.Canvas.showPage(self)
        pdfgen_canvas.Canvas.save(self)

    def _draw_page_number(self, total: int):
        self.saveState()
        self.setFont("Helvetica", 7)
        self.setFillColor(TEXT_MUTED)
        self.drawRightString(
            PAGE_WIDTH - MARGIN,
            MARGIN - 12 * mm,
            f"Página {self._pageNumber} de {total}",
        )
        self.restoreState()


def _draw_footer(canvas, doc):
    canvas.saveState()
    y = MARGIN - 10 * mm

    # Thin rule
    canvas.setStrokeColor(RULE_COLOR)
    canvas.setLineWidth(0.4)
    canvas.line(MARGIN, y + 4 * mm, PAGE_WIDTH - MARGIN, y + 4 * mm)

    # Institutional attribution
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(TEXT_MUTED)
    canvas.drawCentredString(PAGE_WIDTH / 2, y + 1.5 * mm, FOOTER_LINE1)
    canvas.drawCentredString(PAGE_WIDTH / 2, y - 2 * mm, FOOTER_LINE2)

    canvas.restoreState()


# ============================================================
# PAGE BUILDERS
# ============================================================

def _metric_cell(value: str, label: str, styles: dict) -> Table:
    inner = Table(
        [[Paragraph(value, styles["metric_value"])],
         [Paragraph(label, styles["metric_label"])]],
        colWidths=["*"],
    )
    inner.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    return inner


def _build_cover(data: dict, styles: dict, gen_date: str) -> list:
    el = []
    empresa = data.get("empresa", {})

    # Generous top whitespace
    el.append(Spacer(1, 80 * mm))

    # Short bronze rule (40mm) — understated elegance
    avail = PAGE_WIDTH - 2 * MARGIN
    rule_t = Table([["", ""]], colWidths=[40 * mm, avail - 40 * mm])
    rule_t.setStyle(TableStyle([("LINEBELOW", (0, 0), (0, 0), 0.8, ACCENT)]))
    el.append(rule_t)
    el.append(Spacer(1, 6 * mm))

    # Title — left-aligned serif
    el.append(Paragraph(
        "Relatório Executivo de<br/>Oportunidades em Licitações",
        styles["cover_title"],
    ))

    nome = _s(empresa.get("nome_fantasia") or empresa.get("razao_social", ""))
    if nome:
        el.append(Paragraph(nome, styles["cover_subtitle"]))

    el.append(Spacer(1, 12 * mm))

    # Metadata block — clean, sans-serif, left-aligned
    cnpj = _s(empresa.get("cnpj", ""), restore_accents=False)
    setor = _s(data.get("setor", ""))
    uf_sede = _s(empresa.get("uf_sede", ""), restore_accents=False)
    cidade = _s(empresa.get("cidade_sede", ""))

    meta_lines = []
    if cnpj:
        meta_lines.append(f"CNPJ {cnpj}")
    if setor:
        meta_lines.append(setor)
    if cidade and uf_sede:
        meta_lines.append(f"{cidade}, {uf_sede}")
    elif uf_sede:
        meta_lines.append(uf_sede)
    meta_lines.append(gen_date)

    for line in meta_lines:
        el.append(Paragraph(line, styles["cover_info"]))

    el.append(Spacer(1, 30 * mm))

    # Consultant attribution — bottom
    el.append(Paragraph(
        "<b>Tiago Sasaki</b><br/>"
        "Consultor de Inteligência em Licitações<br/>"
        "(48) 9 8834-4559",
        ParagraphStyle(
            "cover_attr", fontName="Helvetica", fontSize=9,
            textColor=TEXT_SECONDARY, alignment=TA_LEFT, leading=13,
        ),
    ))

    el.append(PageBreak())
    return el


def _build_resumo_decisorio(data: dict, styles: dict) -> list:
    """CRÍTICA 10B: Resumo Decisório — the 'nectar' of the report on page 2.

    3-5 short paragraphs that give the reader immediate decision-making value
    without reading the full report. Contains:
    1. How many opportunities, total value at stake
    2. Top 3 recommended opportunities with municipality, object, ROI
    3. Critical alerts (vetos, tight deadlines, fiscal risks)
    4. Final verdict with total estimated ROI
    """
    el: list = []
    editais = data.get("editais", [])

    if not editais:
        return el

    # Bronze accent rule
    avail = PAGE_WIDTH - 2 * MARGIN
    rule_t = Table([["", ""]], colWidths=[40 * mm, avail - 40 * mm])
    rule_t.setStyle(TableStyle([("LINEBELOW", (0, 0), (0, 0), 0.8, ACCENT)]))
    el.append(rule_t)
    el.append(Spacer(1, 4 * mm))

    el.append(Paragraph("O Que Realmente Importa", styles["h1"]))
    el.append(Spacer(1, 3 * mm))

    # Classify editais
    participar = [e for e in editais if (e.get("recomendacao") or "").upper() == "PARTICIPAR"]
    avaliar = [e for e in editais if "AVALIAR" in (e.get("recomendacao") or "").upper()]
    nao_rec = [e for e in editais if (e.get("recomendacao") or "").upper() == "NÃO RECOMENDADO"]
    vetados = [e for e in editais if (e.get("risk_score") or {}).get("vetoed")]

    total_valor = sum(_safe_float_gen(e.get("valor_estimado")) for e in editais)
    valor_participar = sum(_safe_float_gen(e.get("valor_estimado")) for e in participar)

    # Paragraph 1: Overview
    el.append(Paragraph(
        f"Foram identificadas <b>{len(editais)} oportunidades</b> em licitações abertas, "
        f"totalizando <b>{_currency_short(total_valor)}</b> em valor estimado. "
        f"Destas, <b>{len(participar)}</b> são recomendadas para participação imediata"
        f"{f' ({_currency_short(valor_participar)} em valor)' if valor_participar > 0 else ''}, "
        f"<b>{len(avaliar)}</b> merecem avaliação com cautela"
        f"{f' e <b>{len(vetados)}</b> foram eliminadas por impedimento legal' if vetados else ''}"
        f"{f', e <b>{len(nao_rec)}</b> não são recomendadas' if nao_rec else ''}.",
        styles["body"],
    ))
    el.append(Spacer(1, 2 * mm))

    # Paragraph 2: Top 3 recommended (by ROI)
    top_eds = sorted(
        participar,
        key=lambda e: (e.get("roi_potential") or {}).get("roi_max", 0),
        reverse=True,
    )[:3]

    if top_eds:
        lines = []
        for idx, e in enumerate(top_eds, 1):
            mun = _s(e.get("municipio", ""))
            uf = _s(e.get("uf", ""))
            obj_text = _trunc(_s(e.get("objeto", "")), 120)
            roi = e.get("roi_potential", {})
            roi_max = roi.get("roi_max", 0)
            prob = (e.get("win_probability") or {}).get("probability", 0)
            valor_e = _safe_float_gen(e.get("valor_estimado"))
            loc = f"{mun}/{uf}" if mun and uf else (mun or uf or "Local N/I")
            lines.append(
                f"<b>{idx}.</b> <b>{loc}</b> — {obj_text} "
                f"({_currency_short(valor_e)}, probabilidade {_pct(prob)}, "
                f"retorno potencial até {_currency_short(max(roi_max, 0))})"
            )
        el.append(Paragraph(
            "<b>Destaques para ação imediata:</b><br/>" + "<br/>".join(lines),
            styles["body"],
        ))
        el.append(Spacer(1, 2 * mm))

    # Paragraph 3: Critical alerts
    alerts = []
    # Veto alerts
    for e in vetados:
        reasons = (e.get("risk_score") or {}).get("veto_reasons", [])
        if reasons:
            alerts.append(f"ELIMINADO: {_trunc(_s(e.get('objeto', '')), 120)} — {reasons[0]}")
    # Tight deadline alerts
    for e in participar + avaliar:
        dias = e.get("dias_restantes")
        if dias is not None and dias <= 7:
            alerts.append(
                f"URGENTE: {_trunc(_s(e.get('objeto', '')), 120)} encerra em {dias} dia(s)"
            )
    # Fiscal risk alerts
    for e in participar + avaliar:
        fiscal = (e.get("risk_score") or {}).get("fiscal_risk", {})
        if isinstance(fiscal, dict) and fiscal.get("nivel") == "ALTO":
            fiscal_alerts = fiscal.get("alertas", [])
            if fiscal_alerts:
                alerts.append(
                    f"RISCO FISCAL: {_trunc(_s(e.get('objeto', '')), 120)} — {fiscal_alerts[0]}"
                )
    # Sanctions inconclusive alert
    emp_sancoes = data.get("empresa", {}).get("sancoes", {})
    if emp_sancoes.get("inconclusive"):
        alerts.append(
            "Situação de sanções não confirmada — verificar Portal da Transparência antes de submeter propostas"
        )

    if alerts:
        alert_text = "<br/>".join(f"• {a}" for a in alerts[:5])
        el.append(Paragraph(
            f"<b>Alertas críticos:</b><br/>{alert_text}",
            ParagraphStyle(
                "alert_body", parent=styles["body"],
                textColor=SIGNAL_RED, fontSize=9, leading=13,
            ),
        ))
        el.append(Spacer(1, 2 * mm))

    # Paragraph 4: Verdict
    total_roi_max = sum(
        max((e.get("roi_potential") or {}).get("roi_max", 0), 0)
        for e in participar
    )
    acervo_note = ""
    acervo_eds = [e for e in editais if (e.get("risk_score") or {}).get("acervo_confirmado") is False
                  and (e.get("recomendacao") or "").upper() in ("PARTICIPAR", "AVALIAR COM CAUTELA")]
    if acervo_eds:
        acervo_note = (
            f" <b>Atenção:</b> {len(acervo_eds)} edital(is) dependem de verificação prévia "
            f"de atestados técnicos compatíveis com o objeto licitado."
        )

    if participar:
        el.append(Paragraph(
            f"<b>Veredicto:</b> Recomendamos priorizar <b>{len(participar)} edital(is)</b> "
            f"com retorno potencial agregado de até <b>{_currency_short(max(total_roi_max, 0))}</b>. "
            f"O restante do relatório detalha a análise de cada oportunidade, "
            f"inteligência competitiva e plano de ação.{acervo_note}",
            styles["body"],
        ))
    else:
        el.append(Paragraph(
            "<b>Veredicto:</b> Nenhuma oportunidade atende aos critérios mínimos para "
            "participação imediata. As oportunidades listadas como AVALIAR COM CAUTELA "
            "podem ser viáveis mediante verificação adicional de requisitos específicos.",
            styles["body"],
        ))

    el.append(Spacer(1, 4 * mm))
    # Light separator
    sep = Table([["", ""]], colWidths=[avail * 0.3, avail * 0.7])
    sep.setStyle(TableStyle([("LINEBELOW", (0, 0), (0, 0), 0.4, RULE_COLOR)]))
    el.append(sep)
    el.append(Spacer(1, 4 * mm))

    return el


def _safe_float_gen(v: Any) -> float:
    """Safe float conversion for generator expressions."""
    if v is None:
        return 0.0
    try:
        if isinstance(v, str):
            v = v.replace(",", ".")
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _build_exclusive_intelligence(data: dict, styles: dict, sec: dict) -> list:
    """Build 'Inteligência Exclusiva' — 1-page executive summary of what PNCP alone can't provide.

    Four differentials: incumbency mapping, calibrated viability, strategic acervo, regional clusters.
    """
    editais = data.get("editais", [])
    if not editais:
        return []

    el = []
    num = sec["next"]()
    el.extend(_section_heading(f"{num}. Inteligência Exclusiva", styles))

    el.append(Paragraph(
        "Este relatório vai além da listagem de editais. Os quatro blocos abaixo "
        "representam análises que exigem cruzamento de dados históricos, perfil da empresa "
        "e modelagem competitiva — informações que não estão disponíveis em consultas individuais aos portais públicos.",
        styles["body"],
    ))
    el.append(Spacer(1, 5 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN
    empresa = data.get("empresa", {})

    # --- 1. Incumbency Intelligence ---
    el.append(Paragraph("<b>1. Mapeamento de Incumbência</b>", styles["h3"]))
    incumbent_insights = []
    for ed in editais:
        idx = ed.get("_display_idx", 0)
        ci = ed.get("competitive_intel", [])
        wp = ed.get("win_probability", {})
        if not ci and not wp:
            continue
        orgao = _s(ed.get("orgao", ""))[:120]
        hhi = wp.get("hhi", 0)
        top_share = wp.get("top_supplier_share", 0)
        unique = wp.get("n_unique_suppliers", wp.get("unique_suppliers", 0))
        prob = wp.get("probability", 0)

        if top_share > 0.60:
            insight = f"Incumbente dominante (>{_pct(top_share)} do mercado) — entrada difícil"
        elif top_share > 0.40:
            insight = f"Incumbente forte ({_pct(top_share)}) — requer diferenciação"
        elif unique >= 6:
            insight = f"Mercado fragmentado ({unique} fornecedores) — oportunidade aberta"
        elif unique >= 2:
            insight = f"Competição moderada ({unique} fornecedores, HHI {_dec(hhi, 2)})"
        else:
            insight = f"Sem histórico de fornecedores — mercado inexplorado"

        incumbent_insights.append((idx, orgao, insight, prob))

    if incumbent_insights:
        inc_header = [
            Paragraph("#", styles["cell_header_center"]),
            Paragraph("Órgão", styles["cell_header"]),
            Paragraph("Dinâmica Competitiva", styles["cell_header"]),
            Paragraph("Prob.", styles["cell_header_center"]),
        ]
        inc_rows = [inc_header]
        for idx, orgao, insight, prob in incumbent_insights[:8]:
            prob_color = SIGNAL_GREEN if prob >= 0.20 else (SIGNAL_AMBER if prob >= 0.10 else SIGNAL_RED)
            inc_rows.append([
                Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
                Paragraph(orgao, styles["cell"]),
                Paragraph(insight, styles["cell"]),
                Paragraph(
                    f"<b>{_pct(prob)}</b>",
                    ParagraphStyle(f"inc_p_{idx}", parent=styles["cell_center"],
                                   fontName="Helvetica-Bold", textColor=prob_color),
                ),
            ])
        t = _three_rule_table(inc_rows, [avail * 0.06, avail * 0.30, avail * 0.50, avail * 0.14])
        el.append(t)
    else:
        el.append(Paragraph(
            "Dados de incumbência não disponíveis para os editais analisados.",
            styles["body_small"],
        ))
    el.append(Spacer(1, 5 * mm))

    # --- 2. Calibrated Viability ---
    el.append(Paragraph("<b>2. Viabilidade Calibrada ao Perfil</b>", styles["h3"]))

    maturity = data.get("maturity_profile") or empresa.get("maturity_profile", {})
    profile_name = maturity.get("profile", "N/I") if maturity else "N/I"
    profile_labels = {
        "ENTRANTE": "Entrante (0-2 contratos federais)",
        "REGIONAL": "Regional (3-10 contratos, até 3 UFs)",
        "ESTABELECIDO": "Estabelecido (10+ contratos ou 4+ UFs)",
    }
    el.append(Paragraph(
        f"Perfil de maturidade: <b>{profile_labels.get(profile_name, profile_name)}</b> — "
        f"os pesos de viabilidade foram ajustados para refletir este perfil.",
        styles["body"],
    ))

    # Viability ranking
    scored = [(ed.get("_display_idx", i), ed) for i, ed in enumerate(editais, 1)
              if isinstance(ed.get("risk_score"), dict) and ed["risk_score"].get("total", 0) > 0]
    scored.sort(key=lambda x: x[1]["risk_score"]["total"], reverse=True)

    if scored:
        via_header = [
            Paragraph("#", styles["cell_header_center"]),
            Paragraph("Edital", styles["cell_header"]),
            Paragraph("Score", styles["cell_header_center"]),
            Paragraph("Fator Decisivo", styles["cell_header"]),
        ]
        via_rows = [via_header]
        for idx, ed in scored[:6]:
            risk = ed["risk_score"]
            score = _safe_int(risk.get("total"))
            # Find weakest and strongest factor
            components = {
                "Habilitação": risk.get("habilitacao", 50),
                "Financeiro": risk.get("financeiro", 50),
                "Geográfico": risk.get("geografico", 50),
                "Prazo": risk.get("prazo", 50),
                "Competitivo": risk.get("competitivo", 50),
            }
            weakest = min(components, key=lambda k: _safe_int(components[k]))
            strongest = max(components, key=lambda k: _safe_int(components[k]))
            factor = f"Forte: {strongest} ({_safe_int(components[strongest])}) | Fraco: {weakest} ({_safe_int(components[weakest])})"

            score_color = SIGNAL_GREEN if score >= 60 else (SIGNAL_AMBER if score >= 30 else SIGNAL_RED)
            obj_text = _trunc(_s(ed.get("objeto", "")), 150)
            via_link = _fix_pncp_link(ed.get("link", ""))
            if via_link and via_link.startswith("http"):
                obj_text = f'<a href="{via_link}" color="{INK.hexval()}">{obj_text}</a>'
            via_rows.append([
                Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
                Paragraph(obj_text, styles["cell"]),
                Paragraph(
                    f"<b>{score}/100</b>",
                    ParagraphStyle(f"via_s_{idx}", parent=styles["cell_center"],
                                   fontName="Helvetica-Bold", textColor=score_color),
                ),
                Paragraph(factor, styles["cell"]),
            ])
        t = _three_rule_table(via_rows, [avail * 0.06, avail * 0.38, avail * 0.12, avail * 0.44])
        el.append(t)
    el.append(Spacer(1, 5 * mm))

    # --- 3. Strategic Acervo Sequence ---
    el.append(Paragraph("<b>3. Sequência Estratégica de Acervo</b>", styles["h3"]))

    acervo_items = []
    for ed in editais:
        idx = ed.get("_display_idx", 0)
        roi = ed.get("roi_potential", {})
        cat = ed.get("strategic_category", "")
        if roi.get("strategic_reclassification") == "INVESTIMENTO_ESTRATEGICO_ACERVO" or cat == "INVESTIMENTO":
            acervo_items.append((idx, ed))

    participar_items = [(ed.get("_display_idx", i), ed) for i, ed in enumerate(editais, 1)
                        if _normalize_recommendation(_s(ed.get("recomendacao", ""))) == "PARTICIPAR"]

    if acervo_items:
        el.append(Paragraph(
            f"<b>{len(acervo_items)}</b> edital(is) classificado(s) como <b>Investimento Estratégico em Acervo</b> — "
            f"o retorno imediato é marginal, mas a execução constrói atestados e relacionamento "
            f"que desbloqueiam mercados futuros de maior valor.",
            styles["body"],
        ))
        for idx, ed in acervo_items:
            obj = _trunc(_s(ed.get("objeto", "")), 150)
            valor = _currency_short(ed.get("valor_estimado"))
            orgao = _s(ed.get("orgao", ""))[:120]
            el.append(Paragraph(
                f"  <b>{idx}.</b> {obj} ({orgao}) — {valor}",
                styles["body_small"],
            ))
    elif participar_items:
        el.append(Paragraph(
            "Nenhum edital classificado como investimento em acervo. "
            f"Os {len(participar_items)} edital(is) recomendados para participação "
            f"contribuem para o acervo técnico de forma orgânica.",
            styles["body"],
        ))
    else:
        el.append(Paragraph(
            "Nenhuma oportunidade estratégica de construção de acervo identificada nesta janela.",
            styles["body"],
        ))
    el.append(Spacer(1, 5 * mm))

    # --- 4. Regional Clusters ---
    el.append(Paragraph("<b>4. Clusters Geográficos</b>", styles["h3"]))
    clusters_data = data.get("regional_clusters", {})
    clusters = clusters_data.get("clusters", [])
    if clusters:
        el.append(Paragraph(
            f"<b>{len(clusters)}</b> cluster(s) de mobilização compartilhada identificado(s). "
            f"Uma única base operacional pode servir múltiplos editais no mesmo raio.",
            styles["body"],
        ))
        for cl in clusters[:4]:
            center = f"{_s(cl.get('center_municipio', ''))}/{_s(cl.get('center_uf', ''))}"
            n = cl.get("n_editais", 0)
            valor = _currency_short(cl.get("total_valor", 0))
            overlap = "prazos sobrepostos" if cl.get("timeline_overlap") else "prazos independentes"
            el.append(Paragraph(
                f"  — <b>{center}:</b> {n} editais, {valor}, {overlap}",
                styles["body_small"],
            ))
    else:
        el.append(Paragraph(
            "Editais dispersos geograficamente — sem oportunidade de mobilização compartilhada.",
            styles["body"],
        ))

    el.append(Spacer(1, 8 * mm))
    el.append(PageBreak())
    return el


def _build_decision_table(data: dict, styles: dict, sec: dict) -> list:
    """Build 'Decisão em 30 Segundos' — clean three-rule summary table."""
    el = []
    editais = data.get("editais", [])
    if not editais:
        return el

    num = sec["next"]()
    el.extend(_section_heading(f"{num}. Decisão em 30 Segundos", styles))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN
    col_w = [avail * 0.05, avail * 0.28, avail * 0.11, avail * 0.08, avail * 0.20, avail * 0.28]

    # Header
    header = [
        Paragraph("#", styles["cell_header_center"]),
        Paragraph("Objeto", styles["cell_header"]),
        Paragraph("Valor", styles["cell_header_right"]),
        Paragraph("Prazo", styles["cell_header_center"]),
        Paragraph("Recomendação", styles["cell_header"]),
        Paragraph("Diferencial Estratégico", styles["cell_header"]),
    ]
    rows = [header]

    # Sort: PARTICIPAR first, then AVALIAR, then NÃO RECOMENDADO; within each by score desc
    editais_sorted = sorted(
        editais,
        key=lambda e: (
            REC_ORDER.get(_normalize_recommendation(_s(e.get("recomendacao", ""))), 9),
            -(e.get("risk_score", {}).get("total", 0)),
        ),
    )
    # Cap to 30 rows (rest goes to Excel)
    editais_display = editais_sorted[:MAX_OVERVIEW_ROWS]

    for table_row, ed in enumerate(editais_display, 1):
        idx = ed.get("_display_idx", table_row)
        rec = _normalize_recommendation(_s(ed.get("recomendacao", "")))
        risk = ed.get("risk_score", {}) or {}
        vetoed = risk.get("vetoed", False)

        # CRÍTICA 2: Vetoed editais get distinct "ELIMINATÓRIO" label
        if vetoed:
            rec = "NÃO RECOMENDADO"
            rec_display = "ELIMINATÓRIO"
            rec_color = colors.HexColor("#7A1A1A")  # Darker red for elimination
        else:
            rec_style_info = REC_STYLES.get(rec, REC_STYLES["NÃO RECOMENDADO"])
            rec_color = rec_style_info["color"]
            rec_display = rec

        objeto = _trunc(_s(ed.get("objeto", "")), 150)
        link = _fix_pncp_link(ed.get("link", ""))
        if link and link.startswith("http"):
            objeto = f'<a href="{link}" color="{INK.hexval()}">{objeto}</a>'
        valor = _currency_short(ed.get("valor_estimado"))
        prazo = _format_prazo_short(ed.get("dias_restantes"))

        rec_ps = ParagraphStyle(
            f"drec_{idx}", fontName="Helvetica-Bold", fontSize=7.5,
            textColor=rec_color, alignment=TA_LEFT, leading=10,
            wordWrap="CJK",
        )

        # Build strategic differential insight
        diff_parts = []
        wp = ed.get("win_probability", {})
        roi = ed.get("roi_potential", {})
        cat = ed.get("strategic_category", "")

        # Show veto reason as differential for eliminated editais
        if vetoed:
            veto_reasons = risk.get("veto_reasons", [])
            if veto_reasons:
                diff_parts.append(_trunc(veto_reasons[0], 60))
            else:
                diff_parts.append("Impedimento legal")
        else:
            if isinstance(wp, dict) and wp.get("top_supplier_share", 0) < 0.20 and wp.get("n_unique_suppliers", wp.get("unique_suppliers", 0)) >= 2:
                diff_parts.append("Sem incumbente dominante")
            elif isinstance(wp, dict) and wp.get("top_supplier_share", 0) > 0.60:
                diff_parts.append("Incumbente forte")

            if isinstance(risk, dict) and risk.get("total", 0) >= 60:
                diff_parts.append(f"Viab. {risk['total']}/100")
            elif isinstance(risk, dict) and risk.get("total", 0) > 0:
                diff_parts.append(f"Viab. {risk['total']}/100")

            if roi.get("strategic_reclassification") == "INVESTIMENTO_ESTRATEGICO_ACERVO" or cat == "INVESTIMENTO":
                diff_parts.append("Acervo estratégico")
            elif cat == "QUICK_WIN":
                diff_parts.append("Quick Win")

        differential = " · ".join(diff_parts) if diff_parts else "—"

        rows.append([
            Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
            Paragraph(objeto, styles["cell"]),
            Paragraph(valor, styles["cell_right"]),
            Paragraph(prazo, styles["cell_center"]),
            Paragraph(rec_display, rec_ps),
            Paragraph(differential, ParagraphStyle(
                f"ddiff_{idx}", parent=styles["cell"],
                fontName="Helvetica", fontSize=7, textColor=TEXT_SECONDARY,
            )),
        ])

    t = _three_rule_table(rows, col_w)
    el.append(t)

    # Justifications below the table — footnotes aligned to the DISPLAYED rows.
    # CRITICAL: must iterate editais_display (sorted+capped), not editais (original),
    # so that footnote #N corresponds to table row #N.
    has_justif = False
    for table_row, ed in enumerate(editais_display, 1):
        idx = ed.get("_display_idx", table_row)
        justif = _s(ed.get("justificativa", ""))
        if justif:
            if not has_justif:
                el.append(Spacer(1, 4 * mm))
                has_justif = True
            el.append(Paragraph(
                f"<b>{idx}.</b> {_trunc(justif, 280)}",
                styles["caption"],
            ))

    el.append(Spacer(1, 6 * mm))
    return el


def _build_viability_text(risk: dict, styles: dict, _state: dict | None = None) -> list:
    """Build viability indicator with score decomposition using sector-specific weights."""
    score = _safe_int(risk.get("total") if isinstance(risk, dict) else risk)
    if score <= 0:
        return []

    if score >= 60:
        qualif = "Viabilidade Alta"
    elif score >= 30:
        qualif = "Viabilidade Moderada"
    else:
        qualif = "Viabilidade Baixa"

    el = [Paragraph(
        f"<b>Índice de Viabilidade:</b> {score}/100 — {qualif}",
        styles["body"],
    )]

    # Score decomposition — use actual weights from risk dict (sector-specific)
    if isinstance(risk, dict):
        weights = risk.get("weights", {})
        label_map = {
            "habilitacao": "Habilitação",
            "financeiro": "Financeiro",
            "geografico": "Geográfico",
            "prazo": "Prazo",
            "competitivo": "Competitivo",
        }
        components = []
        for key, label in label_map.items():
            val = risk.get(key)
            peso = weights.get(key[0:3] if key == "habilitacao" else key[:3],
                              weights.get(key))
            # Map full key names to weight keys
            weight_key_map = {
                "habilitacao": "hab", "financeiro": "fin",
                "geografico": "geo", "prazo": "prazo", "competitivo": "comp",
            }
            peso = weights.get(weight_key_map.get(key, key))
            if val is not None and peso is not None:
                components.append(f"{label}: {_safe_int(val)} (peso {_pct(peso)})")
            elif val is not None:
                components.append(f"{label}: {_safe_int(val)}")

        if components:
            el.append(Paragraph(
                f"Composição: {' | '.join(components)}",
                styles["caption"],
            ))

    # E8: Maturity adjustment display
    if isinstance(risk, dict) and risk.get("maturity_adjustment"):
        adj = risk["maturity_adjustment"]
        adj_parts = []
        for comp_key, delta in adj.items():
            if not isinstance(delta, (int, float)):
                continue
            if delta != 0:
                sign = "+" if delta > 0 else ""
                comp_label = {"hab": "Habilitação", "fin": "Financeiro", "geo": "Geográfico",
                              "prazo": "Prazo", "comp": "Competitivo"}.get(comp_key, comp_key)
                adj_parts.append(f"{comp_label} {sign}{delta}")
        if adj_parts:
            profile = risk.get("maturity_profile", "")
            el.append(Paragraph(
                f"<i>Ajuste por perfil de maturidade ({profile}): {', '.join(adj_parts)}</i>",
                styles["caption"],
            ))

    if _state is not None and not _state.get("viability_shown"):
        # Dynamic explanation based on weights
        if isinstance(risk, dict) and risk.get("weights"):
            w = risk["weights"]
            parts = []
            name_map = {"hab": "habilitação", "fin": "valor vs. capacidade",
                        "geo": "proximidade geográfica", "prazo": "prazo",
                        "comp": "competitividade"}
            for k in sorted(w, key=lambda x: w[x], reverse=True):
                parts.append(f"{name_map.get(k, k)} ({_pct(w[k])})")
            el.append(Paragraph(
                f"Pesos calibrados para o setor da empresa: {', '.join(parts)}.",
                styles["caption"],
            ))
        else:
            el.append(Paragraph(
                "Índice calculado com base em habilitação, prazo, "
                "valor vs. capacidade, proximidade geográfica e competitividade.",
                styles["caption"],
            ))
        _state["viability_shown"] = True

    return el


def _build_chronogram_table(cronograma: list, styles: dict) -> list:
    if not cronograma:
        return []

    el = []
    el.append(Paragraph("Cronograma Reverso", styles["h3"]))

    avail = PAGE_WIDTH - 2 * MARGIN
    header = [
        Paragraph("Data", styles["cell_header"]),
        Paragraph("Marco", styles["cell_header"]),
        Paragraph("Status", styles["cell_header"]),
    ]
    rows = [header]

    for item in cronograma:
        data_str = _date(item.get("data", ""))
        marco = _s(item.get("marco", ""))
        status = _s(item.get("status", ""))

        # Status color — muted, only red for overdue
        if "atrasado" in status.lower() or "vencido" in status.lower():
            status_color = SIGNAL_RED
        else:
            status_color = TEXT_COLOR

        status_ps = ParagraphStyle(
            f"cs_{len(rows)}", parent=styles["cell"],
            fontName="Helvetica-Bold", textColor=status_color,
        )

        rows.append([
            Paragraph(data_str, styles["cell"]),
            Paragraph(marco, styles["cell"]),
            Paragraph(status, status_ps),
        ])

    t = _three_rule_table(rows, [avail * 0.18, avail * 0.52, avail * 0.30])
    el.append(t)
    el.append(Spacer(1, 3 * mm))
    return el


def _build_roi_text(roi: dict, ed: dict, styles: dict, _state: dict | None = None) -> list:
    """Build ROI indicator with auditable calculation memory (E1)."""
    if not roi or not isinstance(roi, dict):
        return []
    roi_min = roi.get("roi_min", 0)
    roi_max = roi.get("roi_max", 0)
    probability = roi.get("probability", 0)

    el: list = []

    # E1: Strategic reclassification — marginal ROI on substantial contracts
    reclass = roi.get("strategic_reclassification")
    if reclass == "INVESTIMENTO_ESTRATEGICO_ACERVO":
        rationale = _s(roi.get("reclassification_rationale", ""))
        el.append(Paragraph(
            f"<b>Classificação: Investimento Estratégico em Acervo</b>",
            styles["body"],
        ))
        el.append(Paragraph(rationale, styles["body_small"]))
        # Still show calculation memory for auditability
        calc = roi.get("calculation_memory", {})
        if calc:
            el.append(Paragraph(
                f"Memória de cálculo: {_s(calc.get('roi_max_calc', ''))}",
                styles["caption"],
            ))
        el.append(Spacer(1, 2 * mm))
        return el

    if roi_max <= 0:
        return []

    roi_text = f"{_currency_short(roi_min)} — {_currency_short(roi_max)}"
    prob_text = _pct(probability, 1) if probability else "N/I"
    confidence = roi.get("confidence", "")

    conf_label = ""
    if confidence == "alta":
        conf_label = " (confiança alta)"
    elif confidence == "media":
        conf_label = " (confiança média)"
    elif confidence == "baixa":
        conf_label = " (base setorial)"

    el.append(Paragraph(
        f"<b>Resultado Potencial:</b> {roi_text}  |  "
        f"Probabilidade de vitória: {prob_text}{conf_label}",
        styles["body"],
    ))

    # E1: Auditable calculation memory — each factor explicit
    calc = roi.get("calculation_memory", {})
    if calc:
        el.append(Paragraph(
            f"<b>Memória de cálculo</b> (fórmula: {_s(calc.get('formula', 'N/I'))})",
            styles["caption"],
        ))
        el.append(Paragraph(
            f"ROI mínimo: {_s(calc.get('roi_min_calc', 'N/I'))}",
            styles["caption"],
        ))
        el.append(Paragraph(
            f"ROI máximo: {_s(calc.get('roi_max_calc', 'N/I'))}",
            styles["caption"],
        ))
    else:
        # Fallback: reconstruct from available data (backward compat)
        win_prob = ed.get("win_probability", {})
        memo_parts = []
        valor_edital = _safe_float(ed.get("valor_estimado"))
        if valor_edital > 0:
            memo_parts.append(f"Valor: {_currency(valor_edital)}")
        if probability > 0:
            memo_parts.append(f"Prob.: {_dec(probability, 4)}")
        margin_range = roi.get("margin_range", "")
        if margin_range:
            memo_parts.append(f"Margem: {margin_range}")
        if memo_parts:
            el.append(Paragraph(
                "Memória: " + " × ".join(memo_parts),
                styles["caption"],
            ))

    if _state is not None and not _state.get("roi_shown"):
        el.append(Paragraph(
            "Resultado potencial = valor do edital × probabilidade de vitória × margem líquida do setor. "
            "Probabilidade calculada via modelo competitivo (fornecedores históricos, "
            "modalidade, incumbência) ajustado pelo índice de viabilidade.",
            styles["caption"],
        ))
        _state["roi_shown"] = True

    el.append(Spacer(1, 2 * mm))
    return el


def _build_competitive_section(data: dict, styles: dict, sec: dict) -> list:
    entries = []
    seen_orgaos = set()
    for ed in data.get("editais", []):
        ci = ed.get("competitive_intel", {})
        orgao = _s(ed.get("orgao", ""))
        if orgao in seen_orgaos:
            continue
        if isinstance(ci, dict) and ci.get("top_fornecedores"):
            seen_orgaos.add(orgao)
            for c in ci["top_fornecedores"][:5]:
                objs = c.get("objetos", [])
                obj_str = _trunc(_s(objs[0] if objs else ""), 60)
                aditivo = _safe_float(c.get("valor_aditivos"))
                situacao = c.get("situacao_contrato", "")
                if aditivo > 0:
                    obj_str += f" [aditivos: +{_currency_short(aditivo)}]"
                if situacao == "3":  # rescindido
                    obj_str += " [RESCINDIDO]"
                entries.append({
                    "orgao": orgao,
                    "fornecedor": _s(c.get("nome", "")),
                    "objeto": obj_str,
                    "valor": c.get("valor_total"),
                    "data": "",
                })
        elif isinstance(ci, list):
            for c in ci[:5]:
                obj_str = _trunc(_s(c.get("objeto", "")), 60)
                aditivo = _safe_float(c.get("valor_aditivos"))
                situacao = c.get("situacao_contrato", "")
                if aditivo > 0:
                    obj_str += f" [aditivos: +{_currency_short(aditivo)}]"
                if situacao == "3":  # rescindido
                    obj_str += " [RESCINDIDO]"
                entries.append({
                    "orgao": orgao,
                    "fornecedor": _s(c.get("fornecedor", c.get("nome", ""))),
                    "objeto": obj_str,
                    "valor": c.get("valor", c.get("valor_total")),
                    "data": c.get("data", ""),
                })
    if not entries:
        return []

    el = []
    num = sec["next"]()
    el.extend(_section_heading(f"{num}. Mapa Competitivo", styles))
    el.append(Paragraph(
        "Contratos históricos identificados nos órgãos licitantes — "
        "indica incumbentes e valores praticados.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN
    header = [
        Paragraph("Órgão", styles["cell_header"]),
        Paragraph("Fornecedor", styles["cell_header"]),
        Paragraph("Objeto", styles["cell_header"]),
        Paragraph("Valor", styles["cell_header_right"]),
        Paragraph("Data", styles["cell_header_center"]),
    ]
    rows = [header]

    seen = set()
    for e in sorted(entries, key=lambda x: x["data"], reverse=True):
        key = (e["orgao"][:30], e["fornecedor"][:30])
        if key in seen:
            continue
        seen.add(key)
        rows.append([
            Paragraph(_trunc(e["orgao"], 35), styles["cell"]),
            Paragraph(_trunc(e["fornecedor"], 30), styles["cell"]),
            Paragraph(e["objeto"], styles["cell"]),
            Paragraph(_currency_short(e["valor"]), styles["cell_right"]),
            Paragraph(_date(e["data"]), styles["cell_center"]),
        ])
        if len(rows) > 20:
            break

    t = _three_rule_table(rows, [
        avail * 0.20, avail * 0.20, avail * 0.30, avail * 0.15, avail * 0.15,
    ])
    el.append(t)
    el.append(Spacer(1, 6 * mm))

    # E5: Historical dispute stats by typology
    dispute_stats = data.get("dispute_stats", {})
    stats_by_typ = dispute_stats.get("stats_by_typology", {})
    if stats_by_typ:
        el.append(Paragraph("<b>Estatísticas Históricas de Disputas</b>", styles["h3"]))
        el.append(Paragraph(
            "Dados agregados de contratos encerrados — participantes, descontos e taxa de adjudicação por modalidade e faixa de valor.",
            styles["body_small"],
        ))
        el.append(Spacer(1, 3 * mm))

        ds_header = [
            Paragraph("Modalidade / Faixa", styles["cell_header"]),
            Paragraph("Contratos", styles["cell_header_center"]),
            Paragraph("Part. Médio", styles["cell_header_center"]),
            Paragraph("Desc. Médio", styles["cell_header_center"]),
            Paragraph("Adjudicação", styles["cell_header_center"]),
        ]
        ds_rows = [ds_header]

        for key, stats in sorted(stats_by_typ.items()):
            total = stats.get("total", 0)
            if total == 0:
                continue
            avg_p = stats.get("avg_participants")
            avg_d = stats.get("avg_discount")
            adj_r = stats.get("adjudication_rate")
            ds_rows.append([
                Paragraph(_s(key.replace("_", " / ")), styles["cell"]),
                Paragraph(str(total), styles["cell_center"]),
                Paragraph(_dec(avg_p, 1) if avg_p is not None else "—", styles["cell_center"]),
                Paragraph(_pct(avg_d, 1) if avg_d is not None else "—", styles["cell_center"]),
                Paragraph(_pct(adj_r) if adj_r is not None else "—", styles["cell_center"]),
            ])

        if len(ds_rows) > 1:
            ds_t = _three_rule_table(ds_rows, [
                avail * 0.30, avail * 0.15, avail * 0.18, avail * 0.18, avail * 0.19,
            ])
            el.append(ds_t)
            el.append(Spacer(1, 4 * mm))

    # E5: Recurring suppliers with advantage highlighting
    recurring = dispute_stats.get("recurring_suppliers", [])
    if recurring:
        el.append(Paragraph("<b>Fornecedores Recorrentes (Incumbentes)</b>", styles["h3"]))
        # Compute market_share from contract counts if not provided
        total_contracts = sum(sup.get("n_contracts", sup.get("contract_count", 0)) for sup in recurring)
        for sup in recurring[:10]:
            name = _s(sup.get("nome_ou_cnpj", sup.get("fornecedor", "")))
            count = sup.get("n_contracts", sup.get("contract_count", 0))
            region = _s(sup.get("region", ""))
            # Derive share from contract count proportion
            share = sup.get("market_share", count / total_contracts if total_contracts > 0 else 0)

            # Highlight competitive dynamics
            if share > 0.60:
                indicator = f" <font color='{SIGNAL_RED.hexval()}'><b>[DOMINANTE {_pct(share)}]</b></font>"
            elif share > 0.40:
                indicator = f" <font color='{SIGNAL_AMBER.hexval()}'><b>[FORTE {_pct(share)}]</b></font>"
            else:
                indicator = ""

            el.append(Paragraph(
                f"  {name} — {count} contrato(s){f' ({region})' if region else ''}{indicator}",
                styles["body_small"],
            ))
        # F19: Note when supplier list was truncated
        if len(recurring) > 10:
            el.append(Paragraph(
                f"Exibidos os 10 principais fornecedores de {len(recurring)} identificados.",
                ParagraphStyle(
                    "recurring_trunc", parent=styles["caption"],
                    fontName="Helvetica-Oblique", textColor=TEXT_MUTED,
                ),
            ))
        el.append(Spacer(1, 4 * mm))

    # Competitive advantage summary
    editais = data.get("editais", [])
    favorable = []
    for ed in editais:
        idx = ed.get("_display_idx", 0)
        wp = ed.get("win_probability", {})
        if isinstance(wp, dict) and wp.get("top_supplier_share", 1) < 0.20 and wp.get("n_unique_suppliers", wp.get("unique_suppliers", 0)) >= 3:
            favorable.append((idx, _trunc(_s(ed.get("objeto", "")), 60), wp.get("probability", 0)))
    if favorable:
        el.append(Paragraph(
            f"<font color='{SIGNAL_GREEN.hexval()}'><b>Mercados Favoráveis à Entrada</b></font>",
            styles["h3"],
        ))
        el.append(Paragraph(
            f"{len(favorable)} edital(is) sem incumbente dominante (HHI baixo, nenhum fornecedor >20% do mercado):",
            styles["body_small"],
        ))
        for idx, obj, prob in favorable[:5]:
            el.append(Paragraph(
                f"  <b>{idx}.</b> {obj} — prob. vitória: {_pct(prob)}",
                styles["body_small"],
            ))
        el.append(Spacer(1, 4 * mm))

    el.append(Spacer(1, 4 * mm))
    return el


def _section_counter() -> dict:
    state = {"n": 0}

    def _next() -> int:
        state["n"] += 1
        return state["n"]

    return {"next": _next, "current": lambda: state["n"]}


def _build_company_profile_content(data: dict, styles: dict) -> list:
    """Render company profile content without section heading."""
    el = []
    emp = data.get("empresa", {})

    avail = PAGE_WIDTH - 2 * MARGIN

    # Key-value pairs — minimal two-column table with left label column
    info_rows = []
    raw_fields = [
        ("Razão Social", emp.get("razao_social")),
        ("Nome Fantasia", emp.get("nome_fantasia")),
        ("CNPJ", emp.get("cnpj")),
        ("CNAE Principal", emp.get("cnae_principal")),
        ("CNAEs Secundários", _collapse_cnaes(emp.get("cnaes_secundarios"))),
        ("Porte", emp.get("porte")),
        ("Capital Social", _currency(emp.get("capital_social")) if emp.get("capital_social") else None),
        ("Sede", f"{emp.get('cidade_sede', '')} — {emp.get('uf_sede', '')}"),
        ("Situação Cadastral", emp.get("situacao_cadastral")),
    ]
    # Simples Nacional / MEI
    simples = emp.get("simples_nacional")
    mei = emp.get("mei")
    if simples is not None or mei is not None:
        regime_parts = []
        if mei:
            regime_parts.append("MEI")
        elif simples:
            data_opcao = emp.get("data_opcao_simples", "")
            regime_parts.append(f"Simples Nacional{' (desde ' + str(data_opcao)[:10] + ')' if data_opcao else ''}")
        else:
            regime_parts.append("Regime Geral")
        raw_fields.append(("Regime Tributário", " · ".join(regime_parts)))
    for label, value in raw_fields:
        if value and str(value).strip() and value != " — ":
            info_rows.append([
                Paragraph(f"<b>{label}</b>", ParagraphStyle(
                    f"lbl_{label[:6]}", parent=styles["cell"],
                    fontName="Helvetica-Bold", textColor=TEXT_SECONDARY,
                )),
                Paragraph(_s(str(value)), styles["cell"]),
            ])

    if info_rows:
        info_t = Table(info_rows, colWidths=[avail * 0.22, avail * 0.78])
        info_t.setStyle(TableStyle([
            ("LINEABOVE", (0, 0), (-1, 0), 0.6, RULE_HEAVY),
            ("LINEBELOW", (0, -1), (-1, -1), 0.4, RULE_COLOR),
            *[("LINEBELOW", (0, i), (-1, i), 0.2, RULE_COLOR) for i in range(len(info_rows) - 1)],
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ]))
        el.append(info_t)
        el.append(Spacer(1, 4 * mm))

    # QSA
    qsa = emp.get("qsa", [])
    if qsa:
        el.append(Paragraph("Quadro Societário", styles["h3"]))
        for socio in qsa[:5]:
            nome = _s(socio.get("nome", socio) if isinstance(socio, dict) else socio)
            qual = _s(socio.get("qualificacao", "")) if isinstance(socio, dict) else ""
            line = f"— {nome}" + (f" ({qual})" if qual else "")
            el.append(Paragraph(line, styles["bullet"]))
        el.append(Spacer(1, 2 * mm))

    # E8: Maturity profile badge
    maturity = data.get("maturity_profile") or emp.get("maturity_profile", {})
    if maturity and maturity.get("profile"):
        profile = maturity["profile"]
        rationale = _s(maturity.get("rationale", ""))
        profile_labels = {
            "ENTRANTE": ("Entrante — Novo no mercado governamental federal", TEXT_SECONDARY),
            "REGIONAL": ("Regional — Experiência consolidada em âmbito regional", INK),
            "ESTABELECIDO": ("Estabelecido — Portfólio diversificado", colors.HexColor("#2D7D46")),
        }
        label, color = profile_labels.get(profile, (profile, TEXT_SECONDARY))
        el.append(Paragraph(
            f"<b>Perfil de Maturidade Licitatória:</b> <font color='{color.hexval()}'><b>{label}</b></font>",
            styles["body"],
        ))
        if rationale:
            el.append(Paragraph(rationale, styles["caption"]))
        el.append(Spacer(1, 2 * mm))

    # Sanctions — simple text, no colored cards
    sancoes = emp.get("sancoes", {})
    if sancoes:
        is_inconclusive = sancoes.get("inconclusive", False)
        has_sanction = any(sancoes.get(k) for k in ["ceis", "cnep", "cepim", "ceaf"])
        if is_inconclusive:
            sanc_text = (
                f"<b><font color='{SIGNAL_AMBER.hexval()}'>Verificação pendente</font></b> — "
                "recomendamos consulta direta ao Portal da Transparência antes de submeter propostas."
            )
        elif has_sanction:
            sanc_text = f"<b><font color='{SIGNAL_RED.hexval()}'>Atenção:</font></b> Empresa possui sanção ativa — "
            details = []
            for k, label in [("ceis", "CEIS"), ("cnep", "CNEP"), ("cepim", "CEPIM"), ("ceaf", "CEAF")]:
                if sancoes.get(k):
                    details.append(label)
            sanc_text += ", ".join(details)
        else:
            sanc_text = "Sem sanções ativas (CEIS, CNEP, CEPIM, CEAF verificados)"
        el.append(Paragraph(sanc_text, styles["body"]))
        el.append(Spacer(1, 2 * mm))

    # Government contract history
    historico = emp.get("historico_contratos", [])
    if historico:
        el.append(Paragraph("Histórico de Contratos Governamentais", styles["h3"]))
        valor_hist = sum(_safe_float(c.get("valor")) for c in historico)
        hist_text = f"Total: {len(historico)} contrato(s)"
        if valor_hist > 0:
            hist_text += f"  |  Valor acumulado: {_currency(valor_hist)}"
        el.append(Paragraph(hist_text, styles["body"]))
    else:
        el.append(Paragraph(
            "Sem histórico de contratos governamentais federais identificado.",
            styles["body_small"],
        ))

    el.append(Spacer(1, 8 * mm))
    return el


def _build_company_profile(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Build company profile section with heading + content."""
    el = []
    num = sec["next"]() if sec else 1
    el.extend(_section_heading(f"{num}. Perfil da Empresa", styles))
    el.extend(_build_company_profile_content(data, styles))
    return el


def _build_executive_summary(data: dict, styles: dict, sec: dict | None = None) -> list:
    el = []
    editais = data.get("editais", [])
    resumo = data.get("resumo_executivo", {})

    num = sec["next"]() if sec else 2
    el.extend(_section_heading(f"{num}. Resumo Executivo", styles))

    texto = _s(resumo.get("texto", ""))
    if texto:
        el.append(Paragraph(texto, styles["body"]))
        el.append(Spacer(1, 4 * mm))

    # Metrics — clean boxes with thin borders
    total = len(editais)
    participar = sum(1 for e in editais if (e.get("recomendacao") or "").upper().startswith("PARTICIPAR"))
    cautela = sum(1 for e in editais if "CAUTELA" in (e.get("recomendacao") or "").upper() or "AVALIAR" in (e.get("recomendacao") or "").upper())
    valores = [_safe_float(e.get("valor_estimado")) for e in editais if e.get("valor_estimado")]
    valor_total = sum(valores)

    avail = PAGE_WIDTH - 2 * MARGIN
    col_w = avail / 4
    metrics = Table(
        [[
            _metric_cell(str(total), "Oportunidades", styles),
            _metric_cell(str(participar), "Participar", styles),
            _metric_cell(str(cautela), "Avaliar", styles),
            _metric_cell(_currency_short(valor_total), "Valor Total", styles),
        ]],
        colWidths=[col_w] * 4, rowHeights=[20 * mm],
    )
    metrics.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 0.6, RULE_HEAVY),
        ("LINEBELOW", (0, 0), (-1, 0), 0.4, RULE_COLOR),
        ("LINEBEFORE", (1, 0), (1, 0), 0.3, RULE_COLOR),
        ("LINEBEFORE", (2, 0), (2, 0), 0.3, RULE_COLOR),
        ("LINEBEFORE", (3, 0), (3, 0), 0.3, RULE_COLOR),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    el.append(metrics)
    el.append(Spacer(1, 3 * mm))

    # F41: Excel companion reference
    n_editais = len(editais)
    if n_editais > 0:
        el.append(Paragraph(
            f"<i>Este relatório acompanha planilha Excel com todos os {n_editais} editais "
            f"para análise detalhada e filtragem personalizada.</i>",
            ParagraphStyle(
                "excel_ref", parent=styles["body_small"],
                fontName="Helvetica-Oblique", fontSize=8, textColor=TEXT_MUTED,
            ),
        ))
        el.append(Spacer(1, 3 * mm))

    # Discard note — inform reader that irrelevant editais were filtered out
    n_desc = data.get("_descartados_count", 0)
    if n_desc > 0:
        motivos = data.get("_descartados_motivos", "")
        nota = (
            f"{n_desc} licitação(ões) identificada(s) na busca foi(ram) descartada(s) por falta de "
            f"aderência aos CNAEs ou perfil da empresa"
        )
        if motivos:
            nota += f": {motivos}"
        nota += "."
        el.append(Paragraph(
            f"<i>{nota}</i>",
            ParagraphStyle(
                "discard_note", parent=styles["caption"],
                textColor=colors.HexColor("#888888"), fontSize=7.5,
                spaceAfter=4 * mm,
            ),
        ))

    # UF distribution
    uf_counts: dict[str, int] = {}
    for e in editais:
        uf = e.get("uf", "N/I")
        if uf:
            uf_counts[uf] = uf_counts.get(uf, 0) + 1
    if uf_counts and len(uf_counts) > 1:
        el.append(Paragraph("Distribuição por UF", styles["h3"]))
        header = [
            Paragraph("UF", styles["cell_header_center"]),
            Paragraph("Qtd", styles["cell_header_center"]),
            Paragraph("%", styles["cell_header_center"]),
        ]
        rows = [header]
        for uf, cnt in sorted(uf_counts.items(), key=lambda x: -x[1])[:8]:
            pct = cnt / total * 100 if total else 0
            rows.append([
                Paragraph(uf, styles["cell_center"]),
                Paragraph(str(cnt), styles["cell_center"]),
                Paragraph(f"{pct:.0f}%".replace(".", ","), styles["cell_center"]),
            ])
        tw = avail * 0.45 if len(uf_counts) <= 4 else avail * 0.6
        t = _three_rule_table(rows, [tw * 0.30, tw * 0.35, tw * 0.35])
        el.append(t)
        el.append(Spacer(1, 4 * mm))
    elif uf_counts:
        uf_name = list(uf_counts.keys())[0]
        el.append(Paragraph(f"UF: {uf_name} ({total} editais)", styles["body"]))
        el.append(Spacer(1, 3 * mm))

    # Highlights
    destaques = resumo.get("destaques", [])
    if destaques:
        el.append(Paragraph("Destaques", styles["h3"]))
        for d in destaques:
            el.append(Paragraph(f"— {_s(d)}", styles["bullet"]))

    el.append(Spacer(1, 8 * mm))
    return el


def _build_strategic_positioning(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Posicionamento Estrategico — thesis box + 3-signal summary table.

    Renders strategic_thesis data: EXPANDIR/MANTER/REDUZIR thesis with
    trend, concentration (HHI), and price signals.
    """
    thesis_data = data.get("strategic_thesis")
    if not thesis_data:
        return []

    thesis = thesis_data.get("thesis", "")
    rationale = _s(thesis_data.get("rationale", ""))
    # Sanitize: remove nonsensical discount mentions from pre-generated rationale
    import re
    rationale = re.sub(r'desconto médio de apenas -?\d{3,}[,.]?\d*%\s*—\s*margens comprimidas\.?\s*', '', rationale)
    rationale = re.sub(r'desconto médio de apenas N/I%?\s*—\s*margens comprimidas\.?\s*', '', rationale)
    confidence = thesis_data.get("confidence", "")
    signals = thesis_data.get("signals", {})

    if not thesis:
        return []

    el: list = []
    num = sec["next"]() if sec else 2
    el.extend(_section_heading(f"{num}. Posicionamento Estratégico", styles))

    avail = PAGE_WIDTH - 2 * MARGIN

    # Thesis box — color-coded single-row table with left border accent
    thesis_colors = {
        "EXPANDIR": (colors.HexColor("#E8F5E9"), SIGNAL_GREEN),
        "MANTER": (colors.HexColor("#FFF8E1"), SIGNAL_AMBER),
        "REDUZIR": (colors.HexColor("#FFEBEE"), SIGNAL_RED),
    }
    bg_color, border_color = thesis_colors.get(thesis, (BG_SUBTLE, INK))

    confidence_labels = {"alta": "Alta", "media": "Média", "baixa": "Baixa"}
    conf_text = confidence_labels.get(confidence, "")
    conf_suffix = f"  (confiança: {conf_text})" if conf_text else ""

    thesis_cell = Paragraph(
        f"<font color='{border_color.hexval()}'><b>{thesis}</b></font>"
        f"<font size='8' color='{TEXT_SECONDARY.hexval()}'>{conf_suffix}</font>"
        f"<br/><font size='9' color='{TEXT_COLOR.hexval()}'>{rationale}</font>",
        ParagraphStyle(
            "thesis_cell", parent=styles["body"],
            fontName="Times-Roman", fontSize=10, leading=14,
            spaceBefore=0, spaceAfter=0,
        ),
    )
    thesis_t = Table([[thesis_cell]], colWidths=[avail])
    thesis_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), bg_color),
        ("LINEBEFORE", (0, 0), (0, 0), 3, border_color),
        ("TOPPADDING", (0, 0), (0, 0), 6),
        ("BOTTOMPADDING", (0, 0), (0, 0), 6),
        ("LEFTPADDING", (0, 0), (0, 0), 10),
        ("RIGHTPADDING", (0, 0), (0, 0), 8),
    ]))
    el.append(thesis_t)
    el.append(Spacer(1, 4 * mm))

    # 3-column signal summary table
    # Handle both nested-dict and flat-key signal formats
    trend = signals.get("trend", {})
    hhi = signals.get("hhi", {})
    price = signals.get("price", {})
    # Flat-key fallback (collect-report-data.py outputs flat keys)
    if isinstance(trend, str):
        trend = {"growth_rate_pct": signals.get("growth_rate_pct", 0), "label": trend}
    if isinstance(hhi, str):
        hhi = {"classification": hhi, "value": signals.get("hhi_value", 0)}
    if isinstance(price, (int, float)):
        price = {"avg_discount_pct": price}
    if not price and "avg_discount_pct" in signals:
        price = {"avg_discount_pct": signals["avg_discount_pct"]}

    if trend or hhi or price:
        # Row 1: labels
        h_trend = Paragraph("<b>Tendência de Volume</b>", styles["cell_header_center"])
        h_hhi = Paragraph("<b>Concentração</b>", styles["cell_header_center"])
        h_price = Paragraph("<b>Preço vs Estimado</b>", styles["cell_header_center"])

        # Row 2: values
        growth = trend.get("growth_rate_pct", 0)
        growth_sign = "+" if growth > 0 else ""
        growth_text = f"{growth_sign}{_dec(growth)}%"
        v_trend = Paragraph(f"<b>{growth_text}</b>", styles["cell_center"])

        hhi_class = _s(hhi.get("classification", "N/I"))
        hhi_labels = {
            "COMPETITIVO": "Competitivo",
            "MODERADO": "Moderado",
            "CONCENTRADO": "Concentrado",
        }
        v_hhi = Paragraph(
            f"<b>{hhi_labels.get(hhi_class.upper(), hhi_class)}</b>",
            styles["cell_center"],
        )

        discount = price.get("avg_discount_pct", 0)
        # Clamp nonsensical discount values (e.g. -44778%)
        if abs(discount) > 100:
            discount = 0
        discount_text = f"{_dec(discount)}%" if discount else "N/I"
        v_price = Paragraph(f"<b>{discount_text}</b>", styles["cell_center"])

        # Row 3: interpretation
        trend_label = trend.get("trend", "")
        trend_interp = {
            "EXPANSAO": "Mercado em expansão",
            "ESTAVEL": "Volume estável",
            "CONTRACAO": "Volume em contração",
        }
        i_trend = Paragraph(
            trend_interp.get(trend_label, ""),
            ParagraphStyle("i_trend", parent=styles["cell_center"],
                           fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED),
        )

        n_sup = hhi.get("n_suppliers", 0)
        i_hhi_text = f"{n_sup} fornecedores" if n_sup else ""
        hhi_val = hhi.get("hhi", 0)
        if hhi_val:
            i_hhi_text += f" (HHI {_dec(hhi_val, 2)})" if i_hhi_text else f"HHI {_dec(hhi_val, 2)}"
        i_hhi = Paragraph(
            i_hhi_text,
            ParagraphStyle("i_hhi", parent=styles["cell_center"],
                           fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED),
        )

        n_bench = price.get("editais_with_benchmark", 0)
        i_price_text = f"Base: {n_bench} editais" if n_bench else ""
        i_price = Paragraph(
            i_price_text,
            ParagraphStyle("i_price", parent=styles["cell_center"],
                           fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED),
        )

        col_w = avail / 3
        rows = [
            [h_trend, h_hhi, h_price],
            [v_trend, v_hhi, v_price],
            [i_trend, i_hhi, i_price],
        ]
        t = _three_rule_table(rows, [col_w, col_w, col_w])
        el.append(t)

    el.append(Spacer(1, 6 * mm))
    return el


def _build_overview_table(editais_list: list, styles: dict, start_idx: int = 1) -> list:
    avail = PAGE_WIDTH - 2 * MARGIN
    col_widths = [
        avail * 0.06,   # #
        avail * 0.40,   # Objeto + Órgão
        avail * 0.06,   # UF
        avail * 0.14,   # Valor
        avail * 0.10,   # Prazo
        avail * 0.24,   # Recomendação
    ]

    header = [
        Paragraph("#", styles["cell_header_center"]),
        Paragraph("Objeto / Órgão", styles["cell_header"]),
        Paragraph("UF", styles["cell_header_center"]),
        Paragraph("Valor", styles["cell_header_right"]),
        Paragraph("Prazo", styles["cell_header_center"]),
        Paragraph("Recomendação", styles["cell_header"]),
    ]
    rows = [header]

    # Cap overview table to MAX_OVERVIEW_ROWS to prevent page explosion
    display_list = editais_list[:MAX_OVERVIEW_ROWS]
    overflow_count = len(editais_list) - len(display_list)

    for idx, ed in enumerate(display_list, start_idx):
        rec = _normalize_recommendation(_s(ed.get("recomendacao", "")))
        rec_info = REC_STYLES.get(rec, REC_STYLES["NÃO RECOMENDADO"])

        rec_style = ParagraphStyle(
            f"rec_{idx}", parent=styles["cell"],
            fontName="Helvetica-Bold", textColor=rec_info["color"], fontSize=7,
        )

        objeto = _trunc(_s(ed.get("objeto", "")), 150)
        orgao = _trunc(_s(ed.get("orgao", "")), 80)
        ov_link = _fix_pncp_link(ed.get("link", ""))
        if ov_link and ov_link.startswith("http"):
            objeto_orgao = (
                f'<a href="{ov_link}" color="{INK.hexval()}"><b>{objeto}</b></a>'
                f'<br/><font size=\'7\' color=\'{TEXT_MUTED.hexval()}\'>{orgao}</font>'
            )
        else:
            objeto_orgao = f"<b>{objeto}</b><br/><font size='7' color='{TEXT_MUTED.hexval()}'>{orgao}</font>"

        prazo = _format_prazo_short(ed.get("dias_restantes"))
        if prazo == "—":
            enc_date = _date(ed.get("data_encerramento"))
            prazo = enc_date if enc_date != "N/I" else "—"

        rows.append([
            Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
            Paragraph(objeto_orgao, styles["cell"]),
            Paragraph(_s(ed.get("uf", "")), styles["cell_center"]),
            Paragraph(_currency_short(ed.get("valor_estimado")), styles["cell_right"]),
            Paragraph(prazo, styles["cell_center"]),
            Paragraph(rec, ParagraphStyle(
                f"rec2_{idx}", parent=rec_style, wordWrap="CJK",
            )),
        ])

    result = [_three_rule_table(rows, col_widths)]
    if overflow_count > 0:
        result.append(Paragraph(
            f"+ {overflow_count} edital(is) adicionais disponíveis na planilha Excel em anexo",
            ParagraphStyle(
                "ov_overflow_note", parent=styles["caption"],
                fontName="Helvetica-Oblique", textColor=TEXT_MUTED,
                spaceBefore=2 * mm,
            ),
        ))
    return result


def _build_opportunities_overview(data: dict, styles: dict, sec: dict | None = None) -> list:
    el = []
    editais = data.get("editais", [])
    if not editais:
        return el

    num = sec["next"]() if sec else 3
    el.extend(_section_heading(f"{num}. Panorama de Oportunidades", styles))
    el.append(Spacer(1, 2 * mm))

    # Sort: PARTICIPAR first, then AVALIAR, then NÃO RECOMENDADO; within each, by score desc
    editais_sorted = sorted(
        editais,
        key=lambda e: (
            REC_ORDER.get(_normalize_recommendation(_s(e.get("recomendacao", ""))), 9),
            -(e.get("risk_score", {}).get("total", 0)),
        ),
    )
    el.extend(_build_overview_table(editais_sorted, styles, start_idx=1))
    el.append(Spacer(1, 8 * mm))
    return el


def _build_detailed_analysis(data: dict, styles: dict, sec: dict | None = None, _state: dict | None = None) -> list:
    el = []
    editais = data.get("editais", [])
    if not editais:
        return el

    num = sec["next"]() if sec else 4
    el.extend(_section_heading(f"{num}. Análise Detalhada por Edital", styles))
    el.append(Spacer(1, 2 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN

    # Collect PARTICIPAR + AVALIAR editais (NÃO RECOMENDADO / VETOED go to Annex A)
    _detailed_candidates = []
    for ed in editais:
        idx = ed.get("_display_idx", 0)
        rec = _normalize_recommendation(_s(ed.get("recomendacao", "")))
        risk = ed.get("risk_score", {}) or {}
        vetoed = risk.get("vetoed", False)
        if rec == "NÃO RECOMENDADO" or vetoed:
            continue
        score = _safe_int((ed.get("risk_score") or {}).get("total", 0))
        _detailed_candidates.append((idx, ed, score))

    # Sort by risk_score.total descending and apply hard cap
    _detailed_candidates.sort(key=lambda x: x[2], reverse=True)
    all_eligible = _detailed_candidates
    detailed_top = _detailed_candidates[:MAX_DETAILED_EDITAIS]
    detailed_overflow = _detailed_candidates[MAX_DETAILED_EDITAIS:]

    # F18: Truncation warning when editals exceed detail limit
    if len(all_eligible) > MAX_DETAILED_EDITAIS:
        truncation_note = Paragraph(
            f"<i>Análise detalhada dos {MAX_DETAILED_EDITAIS} editais prioritários. "
            f"Os demais {len(all_eligible) - MAX_DETAILED_EDITAIS} editais estão disponíveis "
            f"na planilha Excel anexa para análise complementar.</i>",
            ParagraphStyle(
                "truncation_note", parent=styles["body"],
                fontName="Helvetica-Oblique", fontSize=9, textColor=TEXT_MUTED,
            ),
        )
        el.append(truncation_note)
        el.append(Spacer(1, 4 * mm))

    for idx, ed, _score in detailed_top:
        rec = _normalize_recommendation(_s(ed.get("recomendacao", "")))
        risk = ed.get("risk_score", {}) or {}
        vetoed = risk.get("vetoed", False)

        header_block = []

        objeto = _s(ed.get("objeto", "Sem título"))

        # Edital title — clean serif, no colored background; clickable if link exists
        link = _fix_pncp_link(ed.get("link", ""))
        title_text = f"<b>{num}.{idx}.</b>  {objeto}"
        if link and link.startswith("http"):
            title_text = f'<b>{num}.{idx}.</b>  <a href="{link}" color="{INK.hexval()}">{objeto}</a>'
        header_block.append(Paragraph(
            title_text,
            styles["h2"],
        ))

        # CRÍTICA 5: Acervo confirmation note — warn when technical capacity is unconfirmed
        acervo_confirmado = risk.get("acervo_confirmado", False)
        if not acervo_confirmado:
            header_block.append(Paragraph(
                "⚠ A participação efetiva depende de verificação prévia dos atestados técnicos "
                "compatíveis com o objeto licitado.",
                ParagraphStyle(
                    f"acervo_warn_{idx}", parent=styles["body"],
                    fontName="Helvetica-Oblique", fontSize=8, textColor=SIGNAL_AMBER,
                    spaceBefore=1 * mm, spaceAfter=1 * mm,
                ),
            ))

        # INSIGHT-FIRST: Lead with strategic rationale before technical data
        justificativa = _s(ed.get("justificativa", ""))
        rec_info_block = REC_STYLES.get(rec, REC_STYLES["NÃO RECOMENDADO"])
        if justificativa:
            header_block.append(Paragraph(
                f"<font color='{rec_info_block['color'].hexval()}'><b>{rec}</b></font>"
                f"<font size='9' color='{TEXT_SECONDARY.hexval()}'> — {justificativa}</font>",
                ParagraphStyle(
                    f"insight_{idx}", parent=styles["body"],
                    fontName="Times-Bold", fontSize=10, textColor=rec_info_block["color"],
                    spaceBefore=1 * mm, spaceAfter=2 * mm,
                ),
            ))

        # Key strategic metrics bar (viability + probability + differential)
        wp = ed.get("win_probability", {})
        risk = ed.get("risk_score", {})
        roi = ed.get("roi_potential", {})
        strategic_bar_parts = []
        if isinstance(risk, dict) and risk.get("total", 0) > 0:
            score = _safe_int(risk.get("total"))
            if score >= 60:
                qualif = "Alta"
            elif score >= 30:
                qualif = "Moderada"
            else:
                qualif = "Baixa"
            strategic_bar_parts.append(f"Viabilidade: {score}/100 ({qualif})")
        if isinstance(wp, dict) and wp.get("probability", 0) > 0:
            strategic_bar_parts.append(f"Prob. vitória: {_pct(wp['probability'])}")
        if isinstance(roi, dict) and roi.get("roi_max", 0) > 0:
            strategic_bar_parts.append(f"Resultado potencial: até {_currency_short(roi['roi_max'])}")
        if ed.get("strategic_category"):
            cat_labels = {"QUICK_WIN": "Quick Win", "OPORTUNIDADE": "Oportunidade",
                          "INVESTIMENTO": "Investimento Estratégico", "INACESSÍVEL": "Inacessível",
                          "BAIXA_PRIORIDADE": "Baixa Prioridade"}
            strategic_bar_parts.append(f"Categoria: {cat_labels.get(ed['strategic_category'], ed['strategic_category'])}")

        if strategic_bar_parts:
            header_block.append(Paragraph(
                " &nbsp;|&nbsp; ".join(strategic_bar_parts),
                ParagraphStyle(
                    f"stbar_{idx}", parent=styles["body_small"],
                    fontName="Helvetica-Bold", fontSize=8, textColor=INK,
                    spaceBefore=0, spaceAfter=3 * mm,
                ),
            ))

        # Ficha técnica — minimal key-value
        info_rows = []
        raw_fields = [
            ("Órgão", ed.get("orgao")),
            ("UF / Município", (
                f"{ed.get('uf')} — {ed.get('municipio')}" if ed.get("uf") and ed.get("municipio")
                else f"{ed.get('uf')} — município não informado" if ed.get("uf")
                else f"UF não informada — {ed.get('municipio')}" if ed.get("municipio")
                else "Localização não informada"
            )),
            ("Modalidade", ed.get("modalidade")),
            ("Valor Estimado", _currency(ed.get("valor_estimado")) if ed.get("valor_estimado") else None),
            ("Data de Abertura", _date(ed.get("data_abertura"))),
            ("Data de Encerramento", _date(ed.get("data_encerramento"))),
            ("Situação", _format_dias_restantes(ed.get("dias_restantes"))),
            ("Compatibilidade", ed.get("object_compatibility", {}).get("compatibility")),
            ("Categoria Estratégica", ed.get("strategic_category")),
            ("Fonte", ed.get("fonte")),
        ]
        # IBGE municipal data
        ibge = ed.get("ibge", {})
        if ibge.get("populacao") or ibge.get("pib_mil_reais"):
            ibge_parts = []
            if ibge.get("populacao"):
                ibge_parts.append(f"Pop. {_fmt_pop(ibge['populacao'])}")
            if ibge.get("pib_mil_reais"):
                pib = ibge["pib_mil_reais"]
                if pib >= 1_000_000:
                    ibge_parts.append(f"PIB R${pib/1_000_000:.1f}B")
                elif pib >= 1_000:
                    ibge_parts.append(f"PIB R${pib/1_000:.0f}M")
                else:
                    ibge_parts.append(f"PIB R${pib:.0f}k")
            raw_fields.append(("Município (IBGE)", " · ".join(ibge_parts)))
        link = _fix_pncp_link(ed.get("link", ""))
        if link and link != "N/I" and link.startswith("http"):
            raw_fields.append(("Link", f'<a href="{link}" color="{TEXT_SECONDARY.hexval()}">{link}</a>'))

        for label, value in raw_fields:
            if value and str(value).strip() and value != "N/I" and value != " — N/I" and str(value) != "None":
                info_rows.append([
                    Paragraph(f"<b>{label}</b>", ParagraphStyle(
                        f"ft_lbl_{idx}_{label[:4]}", parent=styles["cell"],
                        fontName="Helvetica-Bold", textColor=TEXT_SECONDARY,
                    )),
                    Paragraph(_s(str(value)), styles["cell"]),
                ])
        if info_rows:
            info_t = Table(info_rows, colWidths=[avail * 0.20, avail * 0.80])
            n_info = len(info_rows)
            info_t.setStyle(TableStyle([
                ("LINEABOVE", (0, 0), (-1, 0), 0.4, RULE_COLOR),
                ("LINEBELOW", (0, n_info - 1), (-1, n_info - 1), 0.4, RULE_COLOR),
                *[("LINEBELOW", (0, i), (-1, i), 0.15, RULE_COLOR) for i in range(n_info - 1)],
                ("BACKGROUND", (0, 0), (0, -1), BG_SUBTLE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ]))
            header_block.append(info_t)
            header_block.append(Spacer(1, 3 * mm))

        # Recommendation already shown at top via insight-first block
        # (justificativa + strategic bar rendered before ficha técnica)

        el.append(KeepTogether(header_block))

        # IBGE fiscal capacity red flag
        ibge = ed.get("ibge", {})
        pop = ibge.get("populacao")
        valor_est = _safe_float(ed.get("valor_estimado"))
        if pop and valor_est and pop < 10_000 and valor_est > 5_000_000:
            el.append(Paragraph(
                f"<font color='{SIGNAL_AMBER.hexval()}'><b>ALERTA</b></font> "
                f"Município com {pop:,} habitantes licitando {_currency_short(valor_est)} — verificar capacidade fiscal",
                styles["body_small"],
            ))

        # Distance
        distancia = ed.get("distancia", {})
        if isinstance(distancia, dict) and distancia.get("km"):
            km = distancia["km"]
            hrs = distancia.get("duracao_horas", "")
            label, _ = _get_source_label(distancia.get("_source"))
            dist_text = f"<b>Distância da sede:</b> {km} km"
            if hrs:
                dist_text += f" (~{hrs}h de carro)"
            dist_text += f"  [{label}]"
            el.append(Paragraph(dist_text, styles["body"]))

        # Viability + ROI — keep together
        metric_block = []
        risk = ed.get("risk_score", {})
        if isinstance(risk, dict) and risk.get("total"):
            metric_block.extend(_build_viability_text(risk, styles, _state))
            metric_block.append(Spacer(1, 1 * mm))

        roi = ed.get("roi_potential", {})
        metric_block.extend(_build_roi_text(roi, ed, styles, _state))

        if metric_block:
            el.append(KeepTogether(metric_block))

        # Habilitação gap analysis
        hab = ed.get("habilitacao_analysis", {})
        el.extend(_build_habilitacao_table(hab, styles))

        # Systemic risk flags
        risk_an = ed.get("risk_analysis", {})
        el.extend(_build_risk_flags(risk_an, styles))

        # E6: Organ risk profile
        organ_risk = ed.get("organ_risk", {})
        if organ_risk and organ_risk.get("organ_track_record") != "INDETERMINADO":
            track = organ_risk.get("organ_track_record", "")
            track_colors = {"BOM": colors.HexColor("#2D7D46"), "REGULAR": colors.HexColor("#B8860B"), "RISCO": SIGNAL_RED}
            track_color = track_colors.get(track, TEXT_SECONDARY)
            el.append(Paragraph(
                f"<b>Risco do Edital (Histórico do Órgão):</b> "
                f"<font color='{track_color.hexval()}'><b>{track}</b></font>",
                styles["body"],
            ))
            details = []
            if organ_risk.get("similar_published", 0) > 0:
                details.append(f"{organ_risk['similar_published']} contratação(ões) similar(es) no histórico")
            adj = organ_risk.get("adjudication_rate")
            if adj is not None:
                details.append(f"Taxa de adjudicação: {_pct(adj)}")
            timeline = organ_risk.get("timeline_assessment", "")
            if timeline and timeline != "INDETERMINADO":
                rationale = _s(organ_risk.get("timeline_rationale", ""))
                details.append(f"Prazo: {timeline} — {rationale}")
            for flag in organ_risk.get("risk_flags", []):
                details.append(f"⚠ {_s(flag)}")
            for d in details:
                el.append(Paragraph(f"  {d}", styles["body_small"]))
            el.append(Spacer(1, 2 * mm))

        # E4: Qualification gap analysis
        qual_gap = ed.get("qualification_gap", {})
        if qual_gap:
            filter_result = qual_gap.get("filter_result", "")
            if filter_result == "INCOMPATÍVEL_CNAE":
                el.append(Paragraph(
                    f"<font color='{SIGNAL_RED.hexval()}'><b>Incompatível com CNAEs da empresa</b></font>",
                    styles["body"],
                ))
                rationale = _s(qual_gap.get("incompatibility_rationale", ""))
                if rationale:
                    el.append(Paragraph(rationale, styles["body_small"]))
                el.append(Spacer(1, 2 * mm))
            elif qual_gap.get("operational_gaps"):
                el.append(Paragraph("<b>Lacunas Operacionais (endereçáveis)</b>", styles["body"]))
                for gap in qual_gap["operational_gaps"]:
                    gap_type = gap.get("gap_type", "")
                    desc = _s(gap.get("description", ""))
                    timeline = gap.get("estimated_timeline", "")
                    action = _s(gap.get("action_required", ""))
                    el.append(Paragraph(
                        f"  [{gap_type}] {desc} — <i>{timeline}</i>",
                        styles["body_small"],
                    ))
                    if action:
                        el.append(Paragraph(f"    Ação: {action}", styles["caption"]))
                el.append(Spacer(1, 2 * mm))

        # Chronogram
        cronograma = ed.get("cronograma", [])
        el.extend(_build_chronogram_table(cronograma, styles))

        # Analysis sections — clean key-value
        analise = ed.get("analise", {})
        analysis_rows = []
        analysis_fields = [
            ("Aderência ao Perfil", "aderencia"),
            ("Análise de Valor", "valor"),
            ("Análise Geográfica", "geografica"),
            ("Análise de Prazo", "prazo"),
            ("Análise de Modalidade", "modalidade"),
            ("Competitividade", "competitividade"),
            ("Riscos e Alertas", "riscos"),
        ]
        for title, key in analysis_fields:
            text = _s(analise.get(key, ""))
            if text:
                analysis_rows.append([
                    Paragraph(f"<b>{title}</b>", ParagraphStyle(
                        f"an_lbl_{idx}_{key[:4]}", parent=styles["cell"],
                        fontName="Helvetica-Bold", textColor=TEXT_SECONDARY,
                    )),
                    Paragraph(text, styles["cell"]),
                ])

        if analysis_rows:
            n_an = len(analysis_rows)
            box_t = Table(analysis_rows, colWidths=[avail * 0.20, avail * 0.80 - 2 * mm])
            box_t.setStyle(TableStyle([
                ("LINEABOVE", (0, 0), (-1, 0), 0.4, RULE_COLOR),
                ("LINEBELOW", (0, n_an - 1), (-1, n_an - 1), 0.4, RULE_COLOR),
                *[("LINEBELOW", (0, i), (-1, i), 0.15, RULE_COLOR) for i in range(n_an - 1)],
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]))
            el.append(box_t)
            el.append(Spacer(1, 3 * mm))

        # Scenarios mini-table + sensitivity + triggers
        scenarios = ed.get("scenarios", {})
        if scenarios:
            sc_base = scenarios.get("base", {})
            sc_opt = scenarios.get("optimistic", {})
            sc_pess = scenarios.get("pessimistic", {})

            def _roi_range(sc: dict) -> str:
                r_min = sc.get("roi_min")
                r_max = sc.get("roi_max")
                if r_min is None or r_max is None:
                    return "—"
                return f"{_currency_short(r_min)} a {_currency_short(r_max)}"

            def _prob_fmt(sc: dict) -> str:
                p = sc.get("prob")
                return _pct(p) if p is not None else "—"

            sc_header = [
                Paragraph("<b>Cenários</b>", styles["cell_header"]),
                Paragraph("<b>Pessimista</b>", styles["cell_header_center"]),
                Paragraph("<b>Base</b>", styles["cell_header_center"]),
                Paragraph("<b>Otimista</b>", styles["cell_header_center"]),
            ]
            sc_rows = [sc_header]
            sc_rows.append([
                Paragraph("Probabilidade", styles["cell"]),
                Paragraph(_prob_fmt(sc_pess), styles["cell_center"]),
                Paragraph(_prob_fmt(sc_base), styles["cell_center"]),
                Paragraph(_prob_fmt(sc_opt), styles["cell_center"]),
            ])
            sc_rows.append([
                Paragraph("Retorno (R$)", styles["cell"]),
                Paragraph(
                    _roi_range(sc_pess),
                    ParagraphStyle(f"sc_p_{idx}", parent=styles["cell_center"],
                                   textColor=SIGNAL_RED, fontSize=7),
                ),
                Paragraph(_roi_range(sc_base), styles["cell_center"]),
                Paragraph(
                    _roi_range(sc_opt),
                    ParagraphStyle(f"sc_o_{idx}", parent=styles["cell_center"],
                                   textColor=SIGNAL_GREEN, fontSize=7),
                ),
            ])
            # Triggers row — only if at least one scenario has a trigger
            pess_trigger = _s(sc_pess.get("trigger", ""))
            opt_trigger = _s(sc_opt.get("trigger", ""))
            if pess_trigger or opt_trigger:
                trigger_style = ParagraphStyle(
                    f"sc_t_{idx}", parent=styles["cell_center"],
                    fontName="Helvetica", fontSize=6.5, textColor=TEXT_MUTED,
                )
                sc_rows.append([
                    Paragraph("Gatilho", styles["cell"]),
                    Paragraph(pess_trigger or "—", trigger_style),
                    Paragraph("—", trigger_style),
                    Paragraph(opt_trigger or "—", trigger_style),
                ])

            sc_col_w = avail * 0.85
            sc_t = Table(
                sc_rows,
                colWidths=[sc_col_w * 0.22, sc_col_w * 0.26, sc_col_w * 0.26, sc_col_w * 0.26],
            )
            n_sc = len(sc_rows)
            sc_t.setStyle(TableStyle([
                ("LINEABOVE", (0, 0), (-1, 0), 0.6, RULE_HEAVY),
                ("LINEBELOW", (0, 0), (-1, 0), 0.4, RULE_HEAVY),
                ("LINEBELOW", (0, n_sc - 1), (-1, n_sc - 1), 0.4, RULE_COLOR),
                *[("LINEBELOW", (0, i), (-1, i), 0.15, RULE_COLOR) for i in range(1, n_sc - 1)],
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            el.append(KeepTogether([sc_t]))
            el.append(Spacer(1, 2 * mm))

        # Sensitivity badge
        sensitivity = ed.get("sensitivity", {})
        if sensitivity:
            stability = sensitivity.get("stability", "")
            if stability == "ROBUSTA":
                score_range = sensitivity.get("score_range", ["", ""])
                el.append(Paragraph(
                    f"<font color='{SIGNAL_GREEN.hexval()}'><b>ROBUSTA</b></font>"
                    f"<font size='8' color='{TEXT_SECONDARY.hexval()}'>"
                    f" — Score {sensitivity.get('original_score', '')} "
                    f"(faixa {score_range[0]}–{score_range[1]})"
                    f"</font>",
                    styles["body_small"],
                ))
            elif stability == "FRAGIL":
                sensitive_to = _s(sensitivity.get("sensitive_to", ""))
                fragil_detail = f" (sensível a {sensitive_to})" if sensitive_to else ""
                score_range = sensitivity.get("score_range", ["", ""])
                el.append(Paragraph(
                    f"<font color='{SIGNAL_AMBER.hexval()}'><b>FRÁGIL</b></font>"
                    f"<font size='8' color='{TEXT_SECONDARY.hexval()}'>"
                    f"{fragil_detail}"
                    f" — Score {sensitivity.get('original_score', '')} "
                    f"(faixa {score_range[0]}–{score_range[1]})"
                    f"</font>",
                    styles["body_small"],
                ))
            el.append(Spacer(1, 1 * mm))

        # Triggers (decision triggers)
        triggers = ed.get("triggers", [])
        if triggers:
            for trig in triggers[:2]:
                condition = _s(trig.get("condition", ""))
                action = _s(trig.get("action", ""))
                if condition and action:
                    el.append(Paragraph(
                        f"<b>Se</b> {condition} → {action}",
                        ParagraphStyle(
                            f"trig_{idx}", parent=styles["body_small"],
                            fontName="Helvetica", fontSize=7.5, textColor=TEXT_SECONDARY,
                            leftIndent=5,
                        ),
                    ))
            el.append(Spacer(1, 2 * mm))

        # Q&A section
        perguntas = ed.get("perguntas_decisor", {})
        if perguntas:
            el.append(Paragraph("Perguntas do Decisor", styles["h3"]))
            for pergunta, resposta in perguntas.items():
                if resposta:
                    el.append(Paragraph(
                        f"<b>{_s(pergunta)}</b>",
                        ParagraphStyle(
                            f"qa_q_{idx}", parent=styles["body"],
                            fontName="Times-Bold", fontSize=9, textColor=TEXT_COLOR,
                            spaceAfter=0.5 * mm,
                        ),
                    ))
                    el.append(Paragraph(
                        _s(resposta),
                        ParagraphStyle(
                            f"qa_a_{idx}", parent=styles["body"],
                            fontName="Times-Roman", fontSize=9, textColor=TEXT_SECONDARY,
                            leftIndent=0, spaceAfter=3 * mm,
                        ),
                    ))

        el.append(Spacer(1, 10 * mm))

    # Condensed overflow table for editais beyond MAX_DETAILED_EDITAIS
    if detailed_overflow:
        el.append(Paragraph(
            f"Demais {len(detailed_overflow)} edital(is) analisado(s) — ver planilha Excel em anexo para detalhamento completo",
            ParagraphStyle(
                "overflow_note", parent=styles["body_small"],
                fontName="Helvetica-Oblique", textColor=TEXT_SECONDARY,
                spaceBefore=4 * mm, spaceAfter=2 * mm,
            ),
        ))
        ov_header = [
            Paragraph("#", styles["cell_header_center"]),
            Paragraph("Objeto / Órgão", styles["cell_header"]),
            Paragraph("UF", styles["cell_header_center"]),
            Paragraph("Valor", styles["cell_header_right"]),
            Paragraph("Score", styles["cell_header_center"]),
            Paragraph("Recomendação", styles["cell_header"]),
        ]
        ov_rows = [ov_header]
        for idx, ed, score in detailed_overflow:
            rec_ov = _normalize_recommendation(_s(ed.get("recomendacao", "")))
            rec_info_ov = REC_STYLES.get(rec_ov, REC_STYLES["NÃO RECOMENDADO"])
            objeto_ov = _s(ed.get("objeto", ""))
            orgao_ov = _s(ed.get("orgao", ""))
            objeto_orgao_ov = f"<b>{_trunc(objeto_ov, 120)}</b><br/><font size='7' color='{TEXT_MUTED.hexval()}'>{_trunc(orgao_ov, 60)}</font>"
            link_ov = _fix_pncp_link(ed.get("link", ""))
            if link_ov and link_ov.startswith("http"):
                objeto_orgao_ov = f'<a href="{link_ov}" color="{INK.hexval()}"><b>{_trunc(objeto_ov, 120)}</b></a><br/><font size=\'7\' color=\'{TEXT_MUTED.hexval()}\'>{_trunc(orgao_ov, 60)}</font>'
            score_color = SIGNAL_GREEN if score >= 60 else (SIGNAL_AMBER if score >= 30 else SIGNAL_RED)
            ov_rows.append([
                Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
                Paragraph(objeto_orgao_ov, styles["cell"]),
                Paragraph(_s(ed.get("uf", "")), styles["cell_center"]),
                Paragraph(_currency_short(ed.get("valor_estimado")), styles["cell_right"]),
                Paragraph(
                    f"<b>{score}</b>",
                    ParagraphStyle(f"ov_sc_{idx}", parent=styles["cell_center"],
                                   fontName="Helvetica-Bold", textColor=score_color),
                ),
                Paragraph(rec_ov, ParagraphStyle(
                    f"ov_rec_{idx}", parent=styles["cell"],
                    fontName="Helvetica-Bold", fontSize=7,
                    textColor=rec_info_ov["color"], wordWrap="CJK",
                )),
            ])
        ov_col_w = [avail * 0.05, avail * 0.38, avail * 0.05, avail * 0.13, avail * 0.09, avail * 0.30]
        el.append(_three_rule_table(ov_rows, ov_col_w))
        el.append(Spacer(1, 6 * mm))

    return el


def _build_market_intelligence(data: dict, styles: dict, sec: dict | None = None) -> list:
    el = []
    intel = data.get("inteligencia_mercado", {})
    if not intel:
        return el

    num = sec["next"]() if sec else 5
    el.extend(_section_heading(f"{num}. Inteligência de Mercado", styles))

    for title, key in [
        ("Panorama Setorial", "panorama"),
        ("Tendências", "tendencias"),
        ("Vantagens Competitivas da Empresa", "vantagens"),
        ("Recomendação Geral", "recomendacao_geral"),
    ]:
        text = _s(intel.get(key, ""))
        if text:
            el.append(Paragraph(title, styles["h3"]))
            for paragraph in text.split("\n"):
                paragraph = paragraph.strip()
                if paragraph:
                    if paragraph.startswith("•") or paragraph.startswith("-"):
                        # Replace bullets with em-dash
                        paragraph = "— " + paragraph.lstrip("•- ")
                        el.append(Paragraph(paragraph, styles["bullet"]))
                    else:
                        el.append(Paragraph(paragraph, styles["body"]))

    el.append(Spacer(1, 8 * mm))
    return el


def _build_querido_diario_content(data: dict, styles: dict) -> list:
    """Render Querido Diario mentions without section heading."""
    el = []
    mencoes = data.get("querido_diario", [])
    if not mencoes:
        return el

    # Filter QD: keep only mentions from UFs where editais exist or empresa is based
    empresa = data.get("empresa", {})
    uf_sede = (empresa.get("uf_sede", "") or "").upper()
    editais_ufs = {(e.get("uf", "") or "").upper() for e in data.get("editais", [])}
    relevant_ufs = editais_ufs | ({uf_sede} if uf_sede else set())

    if relevant_ufs:
        filtered = []
        for m in mencoes:
            territorio = (m.get("territorio", "") or "").upper()
            if any(uf in territorio for uf in relevant_ufs if uf):
                filtered.append(m)
        if not filtered:
            filtered = mencoes[:3]
        mencoes = filtered

    if not mencoes:
        return el

    el.append(Paragraph(
        "Publicações em diários oficiais de municípios relevantes para o perfil da empresa.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN

    header = [
        Paragraph("#", styles["cell_header_center"]),
        Paragraph("Data / Município", styles["cell_header"]),
        Paragraph("Trecho Relevante", styles["cell_header"]),
    ]
    rows = [header]

    for idx, m in enumerate(mencoes[:10], 1):
        data_str = _date(m.get("data"))
        territorio = _s(m.get("territorio", ""))
        local_info = f"<b>{data_str}</b><br/><font size='7' color='{TEXT_MUTED.hexval()}'>{territorio}</font>"

        excerpts = m.get("excerpts", [])
        excerpt_texts = []
        for exc in excerpts[:2]:
            text = _s(exc.get("text", exc) if isinstance(exc, dict) else exc)
            if text:
                excerpt_texts.append(f"<i>\"{_trunc(text, 200)}\"</i>")

        trecho = "<br/>".join(excerpt_texts) if excerpt_texts else "<i>Sem trecho disponível</i>"

        rows.append([
            Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
            Paragraph(local_info, styles["cell"]),
            Paragraph(f"<font size='7'>{trecho}</font>", styles["cell"]),
        ])

    t = _three_rule_table(rows, [avail * 0.05, avail * 0.22, avail * 0.73])
    el.append(t)
    el.append(Spacer(1, 8 * mm))
    return el


def _build_querido_diario(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Build Querido Diario section with heading + content."""
    mencoes = data.get("querido_diario", [])
    if not mencoes:
        return []
    el = []
    num = sec["next"]() if sec else 6
    el.extend(_section_heading(f"{num}. Menções em Diários Oficiais", styles))
    el.extend(_build_querido_diario_content(data, styles))
    return el


def _build_prioritization(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Consolidated ranking of PARTICIPAR editais — hierarchized for the decision-maker."""
    editais = data.get("editais", [])
    participar = []
    for ed in editais:
        idx = ed.get("_display_idx", 0)
        rec = _normalize_recommendation(_s(ed.get("recomendacao", "")))
        if rec != "PARTICIPAR":
            continue
        risk = ed.get("risk_score", {})
        roi = ed.get("roi_potential", {})
        wp = ed.get("win_probability", {})
        score = _safe_int(risk.get("total")) if isinstance(risk, dict) else 0
        roi_max = _safe_float(roi.get("roi_max", 0))
        prob = _safe_float(wp.get("probability", 0))
        # Composite ranking: weighted by ROI, viability, and probability
        rank_score = roi_max * 0.4 + score * prob * 100 * 0.6
        participar.append((idx, ed, score, roi_max, prob, rank_score))

    if len(participar) < 2:
        return []  # No need to prioritize 0-1 editais

    participar.sort(key=lambda x: x[5], reverse=True)

    el = []
    num = sec["next"]() if sec else 12
    el.extend(_section_heading(f"{num}. Priorização Consolidada", styles))
    el.append(Paragraph(
        f"Ranking dos {len(participar)} editais recomendados para participação, "
        f"ordenados por retorno esperado ajustado à viabilidade. "
        f"A empresa deve focar nos primeiros da lista e descer conforme capacidade operacional.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN
    header = [
        Paragraph("Rank", styles["cell_header_center"]),
        Paragraph("Edital", styles["cell_header"]),
        Paragraph("Viab.", styles["cell_header_center"]),
        Paragraph("Prob.", styles["cell_header_center"]),
        Paragraph("Resultado", styles["cell_header_right"]),
        Paragraph("Ação Imediata", styles["cell_header"]),
    ]
    rows = [header]

    for rank, (idx, ed, score, roi_max, prob, _) in enumerate(participar, 1):
        obj = _trunc(_s(ed.get("objeto", "")), 150)
        pri_link = _fix_pncp_link(ed.get("link", ""))
        if pri_link and pri_link.startswith("http"):
            obj = f'<a href="{pri_link}" color="{INK.hexval()}">{obj}</a>'
        dias = ed.get("dias_restantes")
        urgency = f"({_format_prazo_short(dias)})" if dias is not None else ""

        if rank <= 3:
            action = f"Preparar proposta imediatamente {urgency}"
        elif rank <= 6:
            action = f"Preparar se houver capacidade {urgency}"
        else:
            action = f"Avaliar custo-benefício {urgency}"

        score_color = SIGNAL_GREEN if score >= 60 else (SIGNAL_AMBER if score >= 30 else SIGNAL_RED)

        rows.append([
            Paragraph(f"<b>{rank}º</b>", styles["cell_center"]),
            Paragraph(f"<b>{idx}.</b> {obj}", styles["cell"]),
            Paragraph(
                f"<b>{score}</b>",
                ParagraphStyle(f"pri_v_{rank}", parent=styles["cell_center"],
                               fontName="Helvetica-Bold", textColor=score_color),
            ),
            Paragraph(_pct(prob), styles["cell_center"]),
            Paragraph(_currency_short(roi_max), styles["cell_right"]),
            Paragraph(action, styles["cell"]),
        ])

    t = _three_rule_table(rows, [
        avail * 0.06, avail * 0.30, avail * 0.08, avail * 0.08, avail * 0.14, avail * 0.34,
    ])
    el.append(t)
    el.append(Spacer(1, 6 * mm))
    return el


def _build_development_path(data: dict, styles: dict, sec: dict | None = None) -> list:
    """For NÃO RECOMENDADO editais — what the company needs to build in 24 months to compete."""
    editais = data.get("editais", [])
    nao_rec = []
    for ed in editais:
        idx = ed.get("_display_idx", 0)
        rec = _normalize_recommendation(_s(ed.get("recomendacao", "")))
        if rec != "NÃO RECOMENDADO":
            continue
        nao_rec.append((idx, ed))

    if not nao_rec:
        return []

    el = []
    num = sec["next"]() if sec else 13
    el.extend(_section_heading(f"{num}. Caminho de Desenvolvimento (24 meses)", styles))
    el.append(Paragraph(
        "Editais não recomendados no momento atual — mas que representam mercados acessíveis "
        "com investimento em capacitação. Abaixo, o que a empresa precisa construir para "
        "participar dessas oportunidades em futuras janelas.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN

    for idx, ed in nao_rec[:8]:  # Top 8
        obj = _trunc(_s(ed.get("objeto", "")), 150)
        dev_link = _fix_pncp_link(ed.get("link", ""))
        if dev_link and dev_link.startswith("http"):
            obj = f'<a href="{dev_link}" color="{INK.hexval()}">{obj}</a>'
        justif = _s(ed.get("justificativa", ""))
        valor = _currency_short(ed.get("valor_estimado"))

        # Derive what's needed from qualification gaps and risk factors
        qual_gap = ed.get("qualification_gap", {})
        risk_an = ed.get("risk_analysis", {})
        gaps = qual_gap.get("operational_gaps", [])
        flags = risk_an.get("flags", []) if isinstance(risk_an, dict) else []

        actions = []
        for g in gaps:
            if g.get("addressable") and g.get("gap_type") != "ACERVO_EXISTENTE":
                actions.append(f"{g['gap_type']}: {_trunc(_s(g.get('description', '')), 60)}")
        for f in flags:
            if f.get("severity") == "ALTA":
                actions.append(f"Resolver: {_trunc(_s(f.get('flag', '')), 60)}")

        if not actions:
            # Infer from justificativa
            if justif:
                actions.append(f"Superar: {_trunc(justif, 80)}")
            else:
                actions.append("Avaliar requisitos específicos do edital")

        el.append(Paragraph(
            f"<b>{idx}.</b> {obj} — {valor}",
            styles["body"],
        ))
        if justif:
            el.append(Paragraph(
                f"<i>Motivo atual: {_trunc(justif, 120)}</i>",
                styles["caption"],
            ))
        for action in actions[:3]:
            el.append(Paragraph(f"  → {action}", styles["body_small"]))
        el.append(Spacer(1, 2 * mm))

    el.append(Spacer(1, 6 * mm))
    return el


def _build_next_steps(data: dict, styles: dict, sec: dict | None = None) -> list:
    el = []
    proximos = data.get("proximos_passos", [])

    num = sec["next"]() if sec else 7
    el.extend(_section_heading(f"{num}. Próximos Passos", styles))

    avail = PAGE_WIDTH - 2 * MARGIN

    if proximos:
        header = [
            Paragraph("Ação", styles["cell_header"]),
            Paragraph("Prazo", styles["cell_header_center"]),
            Paragraph("Prioridade", styles["cell_header_center"]),
        ]
        rows = [header]
        for step in proximos:
            if isinstance(step, dict):
                acao = _s(step.get("acao", ""))
                prazo = _s(step.get("prazo", ""))
                prioridade = _s(step.get("prioridade", ""))
                prio_upper = prioridade.upper()
                prio_color = SIGNAL_RED if "URGENTE" in prio_upper or "ALTA" in prio_upper else TEXT_COLOR
                rows.append([
                    Paragraph(acao, styles["cell"]),
                    Paragraph(prazo, styles["cell_center"]),
                    Paragraph(f"<font color='{prio_color.hexval()}'><b>{prioridade}</b></font>", styles["cell_center"]),
                ])
            else:
                rows.append([
                    Paragraph(_s(step), styles["cell"]),
                    Paragraph("—", styles["cell_center"]),
                    Paragraph("—", styles["cell_center"]),
                ])

        t = _three_rule_table(rows, [avail * 0.60, avail * 0.22, avail * 0.18])
        el.append(t)
    else:
        for i, text in enumerate([
            "Revisar os editais marcados como PARTICIPAR e iniciar preparação documental.",
            "Avaliar os editais marcados como AVALIAR COM CAUTELA conforme capacidade operacional.",
            "Monitorar novos editais semanalmente para oportunidades adicionais.",
        ], 1):
            el.append(Paragraph(f"{i}. {text}", styles["bullet"]))

    el.append(Spacer(1, 8 * mm))

    # Contact card
    contact_t = Table(
        [[Paragraph(
            "Para dúvidas ou acompanhamento:<br/>"
            "<b>Tiago Sasaki</b> — Consultor de Inteligência em Licitações<br/>"
            "(48) 9 8834-4559",
            styles["body"],
        )]],
        colWidths=[avail],
    )
    contact_t.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 0.4, RULE_COLOR),
        ("LINEBELOW", (0, 0), (-1, 0), 0.4, RULE_COLOR),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    el.append(contact_t)

    el.append(Spacer(1, 8 * mm))
    return el


# ============================================================
# SICAF & SOURCE CONFIDENCE
# ============================================================

def _build_sicaf_content(data: dict, styles: dict) -> list:
    """Render SICAF content without section heading."""
    el = []
    sicaf = data.get("sicaf", {})
    if not sicaf:
        return el

    label, label_color = _get_source_label(sicaf.get("_source"))

    # E2: Handle SICAF collection failure explicitly
    sicaf_status = sicaf.get("status", "")
    if sicaf_status == "FALHA_COLETA":
        attempted_at = sicaf.get("attempted_at", "N/I")
        error_detail = _s(sicaf.get("error_detail", "erro não especificado"))
        amber = colors.HexColor("#D4760A")
        el.append(Paragraph(
            f"<font color='{amber.hexval()}'><b>ANÁLISE DE HABILITAÇÃO INCOMPLETA</b></font>",
            styles["body"],
        ))
        el.append(Paragraph(
            f"Coleta SICAF falhou em <b>{attempted_at}</b>.",
            styles["body"],
        ))
        el.append(Paragraph(
            f"Motivo: <i>{error_detail}</i>",
            styles["body"],
        ))
        el.append(Spacer(1, 2 * mm))
        el.append(Paragraph(
            "Esta seção não foi omitida por irrelevância. A regularidade fiscal é determinante "
            "para a recomendação final. A ausência de dados SICAF significa que não foi possível "
            "confirmar a situação cadastral da empresa — <b>não</b> que ela foi verificada e está regular.",
            styles["body_small"],
        ))
        el.append(Spacer(1, 8 * mm))
        return el

    crc = sicaf.get("crc", {})
    restricao = sicaf.get("restricao", {})

    if crc or restricao:
        if crc:
            status_cad = _s(crc.get("status_cadastral", ""))
            color = INK if status_cad == "CADASTRADO" else SIGNAL_RED
            el.append(Paragraph(
                f"<b>Status Cadastral (CRC):</b> <font color='{color.hexval()}'><b>{status_cad}</b></font>",
                styles["body"],
            ))
            for field_label, key in [
                ("Razão Social", "razao_social"),
                ("CNAE", "atividade_principal"),
                ("Endereço", "endereco"),
                ("Emissão CRC", "data_emissao"),
            ]:
                val = crc.get(key)
                if val:
                    rendered = _date(val) if key == "data_emissao" else _s(val)
                    el.append(Paragraph(f"<b>{field_label}:</b> {rendered}", styles["body"]))

            hab = crc.get("habilitacao", {})
            if hab:
                el.append(Spacer(1, 2 * mm))
                el.append(Paragraph("Habilitações SICAF", styles["h3"]))
                for field_label, key in [
                    ("Habilitação Jurídica", "habilitacao_juridica"),
                    ("Fiscal Federal", "regularidade_fiscal_federal"),
                    ("Fiscal Estadual", "regularidade_fiscal_estadual"),
                    ("Fiscal Municipal", "regularidade_fiscal_municipal"),
                    ("Trabalhista", "regularidade_trabalhista"),
                    ("Qualificação Econômica", "qualificacao_economica"),
                ]:
                    val = hab.get(key)
                    if val:
                        hcolor = INK if val.lower() == "regular" else SIGNAL_RED
                        el.append(Paragraph(
                            f"— {field_label}: <font color='{hcolor.hexval()}'><b>{_s(val)}</b></font>",
                            styles["body_small"],
                        ))

            detalhe = crc.get("detalhe")
            if detalhe and status_cad != "CADASTRADO":
                el.append(Paragraph(f"<i>{_s(detalhe)}</i>", styles["body_small"]))

        el.append(Spacer(1, 3 * mm))

        if restricao:
            possui = restricao.get("possui_restricao", False)
            if possui:
                el.append(Paragraph(
                    f"<b>Restrições:</b> <font color='{SIGNAL_RED.hexval()}'><b>Verificar detalhes</b></font>",
                    styles["body"],
                ))
                for r in restricao.get("restricoes", []):
                    el.append(Paragraph(
                        f"— {_s(r.get('tipo', ''))} — {_s(r.get('detalhe', ''))}",
                        styles["body_small"],
                    ))
            else:
                el.append(Paragraph("Restrições: Nenhuma identificada.", styles["body"]))
    else:
        status = _s(sicaf.get("status", ""))
        instrucao = _s(sicaf.get("instrucao", ""))
        url = sicaf.get("url", "")

        el.append(Paragraph(f"<b>{status}</b>", styles["body"]))
        if instrucao:
            el.append(Paragraph(instrucao, styles["body"]))
        if url:
            el.append(Paragraph(f"Portal: {url}", styles["body"]))

    el.append(Spacer(1, 3 * mm))
    el.append(Paragraph(f"[{label}]", styles["caption"]))
    el.append(Spacer(1, 8 * mm))
    return el


def _build_sicaf_section(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Build SICAF section with heading + content."""
    sicaf = data.get("sicaf", {})
    if not sicaf:
        return []
    el = []
    num = sec["next"]() if sec else 8
    el.extend(_section_heading(f"{num}. Verificação SICAF", styles))
    el.extend(_build_sicaf_content(data, styles))
    return el


def _build_data_sources_content(data: dict, styles: dict) -> list:
    """Render data sources content without section heading."""
    el = []
    metadata = data.get("_metadata", {})
    sources = metadata.get("sources", {})
    if not sources:
        return el

    el.append(Paragraph(
        "Cada dado neste relatório foi obtido de fontes públicas oficiais. "
        "A tabela abaixo indica o status de cada consulta no momento da coleta.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN
    header = [
        Paragraph("Fonte", styles["cell_header"]),
        Paragraph("Status", styles["cell_header"]),
        Paragraph("Detalhe", styles["cell_header"]),
    ]
    rows = [header]

    source_labels = {
        "opencnpj": "Receita Federal (perfil da empresa)",
        "portal_transparencia_sancoes": "Portal da Transparência (sanções)",
        "portal_transparencia_contratos": "Portal da Transparência (contratos)",
        "pncp": "Portal Nacional de Contratações Públicas",
        "pcp_v2": "Portal de Compras Públicas (complementar)",
        "querido_diario": "Diários Oficiais Municipais",
        "sicaf": "Sistema de Cadastro de Fornecedores",
    }

    def _sanitize_detail(detail: str) -> str:
        """Remove technical terms and English from source details."""
        if not detail:
            return ""
        replacements = [
            ("raw", "brutos"), ("filtered", "filtrados"), ("pages", "páginas"),
            ("errors", "erros"), ("via Playwright", ""), ("via playwright", ""),
            ("SICAF completo", "Consulta realizada com sucesso"),
            ("Sem chave API", "Consulta não realizada"),
            ("Skipped", "Não consultado"), ("skipped", "não consultado"),
            ("men\u00e7\u00f5es encontradas", "menções encontradas"),
            ("mencoes encontradas", "menções encontradas"),
        ]
        for eng, pt in replacements:
            detail = detail.replace(eng, pt)
        detail = re.sub(r'\b\d{3}\s*OK\b', 'consultado', detail)
        detail = re.sub(r'\bhttpx?\b', '', detail, flags=re.IGNORECASE)
        detail = re.sub(r'\bGET\b', '', detail)
        detail = re.sub(r'\bPOST\b', '', detail)
        return detail.strip().strip(",").strip()

    for key, src_label in source_labels.items():
        src = sources.get(key, {})
        label, label_color = _get_source_label(src)
        detail = src.get("detail", "") if isinstance(src, dict) else ""
        detail = _sanitize_detail(detail)

        status_style = ParagraphStyle(
            f"src_{key}", parent=styles["cell"],
            fontName="Helvetica-Bold", textColor=label_color,
        )
        rows.append([
            Paragraph(src_label, styles["cell"]),
            Paragraph(label, status_style),
            Paragraph(_trunc(_s(detail), 120), styles["cell"]),
        ])

    t = _three_rule_table(rows, [avail * 0.35, avail * 0.20, avail * 0.45])
    el.append(t)

    el.append(Spacer(1, 4 * mm))
    gen_at = _date(metadata.get("generated_at", ""))
    if gen_at:
        el.append(Paragraph(
            f"Dados coletados em {gen_at}.",
            styles["caption"],
        ))

    el.append(Spacer(1, 8 * mm))
    return el


def _build_data_sources_section(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Build data sources section with heading + content."""
    metadata = data.get("_metadata", {})
    sources = metadata.get("sources", {})
    if not sources:
        return []
    el = []
    num = sec["next"]() if sec else 9
    el.extend(_section_heading(f"{num}. Fontes de Dados e Confiabilidade", styles))
    el.extend(_build_data_sources_content(data, styles))
    return el


# ============================================================
# SESSION 2-4 PDF SECTIONS
# ============================================================

_HAB_STATUS_COLORS = {
    "OK": colors.HexColor("#2D7D46"),
    "ATENÇÃO": colors.HexColor("#B8860B"),
    "INCOMPLETO": colors.HexColor("#D4760A"),  # E2: amber for collection failure
    "CRÍTICO": colors.HexColor("#B5342A"),
    "VERIFICAR": colors.HexColor("#5A6577"),
}

_HAB_OVERALL_LABELS = {
    "APTA": ("Apta a Participar", colors.HexColor("#2D7D46")),
    "PARCIALMENTE_APTA": ("Parcialmente Apta — Verificações Necessárias", colors.HexColor("#B8860B")),
    "INAPTA": ("Inapta — Impedimentos Identificados", colors.HexColor("#B5342A")),
}


def _build_habilitacao_table(hab: dict, styles: dict) -> list:
    """Render habilitação gap analysis as a status-colored table."""
    if not hab or not hab.get("dimensions"):
        return []

    el = []
    overall = hab.get("status", "?")
    label, label_color = _HAB_OVERALL_LABELS.get(overall, (overall, TEXT_SECONDARY))

    el.append(Paragraph(
        f"<b>Habilitação:</b> <font color='{label_color.hexval()}'>{label}</font>",
        ParagraphStyle("hab_title", parent=styles["body"], fontName="Helvetica-Bold", fontSize=9),
    ))

    avail = PAGE_WIDTH - 2 * MARGIN
    rows = [[
        Paragraph("<b>Dimensão</b>", styles["cell_header"]),
        Paragraph("<b>Status</b>", styles["cell_header"]),
        Paragraph("<b>Detalhe</b>", styles["cell_header"]),
    ]]

    for dim in hab["dimensions"]:
        status = dim.get("status", "?")
        status_color = _HAB_STATUS_COLORS.get(status, TEXT_SECONDARY)
        rows.append([
            Paragraph(dim.get("dimension", ""), styles["cell"]),
            Paragraph(f"<b>{status}</b>", ParagraphStyle(
                f"hab_s_{status[:3]}", parent=styles["cell"],
                fontName="Helvetica-Bold", textColor=status_color,
            )),
            Paragraph(_s(dim.get("detail", "")), styles["cell"]),
        ])

    t = _three_rule_table(rows, [avail * 0.18, avail * 0.12, avail * 0.70])
    el.append(t)
    el.append(Spacer(1, 2 * mm))
    return el


def _build_risk_flags(risk_analysis: dict, styles: dict) -> list:
    """Render systemic risk flags as colored bullets."""
    if not risk_analysis or not risk_analysis.get("flags"):
        return []

    el = []
    severity_markers = {
        "ALTA": f"<font color='{SIGNAL_RED.hexval()}'><b>[ALTO]</b></font>",
        "MEDIA": f"<font color='{colors.HexColor('#B8860B').hexval()}'><b>[MÉDIO]</b></font>",
        "BAIXA": f"<font color='{TEXT_SECONDARY.hexval()}'>[BAIXO]</font>",
    }

    risk_level = risk_analysis.get("risk_level", "")
    el.append(Paragraph(
        f"<b>Riscos Sistêmicos</b> — Nível geral: {risk_level}",
        ParagraphStyle("risk_title", parent=styles["body"], fontName="Helvetica-Bold", fontSize=9),
    ))

    for flag in risk_analysis["flags"]:
        marker = severity_markers.get(flag.get("severity", ""), "")
        el.append(Paragraph(
            f"  {marker}  {_s(flag.get('flag', ''))}",
            ParagraphStyle("risk_bullet", parent=styles["body_small"], leftIndent=8),
        ))

    el.append(Spacer(1, 2 * mm))
    return el


def _build_portfolio_section(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Render portfolio strategic matrix — new section between competitive and market intel."""
    editais = data.get("editais", [])
    if not editais:
        return []

    # Check if any edital has strategic_category
    if not any(ed.get("strategic_category") for ed in editais):
        return []

    el = []
    num = sec["next"]() if sec else 7
    el.extend(_section_heading(f"{num}. Matriz Estratégica de Portfólio", styles))

    el.append(Paragraph(
        "Classificação das oportunidades por potencial estratégico, "
        "combinando probabilidade de vitória, viabilidade técnica e valor de investimento.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    # Categorize
    categories = {
        "QUICK_WIN": {"label": "Quick Wins", "color": colors.HexColor("#2D7D46"), "editais": []},
        "OPORTUNIDADE": {"label": "Oportunidades", "color": colors.HexColor("#2563EB"), "editais": []},
        "INVESTIMENTO": {"label": "Investimentos Estratégicos", "color": colors.HexColor("#B8860B"), "editais": []},
        "INACESSÍVEL": {"label": "Inacessíveis", "color": SIGNAL_RED, "editais": []},
        "BAIXA_PRIORIDADE": {"label": "Baixa Prioridade", "color": TEXT_MUTED, "editais": []},
    }

    for ed in editais:
        idx = ed.get("_display_idx", 0)
        cat = ed.get("strategic_category", "BAIXA_PRIORIDADE")
        if cat in categories:
            categories[cat]["editais"].append((idx, ed))

    avail = PAGE_WIDTH - 2 * MARGIN

    # Summary counts row
    summary_rows = [[
        Paragraph("<b>Categoria</b>", styles["cell_header"]),
        Paragraph("<b>Qtd</b>", styles["cell_header"]),
        Paragraph("<b>Descrição</b>", styles["cell_header"]),
    ]]

    cat_descriptions = {
        "QUICK_WIN": "Alta probabilidade + alta viabilidade — prioridade máxima",
        "OPORTUNIDADE": "Probabilidade moderada — vale a preparação cuidadosa",
        "INVESTIMENTO": "Baixa probabilidade imediata, alto valor para acervo e relacionamento",
        "INACESSÍVEL": "Impedimentos de habilitação — não participar",
        "BAIXA_PRIORIDADE": "Baixo retorno esperado para o perfil atual",
    }

    for cat_key, cat_info in categories.items():
        n = len(cat_info["editais"])
        if n == 0:
            continue
        summary_rows.append([
            Paragraph(
                f"<b>{cat_info['label']}</b>",
                ParagraphStyle(f"port_{cat_key[:4]}", parent=styles["cell"],
                               fontName="Helvetica-Bold", textColor=cat_info["color"]),
            ),
            Paragraph(str(n), styles["cell"]),
            Paragraph(cat_descriptions.get(cat_key, ""), styles["cell"]),
        ])

    if len(summary_rows) > 1:
        t = _three_rule_table(summary_rows, [avail * 0.25, avail * 0.08, avail * 0.67])
        el.append(t)
        el.append(Spacer(1, 4 * mm))

    # Quick Wins detail
    qw = categories["QUICK_WIN"]["editais"]
    if qw:
        el.append(Paragraph("<b>Quick Wins — Ação Imediata Recomendada</b>", styles["h3"]))
        for idx, ed in qw:
            prob = ed.get("win_probability", {}).get("probability", 0)
            valor = ed.get("valor_estimado", 0)
            obj = _s((ed.get("objeto") or "")[:100])
            qw_link = _fix_pncp_link(ed.get("link", ""))
            if qw_link and qw_link.startswith("http"):
                obj = f'<a href="{qw_link}" color="{INK.hexval()}">{obj}</a>'
            el.append(Paragraph(
                f"<b>{idx}.</b> {obj} — Prob. {_pct(prob)}, Valor {_currency(valor)}",
                styles["body_small"],
            ))
        el.append(Spacer(1, 3 * mm))

    # Investments detail — with acervo unlock information
    inv = categories["INVESTIMENTO"]["editais"]
    if inv:
        el.append(Paragraph("<b>Investimentos Estratégicos — Construção de Acervo</b>", styles["h3"]))
        el.append(Paragraph(
            "Contratos cujo valor principal não é o retorno financeiro imediato, "
            "mas os atestados técnicos e relacionamentos que desbloqueiam mercados futuros.",
            styles["body_small"],
        ))
        el.append(Spacer(1, 2 * mm))

        inv_header = [
            Paragraph("#", styles["cell_header_center"]),
            Paragraph("Edital / Órgão", styles["cell_header"]),
            Paragraph("Valor", styles["cell_header_right"]),
            Paragraph("Acervo Desbloqueado", styles["cell_header"]),
        ]
        inv_rows = [inv_header]
        for idx, ed in inv:
            obj = _trunc(_s(ed.get("objeto") or ""), 70)
            orgao = _trunc(_s(ed.get("orgao") or ""), 40)
            valor = _currency_short(ed.get("valor_estimado"))
            # Derive acervo unlock from object keywords and qualification gaps
            acervo_parts = []
            qual_gap = ed.get("qualification_gap", {})
            for gap in qual_gap.get("operational_gaps", []):
                desc = _s(gap.get("description", ""))
                if desc:
                    acervo_parts.append(_trunc(desc, 50))
            roi = ed.get("roi_potential", {})
            rationale = _s(roi.get("reclassification_rationale", ""))
            if not acervo_parts and rationale:
                acervo_parts.append(_trunc(rationale, 80))
            if not acervo_parts:
                # Infer from object
                acervo_parts.append(f"Atestado: {_trunc(obj, 60)}")

            inv_rows.append([
                Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
                Paragraph(f"<b>{obj}</b><br/><font size='7' color='{TEXT_MUTED.hexval()}'>{orgao}</font>", styles["cell"]),
                Paragraph(valor, styles["cell_right"]),
                Paragraph("; ".join(acervo_parts[:2]), styles["cell"]),
            ])

        t = _three_rule_table(inv_rows, [avail * 0.06, avail * 0.36, avail * 0.14, avail * 0.44])
        el.append(t)
        el.append(Spacer(1, 3 * mm))

    # --- Portfólio Recomendado (optimal set from portfolio analysis) ---
    portfolio = data.get("portfolio", {})
    optimal_set_raw = portfolio.get("optimal_set", [])
    # Handle nested dict format: {"optimal_set": {"optimal_set": [...], ...}}
    if isinstance(optimal_set_raw, dict):
        optimal_set = optimal_set_raw.get("optimal_set", [])
    else:
        optimal_set = optimal_set_raw if isinstance(optimal_set_raw, list) else []
    capacity = portfolio.get("capacity", {})
    correlation = portfolio.get("correlation", {})

    if optimal_set:
        el.append(Spacer(1, 4 * mm))
        el.append(Paragraph("Portfólio Recomendado", styles["h3"]))

        # Capacity + diversification note
        cap_parts = []
        max_bids = capacity.get("max_simultaneous_bids")
        if max_bids:
            cap_parts.append(f"Capacidade estimada: {max_bids} participações simultâneas")
        div_score = correlation.get("diversification_score")
        if div_score is not None:
            cap_parts.append(f"Diversificação: {_dec(div_score)}/1,0")
        if cap_parts:
            el.append(Paragraph(
                " &nbsp;|&nbsp; ".join(cap_parts),
                ParagraphStyle(
                    "cap_note", parent=styles["body_small"],
                    fontName="Helvetica", fontSize=8, textColor=TEXT_SECONDARY,
                    spaceAfter=2 * mm,
                ),
            ))

        # Table
        opt_header = [
            Paragraph("<b>Prio.</b>", styles["cell_header_center"]),
            Paragraph("<b>Edital</b>", styles["cell_header"]),
            Paragraph("<b>Valor</b>", styles["cell_header_right"]),
            Paragraph("<b>Custo</b>", styles["cell_header_right"]),
            Paragraph("<b>Retorno</b>", styles["cell_header_right"]),
            Paragraph("<b>Acumulado</b>", styles["cell_header_right"]),
        ]
        opt_rows = [opt_header]
        for item in optimal_set:
            prio = item.get("priority", "")
            obj_resumo = _trunc(_s(item.get("objeto_resumo", "")), 50)
            valor = item.get("valor")
            custo = item.get("custo")
            roi_exp = item.get("roi_expected", 0)
            roi_cum = item.get("roi_cumulative", 0)

            roi_color = SIGNAL_GREEN if roi_exp >= 0 else SIGNAL_RED
            cum_color = SIGNAL_GREEN if roi_cum >= 0 else SIGNAL_RED

            opt_rows.append([
                Paragraph(f"<b>{prio}</b>", styles["cell_center"]),
                Paragraph(obj_resumo, styles["cell"]),
                Paragraph(_currency_short(valor), styles["cell_right"]),
                Paragraph(_currency_short(custo), styles["cell_right"]),
                Paragraph(
                    f"<b>{_currency_short(roi_exp)}</b>",
                    ParagraphStyle(f"roi_{prio}", parent=styles["cell_right"],
                                   fontName="Helvetica-Bold", textColor=roi_color),
                ),
                Paragraph(
                    f"<b>{_currency_short(roi_cum)}</b>",
                    ParagraphStyle(f"cum_{prio}", parent=styles["cell_right"],
                                   fontName="Helvetica-Bold", textColor=cum_color),
                ),
            ])

        t = _three_rule_table(
            opt_rows,
            [avail * 0.07, avail * 0.33, avail * 0.14, avail * 0.12, avail * 0.17, avail * 0.17],
        )
        el.append(t)

        # Correlation note
        corr_note = item.get("correlation_note") if optimal_set else None
        cap_warning = capacity.get("capacity_overflow_warning")
        if cap_warning:
            el.append(Paragraph(
                f"<i>{_s(cap_warning)}</i>",
                ParagraphStyle("cap_warn", parent=styles["caption"],
                               textColor=SIGNAL_AMBER, fontSize=7.5),
            ))
        el.append(Spacer(1, 3 * mm))

    elif portfolio and not optimal_set and (capacity or correlation):
        # Portfolio data exists but no optimal set
        el.append(Spacer(1, 4 * mm))
        el.append(Paragraph("Portfólio Recomendado", styles["h3"]))
        el.append(Paragraph(
            "Nenhum edital apresenta retorno esperado positivo neste ciclo. "
            "Participação representa investimento estratégico.",
            styles["body_small"],
        ))
        el.append(Spacer(1, 3 * mm))

    el.append(Spacer(1, 6 * mm))
    return el


# ============================================================
# E3: COVERAGE WARNING
# ============================================================

def _build_coverage_warning(data: dict, styles: dict) -> list:
    """Render coverage diagnostic warning if capture rate < 70%."""
    cov = data.get("coverage_diagnostic", {})
    if not cov:
        return []

    rate_raw = cov.get("coverage_rate")
    total = cov.get("total_estimated", 0)
    captured = cov.get("captured_count", 0)
    warning = cov.get("warning")

    # Suppress entirely when coverage data is invalid/missing/zero — avoid "258 de 0" display
    if rate_raw is None or total is None or total == 0:
        return []
    try:
        rate = float(rate_raw)
    except (ValueError, TypeError):
        return []
    if rate <= 0:
        return []

    if not warning and rate >= 0.70:
        return []  # Coverage is acceptable, render note in data sources instead

    el = []
    amber = colors.HexColor("#D4760A")

    el.append(Paragraph(
        f"<font color='{amber.hexval()}'><b>AVISO DE COBERTURA</b></font>",
        styles["h3"],
    ))

    el.append(Paragraph(
        f"Taxa de captura: <b>{_pct(rate)}</b> ({captured} de {total} editais estimados). "
        f"Este relatório pode não representar a totalidade das oportunidades disponíveis. "
        f"A análise abaixo deve ser interpretada com essa limitação.",
        styles["body"],
    ))

    # Per-UF breakdown
    per_uf = cov.get("per_uf", [])
    low_ufs = [p for p in per_uf if (p.get("rate") if p.get("rate") is not None else 0.0) < 0.70 and (p.get("estimated_total") or 0) > 0]
    if low_ufs:
        uf_detail = "; ".join(
            f"{p['uf']}: {p['captured']}/{p['estimated_total']} ({_pct(p['rate'])})"
            for p in low_ufs
        )
        el.append(Paragraph(f"UFs com baixa cobertura: {uf_detail}", styles["caption"]))

    el.append(Spacer(1, 6 * mm))
    return el


def _build_sector_divergence_alert(data: dict, styles: dict) -> list:
    """Render critical alert when company's CNAE diverges from its actual contract history.

    This is a top-level finding: a company registered for construction but whose
    700 contracts are all food/merchandise should see this BEFORE any edital analysis.
    """
    empresa = data.get("empresa", {})
    divergence = empresa.get("_sector_divergence")
    if not divergence:
        return []

    el = []
    total = divergence.get("total_contracts", 0)
    sector = divergence.get("sector_contracts", 0)

    el.append(Paragraph(
        f"<font color='{SIGNAL_RED.hexval()}'><b>ALERTA CRÍTICO — CNAE INCONSISTENTE COM HISTÓRICO</b></font>",
        ParagraphStyle(
            "sector_alert_title", parent=styles["h3"],
            fontName="Helvetica-Bold", textColor=SIGNAL_RED,
        ),
    ))

    el.append(Paragraph(
        f"A empresa possui <b>{total} contratos governamentais</b> registrados no "
        f"Portal Nacional de Contratações Públicas, porém "
        f"<b>{'nenhum' if sector == 0 else f'apenas {sector}'}</b> "
        f"está relacionado ao setor de atuação dos editais analisados neste relatório.",
        styles["body"],
    ))

    el.append(Paragraph(
        "O CNAE registrado indica capacidade formal para atuar no setor, mas o histórico de "
        "contratos revela atuação predominante em segmento distinto. Isso significa que:",
        styles["body"],
    ))

    implications = [
        "A empresa provavelmente <b>não possui atestados de capacidade técnica</b> no setor — "
        "requisito eliminatório na fase de habilitação da maioria das licitações",
        "A <b>Certidão de Acervo Técnico (CAT)</b> junto ao CREA/CAU pode estar vazia para este tipo de obra",
        "Todas as recomendações deste relatório devem ser lidas com esta ressalva: "
        "<b>a participação efetiva depende de verificação prévia do acervo real da empresa</b>",
    ]
    for imp in implications:
        el.append(Paragraph(
            f"— {imp}",
            ParagraphStyle(
                "sector_imp", parent=styles["bullet"],
                fontName="Times-Roman", textColor=TEXT_COLOR,
            ),
        ))

    el.append(Spacer(1, 2 * mm))
    el.append(Paragraph(
        "<b>Ação imediata recomendada:</b> Confirmar junto à empresa se existem contratos no setor "
        "não registrados no PNCP (contratos estaduais/municipais anteriores a 2021, contratos privados). "
        "Se confirmada a ausência de acervo, consultar a seção \"Plano de Desenvolvimento\" para "
        "o roteiro de construção de acervo técnico.",
        ParagraphStyle(
            "sector_action", parent=styles["body"],
            fontName="Helvetica", fontSize=9, textColor=INK,
        ),
    ))

    el.append(Spacer(1, 8 * mm))
    return el


# ============================================================
# E7: REGIONAL CLUSTER ANALYSIS
# ============================================================

def _build_regional_analysis(data: dict, styles: dict, sec: dict | None = None) -> list:
    """Render regional portfolio analysis with geographic clusters."""
    clusters_data = data.get("regional_clusters", {})
    clusters = clusters_data.get("clusters", [])
    if not clusters:
        return []

    el = []
    num = sec["next"]() if sec else 9
    el.extend(_section_heading(f"{num}. Análise Regional de Portfólio", styles))

    el.append(Paragraph(
        "Editais compatíveis agrupados por proximidade geográfica (raio de 150km). "
        "Clusters identificam oportunidades de mobilização compartilhada — "
        "uma única estrutura operacional servindo múltiplas frentes.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    editais = data.get("editais", [])

    for cl in clusters:
        center = f"{_s(cl.get('center_municipio', ''))}/{_s(cl.get('center_uf', ''))}"
        n = cl.get("n_editais", 0)
        radius = cl.get("radius_km", 0)
        total_valor = cl.get("total_valor", 0)
        timeline = "sim" if cl.get("timeline_overlap") else "não"

        el.append(Paragraph(
            f"<b>Cluster {cl.get('id', '')}: {center}</b> — "
            f"{n} editais, raio {radius}km, valor total {_currency(total_valor)}",
            styles["body"],
        ))
        el.append(Paragraph(
            f"Sobreposição de prazos: {timeline}",
            styles["body_small"],
        ))

        # List editais in cluster with viability analysis
        indices = cl.get("editais_indices", [])
        cluster_editais = []
        cluster_roi_sum = 0
        cluster_mobilization_cost = 0
        for idx in indices:
            if idx < len(editais):
                ed = editais[idx]
                cluster_editais.append(ed)
                obj = _trunc(_s(ed.get("objeto") or ""), 150)
                mun = _s(ed.get("municipio", ""))
                valor = _currency_short(_safe_float(ed.get("valor_estimado", 0)))
                roi = ed.get("roi_potential", {})
                roi_max = _safe_float(roi.get("roi_max", 0))
                cluster_roi_sum += roi_max
                participation_cost = roi.get("calculation_memory", {}).get("custo_participacao", 0)
                cluster_mobilization_cost += _safe_float(participation_cost)
                el.append(Paragraph(
                    f"  — {obj} ({mun}) — {valor}",
                    styles["body_small"],
                ))

        # Joint viability analysis
        if cluster_editais:
            # Check if cluster is primarily strategic (acervo) investments
            n_cluster = len(cluster_editais)
            n_acervo = sum(
                1 for ed in cluster_editais
                if (ed.get("roi_potential") or {}).get("strategic_reclassification")
                == "INVESTIMENTO_ESTRATEGICO_ACERVO"
            )

            # Shared mobilization savings: only 1 trip, not N
            shared_savings = cluster_mobilization_cost * (1 - 1.0 / max(n_cluster, 1))
            net_roi_cluster = cluster_roi_sum + shared_savings

            if n_acervo == n_cluster:
                # ALL editais are strategic acervo → financial ROI is irrelevant
                viability = "ESTRATÉGICO"
                viab_color = ACCENT  # Bronze — neither green nor red
                el.append(Paragraph(
                    f"<b>Viabilidade conjunta:</b> <font color='{viab_color.hexval()}'>"
                    f"<b>{viability}</b></font>"
                    f" — Cluster de investimento em acervo técnico"
                    f" ({n_cluster} editais, mobilização compartilhada: {_currency_short(shared_savings)})",
                    styles["body_small"],
                ))
            elif n_acervo > 0:
                # Mixed: some acervo, some financial
                n_financial = n_cluster - n_acervo
                viability = "VIÁVEL" if net_roi_cluster > 0 else "MISTO"
                viab_color = SIGNAL_GREEN if net_roi_cluster > 0 else SIGNAL_AMBER
                el.append(Paragraph(
                    f"<b>Viabilidade conjunta:</b> <font color='{viab_color.hexval()}'>"
                    f"<b>{viability}</b></font>"
                    f" — {n_financial} edital(is) com retorno financeiro"
                    f" + {n_acervo} investimento(s) em acervo"
                    f" (economia de mobilização: {_currency_short(shared_savings)})",
                    styles["body_small"],
                ))
            else:
                # Pure financial analysis
                viability = "VIÁVEL" if net_roi_cluster > 0 else "INVIÁVEL"
                viab_color = SIGNAL_GREEN if net_roi_cluster > 0 else SIGNAL_RED
                el.append(Paragraph(
                    f"<b>Viabilidade conjunta:</b> <font color='{viab_color.hexval()}'>"
                    f"<b>{viability}</b></font>"
                    f" — Resultado conjunto: {_currency_short(net_roi_cluster)}"
                    f" (economia de mobilização compartilhada: {_currency_short(shared_savings)})",
                    styles["body_small"],
                ))

        rec = _s(cl.get("recommendation", ""))
        if rec:
            el.append(Paragraph(f"<i>{rec}</i>", styles["caption"]))
        el.append(Spacer(1, 3 * mm))

    if len(clusters) >= 2:
        el.append(Paragraph(
            "A decisão não é participar de cada edital individualmente — "
            "é decidir se vale estabelecer presença operacional em cada micro-região.",
            styles["body_small"],
        ))

    el.append(Spacer(1, 6 * mm))
    return el


def _build_development_plan(data: dict, styles: dict, sec: dict | None = None) -> list:
    """E4: Consolidated development plan aggregating operational gaps across all editais."""
    editais = data.get("editais", [])
    all_gaps = []
    for ed in editais:
        qual_gap = ed.get("qualification_gap", {})
        for gap in qual_gap.get("operational_gaps", []):
            gap_copy = dict(gap)
            gap_copy["edital_objeto"] = _trunc(_s(ed.get("objeto", "")), 60)
            all_gaps.append(gap_copy)

    if not all_gaps:
        return []

    # Deduplicate by (gap_type, description) — keep first occurrence
    seen = set()
    unique_gaps = []
    for g in all_gaps:
        key = (g.get("gap_type", ""), g.get("description", ""))
        if key not in seen:
            seen.add(key)
            unique_gaps.append(g)

    if not unique_gaps:
        return []

    el = []
    num = sec["next"]() if sec else 10
    el.extend(_section_heading(f"{num}. Plano de Desenvolvimento (12 meses)", styles))

    el.append(Paragraph(
        "Consolidação das lacunas operacionais identificadas nos editais analisados. "
        "Cada item representa uma capacidade que, uma vez construída, abre acesso a novos mercados.",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN
    header = [
        Paragraph("Tipo", styles["cell_header"]),
        Paragraph("Descrição", styles["cell_header"]),
        Paragraph("Prazo", styles["cell_header_center"]),
        Paragraph("Ação Necessária", styles["cell_header"]),
    ]
    rows = [header]

    for g in unique_gaps:
        rows.append([
            Paragraph(_s(g.get("gap_type", "")), styles["cell"]),
            Paragraph(_s(g.get("description", "")), styles["cell"]),
            Paragraph(_s(g.get("estimated_timeline", "")), styles["cell_center"]),
            Paragraph(_s(g.get("action_required", "")), styles["cell"]),
        ])

    t = _three_rule_table(rows, [
        avail * 0.15, avail * 0.30, avail * 0.15, avail * 0.40,
    ])
    el.append(t)
    el.append(Spacer(1, 6 * mm))
    return el


# ============================================================
# E10: REPORT VALIDATION
# ============================================================

def validate_report_completeness(data: dict) -> tuple[list[str], list[str]]:
    """Validate report data completeness.

    Returns (blocking_errors, warnings).
    Blocking errors prevent PDF generation — the JSON must be re-enriched first.
    Warnings are informational (printed to console but don't block).

    CRÍTICA 10A: Enhanced consistency validation — detects:
    - Recommendations inconsistent with risk_score
    - Alerts identified but not reflected in score
    - Empty sections that should have content
    """
    errors = []
    warnings = []

    editais = data.get("editais", [])

    for i, ed in enumerate(editais, 1):
        idx = ed.get("_display_idx", i)
        obj = _trunc(_s(ed.get("objeto", "")), 50)

        # BLOCKING: Every edital must have recommendation + justification
        rec = ed.get("recomendacao") or ed.get("recommendation")
        justif = ed.get("justificativa") or ed.get("recommendation_justification")
        if not justif and rec:
            errors.append(f"Edital {idx} ({obj}): recomendação '{rec}' sem justificativa")

        # BLOCKING: risk_score required for viability section
        rs = ed.get("risk_score", {})
        if not rs:
            errors.append(f"Edital {idx} ({obj}): risk_score ausente — execute --re-enrich")

        # BLOCKING: win_probability required for incumbency/competitive sections
        if not ed.get("win_probability"):
            errors.append(f"Edital {idx} ({obj}): win_probability ausente — execute --re-enrich")

        # BLOCKING: strategic_category required for portfolio matrix
        if not ed.get("strategic_category"):
            errors.append(f"Edital {i} ({obj}): strategic_category ausente — execute --re-enrich")

        # CRÍTICA 10A: Consistency checks — recommendation vs score
        rec_upper = (rec or "").upper()
        score_total = rs.get("total", -1) if isinstance(rs, dict) else -1
        vetoed = rs.get("vetoed", False) if isinstance(rs, dict) else False

        if vetoed and rec_upper == "PARTICIPAR":
            errors.append(
                f"Edital {i} ({obj}): PARTICIPAR mas edital VETADO — "
                f"motivo: {', '.join(rs.get('veto_reasons', ['?']))}"
            )

        if score_total >= 0 and score_total > 60 and rec_upper == "NÃO RECOMENDADO" and not vetoed:
            warnings.append(
                f"Edital {i} ({obj}): score alto ({score_total}) mas NÃO RECOMENDADO — "
                f"verificar justificativa"
            )

        if score_total >= 0 and score_total < 30 and rec_upper == "PARTICIPAR" and not vetoed:
            warnings.append(
                f"Edital {i} ({obj}): score baixo ({score_total}) mas PARTICIPAR — "
                f"verificar justificativa"
            )

        # WARNING: ROI calculation memory (desirable but not blocking)
        roi = ed.get("roi_potential", {})
        if roi and not roi.get("calculation_memory"):
            warnings.append(f"Edital {i}: ROI sem memória de cálculo (auditabilidade reduzida)")

        # WARNING: organ_risk (desirable)
        if not ed.get("organ_risk"):
            warnings.append(f"Edital {i}: organ_risk ausente (seção de risco do órgão incompleta)")

        # CRÍTICA 10A: Fiscal risk alerts identified but not in justificativa
        fiscal = rs.get("fiscal_risk", {}) if isinstance(rs, dict) else {}
        if isinstance(fiscal, dict) and fiscal.get("nivel") == "ALTO" and justif:
            if "fiscal" not in (justif or "").lower() and "inadimplência" not in (justif or "").lower():
                warnings.append(
                    f"Edital {i}: risco fiscal ALTO identificado mas não mencionado na justificativa"
                )

    # BLOCKING: top-level computed fields
    if not data.get("maturity_profile") and not data.get("empresa", {}).get("maturity_profile"):
        errors.append("maturity_profile ausente — execute --re-enrich no JSON")

    if not data.get("dispute_stats"):
        errors.append("dispute_stats ausente — execute --re-enrich no JSON")

    # WARNING: coverage_diagnostic (desirable)
    if not data.get("coverage_diagnostic"):
        warnings.append("Diagnóstico de cobertura ausente — seção E3 será omitida")

    # WARNING: regional_clusters (desirable)
    if not data.get("regional_clusters"):
        warnings.append("regional_clusters ausente — seção E7 será omitida")

    # WARNING: SICAF must not be UNAVAILABLE (E2)
    sicaf = data.get("sicaf", {})
    src = sicaf.get("_source", {})
    if isinstance(src, dict) and src.get("status") == "UNAVAILABLE":
        warnings.append("SICAF marcado como UNAVAILABLE — deve ser FALHA_COLETA ou coletado")

    # CROSS-REFERENCE: Detect contradictions between sections
    # If next_steps or prioritization reference descartados, flag it
    descartados_objs = set()
    for ed in editais:
        rec_norm = _normalize_recommendation(ed.get("recomendacao", ""))
        if rec_norm in ("NÃO RECOMENDADO", "DESCARTADO"):
            obj_trunc = _trunc(_s(ed.get("objeto", "")), 40).lower()
            mun = _s(ed.get("municipio", "")).lower()
            if obj_trunc:
                descartados_objs.add(obj_trunc)
            if mun:
                descartados_objs.add(mun)

    # Check PARTICIPAR with score < 20 (likely error)
    for ed in editais:
        idx = ed.get("_display_idx", 0)
        rec_norm = _normalize_recommendation(ed.get("recomendacao", ""))
        rs = ed.get("risk_score", {})
        total = rs.get("total", -1) if isinstance(rs, dict) else -1
        if rec_norm == "PARTICIPAR" and total >= 0 and total < 20:
            errors.append(
                f"Edital {idx} ({_trunc(_s(ed.get('objeto', '')), 50)}): "
                f"PARTICIPAR com score {total} — provável erro de classificação"
            )

    return errors, warnings


# ============================================================
# ANNEXES (3-layer architecture: Camada 3)
# ============================================================

def _build_annex_nao_recomendado(data: dict, styles: dict, sec: dict) -> list:
    """Annex A: Condensed table of non-recommended + vetoed editais."""
    editais = data.get("editais", [])
    nr_editais = [
        (ed.get("_display_idx", i), ed) for i, ed in enumerate(editais, 1)
        if _normalize_recommendation(_s(ed.get("recomendacao", ""))) == "NÃO RECOMENDADO"
        or (ed.get("risk_score") or {}).get("vetoed", False)
    ]
    if not nr_editais:
        return []

    el = []
    el.append(PageBreak())
    num = sec["next"]()
    # Count vetoed separately for clarity in heading
    n_vetados = sum(1 for _, ed in nr_editais if (ed.get("risk_score") or {}).get("vetoed", False))
    heading_detail = f"{len(nr_editais)}"
    if n_vetados:
        heading_detail += f", sendo {n_vetados} eliminatório(s)"
    el.extend(_section_heading(f"Anexo A — Editais Não Recomendados ({heading_detail})", styles))

    avail = PAGE_WIDTH - 2 * MARGIN
    header = [
        Paragraph("#", styles["cell_header_center"]),
        Paragraph("Município / Objeto", styles["cell_header"]),
        Paragraph("Valor", styles["cell_header_right"]),
        Paragraph("Motivo", styles["cell_header"]),
    ]
    rows = [header]
    for idx, ed in nr_editais:
        mun = _s(ed.get("municipio", ""))
        uf = _s(ed.get("uf", ""))
        obj = _trunc(_s(ed.get("objeto", "")), 120)
        valor = _currency_short(ed.get("valor_estimado"))
        link = _fix_pncp_link(ed.get("link", ""))

        # For vetoed editais, show veto reason as motivo (more specific than justificativa)
        risk = ed.get("risk_score", {}) or {}
        if risk.get("vetoed"):
            veto_reasons = risk.get("veto_reasons", [])
            justif = "ELIMINATÓRIO: " + "; ".join(veto_reasons) if veto_reasons else "Impedimento legal"
            justif = _trunc(justif, 100)
        else:
            justif = _trunc(_s(ed.get("justificativa", "")), 80)

        # Make object clickable
        obj_text = f'{mun}/{uf} — {obj}'
        if link and link.startswith("http"):
            obj_text = f'<a href="{link}" color="{TEXT_SECONDARY.hexval()}">{obj_text}</a>'

        rows.append([
            Paragraph(f"<b>{idx}</b>", styles["cell_center"]),
            Paragraph(obj_text, styles["cell"]),
            Paragraph(valor, styles["cell_right"]),
            Paragraph(justif, styles["cell"]),
        ])

    t = _three_rule_table(rows, [avail * 0.06, avail * 0.34, avail * 0.12, avail * 0.48])
    el.append(t)
    el.append(Spacer(1, 8 * mm))
    return el


def _build_annex_company(data: dict, styles: dict, sec: dict) -> list:
    """Annex B: Company profile + SICAF (reference material)."""
    el = []
    el.append(PageBreak())
    num = sec["next"]()
    el.extend(_section_heading(f"Anexo B — Perfil da Empresa", styles))
    el.extend(_build_company_profile_content(data, styles))
    el.append(Spacer(1, 5 * mm))
    el.extend(_build_sicaf_content(data, styles))
    return el


def _build_methodology_content(styles: dict) -> list:
    """Render methodology subsection for Annex C."""
    el = []
    el.append(Paragraph("<b>Metodologia de Análise</b>", styles["h3"]))

    el.append(Paragraph(
        "O índice de viabilidade combina cinco dimensões com pesos calibrados "
        "para refletir os fatores mais determinantes na decisão de participação:",
        styles["body_small"],
    ))
    el.append(Spacer(1, 3 * mm))

    avail = PAGE_WIDTH - 2 * MARGIN
    header = [
        Paragraph("Dimensão", styles["cell_header"]),
        Paragraph("Peso", styles["cell_header_center"]),
        Paragraph("O que avalia", styles["cell_header"]),
    ]
    rows = [header]
    dimensions = [
        ("Habilitação", "30%", "Capacidade técnica, atestados, capital mínimo, certidões"),
        ("Financeiro", "25%", "Valor do edital vs. capacidade da empresa, regime tributário"),
        ("Geográfico", "20%", "Distância rodoviária até o local de execução"),
        ("Prazo", "15%", "Dias restantes para preparação da proposta"),
        ("Competitivo", "10%", "Histórico de fornecedores do órgão, concentração de mercado"),
    ]
    for dim, peso, desc in dimensions:
        rows.append([
            Paragraph(dim, styles["cell"]),
            Paragraph(peso, styles["cell_center"]),
            Paragraph(desc, styles["cell"]),
        ])
    t = _three_rule_table(rows, [avail * 0.18, avail * 0.12, avail * 0.70])
    el.append(t)
    el.append(Spacer(1, 5 * mm))

    # ROI Formula
    el.append(Paragraph("<b>Resultado Potencial (ROI)</b>", styles["h3"]))
    el.append(Paragraph(
        "O resultado potencial de cada edital é calculado pela fórmula:",
        styles["body_small"],
    ))
    el.append(Spacer(1, 2 * mm))
    roi_formula = ParagraphStyle(
        "roi_formula", parent=styles["body"],
        fontName="Helvetica-Bold", fontSize=9, textColor=INK,
        alignment=TA_CENTER, spaceBefore=2 * mm, spaceAfter=2 * mm,
    )
    el.append(Paragraph(
        "(Valor do edital × Probabilidade de vitória × Margem líquida setorial) "
        "− Custo estimado de participação",
        roi_formula,
    ))
    el.append(Spacer(1, 3 * mm))

    # Probability calibration
    el.append(Paragraph("<b>Calibração da Probabilidade</b>", styles["h3"]))
    el.append(Paragraph(
        "A probabilidade de vitória é calculada com base no histórico de contratos "
        "do órgão licitante obtido via Portal Nacional de Contratações Públicas. "
        "O nível de confiança da estimativa varia conforme a amostra disponível:",
        styles["body_small"],
    ))
    el.append(Spacer(1, 2 * mm))
    conf_header = [
        Paragraph("Amostra", styles["cell_header"]),
        Paragraph("Confiança", styles["cell_header"]),
    ]
    conf_rows = [conf_header]
    conf_data = [
        ("Mais de 20 contratos", "Alta"),
        ("5 a 20 contratos", "Média"),
        ("Menos de 5 contratos", "Baixa"),
    ]
    for amostra, confianca in conf_data:
        conf_rows.append([
            Paragraph(amostra, styles["cell"]),
            Paragraph(confianca, styles["cell"]),
        ])
    t2 = _three_rule_table(conf_rows, [avail * 0.50, avail * 0.50])
    el.append(t2)
    el.append(Spacer(1, 5 * mm))

    # Disclaimer
    disclaimer_style = ParagraphStyle(
        "disclaimer_meth", parent=styles["body_small"],
        fontName="Helvetica-Oblique", fontSize=8, textColor=TEXT_MUTED,
        alignment=TA_JUSTIFY,
    )
    el.append(Paragraph(
        "Este relatório tem caráter informativo e não substitui análise jurídica do edital. "
        "Probabilidades são estimativas baseadas em dados históricos e não representam "
        "garantia de resultado.",
        disclaimer_style,
    ))
    return el


def _build_annex_sources(data: dict, styles: dict, sec: dict) -> list:
    """Annex C: Data sources + methodology + gazette mentions + audit trail."""
    el = []
    el.append(PageBreak())
    num = sec["next"]()
    el.extend(_section_heading(f"Anexo C — Fontes de Dados e Metodologia", styles))
    el.extend(_build_data_sources_content(data, styles))
    el.append(Spacer(1, 5 * mm))
    # Methodology subsection
    el.extend(_build_methodology_content(styles))
    el.append(Spacer(1, 5 * mm))
    # Querido Diário — only if there are mentions
    mencoes = data.get("querido_diario", [])
    if mencoes:
        el.append(Paragraph("<b>Menções em Diários Oficiais</b>", styles["h3"]))
        el.extend(_build_querido_diario_content(data, styles))

    # F36: Audit trail
    import subprocess
    import hashlib
    try:
        git_hash = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=str(Path(__file__).parent.parent),
            stderr=subprocess.DEVNULL,
        ).decode().strip()
    except Exception:
        git_hash = "N/D"

    gen_ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    json_hash = hashlib.sha256(
        json.dumps(data, sort_keys=True, ensure_ascii=False).encode()
    ).hexdigest()[:12]

    audit_text = (
        f"Gerado em {gen_ts} | Script v.{git_hash} | "
        f"Dados SHA-256: {json_hash}"
    )
    el.append(Spacer(1, 5 * mm))
    el.append(Paragraph(
        _s(audit_text, restore_accents=False),
        ParagraphStyle(
            "audit_trail", parent=styles["caption"],
            fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED,
        ),
    ))

    return el


# ============================================================
# MAIN
# ============================================================

def _sanitize_links(data: dict) -> dict:
    editais = data.get("editais", [])
    fixed = 0
    for ed in editais:
        original = ed.get("link", "")
        corrected = _fix_pncp_link(original)
        if corrected != original:
            ed["link"] = corrected
            fixed += 1
    if fixed:
        print(f"Links corrected: {fixed}/{len(editais)}")
    return data


def _backfill_recommendations(data: dict) -> None:
    """Derive recomendacao/justificativa from risk_score if missing.

    Supports JSONs generated by older collector versions that didn't
    populate these fields.
    """
    editais = data.get("editais", [])
    backfilled = 0
    for ed in editais:
        if ed.get("recomendacao"):
            continue  # already set
        rs = ed.get("risk_score", {})
        total = rs.get("total", 0)
        vetoed = rs.get("vetoed", False)
        veto_reasons = rs.get("veto_reasons", [])

        if vetoed:
            ed["recomendacao"] = "NÃO RECOMENDADO"
            ed["justificativa"] = "; ".join(veto_reasons) if veto_reasons else "Edital vetado por impedimento legal."
        elif total >= 70:
            ed["recomendacao"] = "PARTICIPAR"
        elif total >= 40:
            ed["recomendacao"] = "AVALIAR COM CAUTELA"
        else:
            ed["recomendacao"] = "NÃO RECOMENDADO"

        # Build justificativa from score components
        if not ed.get("justificativa"):
            parts = []
            hab = rs.get("habilitacao", 0)
            fin = rs.get("financeiro", 0)
            geo = rs.get("geografico", 0)
            prazo = rs.get("prazo", 0)
            if hab >= 80:
                parts.append("Habilitação compatível")
            elif hab < 40:
                parts.append("Risco de inabilitação")
            if fin >= 80:
                parts.append("valor adequado ao porte")
            elif fin < 40:
                parts.append("valor acima da capacidade financeira")
            if geo >= 60:
                parts.append("proximidade geográfica favorável")
            elif geo < 20:
                parts.append("distância geográfica desfavorável")
            if prazo >= 80:
                parts.append("prazo confortável")
            elif prazo < 30:
                parts.append("prazo insuficiente")
            ed["justificativa"] = ". ".join(parts) + "." if parts else "Análise baseada em scoring multifatorial."

        backfilled += 1

    if backfilled:
        print(f"  Backfill: {backfilled} editais sem recomendação — derivadas de risk_score")


def generate_report_b2g(data: dict) -> BytesIO:
    """Generate the full B2G report PDF from structured data.

    Raises ValueError if blocking validation errors are found
    (e.g. recommendations without justificativa).
    """
    warnings, errors = _validate_json(data)
    if errors:
        raise ValueError(
            f"Geração do PDF bloqueada — {len(errors)} edital(is) com recomendação "
            f"sem justificativa. Preencha 'justificativa' em cada edital antes de gerar.\n"
            + "\n".join(f"  - {e}" for e in errors)
        )

    # Per-report state (passed to functions that need first-time-only rendering)
    _state = {"viability_shown": False, "roi_shown": False}

    data = _sanitize_links(data)

    # Backfill recommendations if missing (for JSONs generated by older collector)
    _backfill_recommendations(data)

    # Drop ENCERRADO editais + missing status + explicitly irrelevant + expired
    def _is_eligible(e: dict) -> bool:
        status = e.get("status_edital")
        if status in ("ENCERRADO", None, ""):
            return False
        if e.get("relevante") is False:
            return False
        dias = e.get("dias_restantes")
        if dias is not None and _safe_int(dias) <= 0:
            return False
        return True
    data["editais"] = [e for e in data.get("editais", []) if _is_eligible(e)]

    # Separate discarded editais (irrelevant to company profile) before rendering
    all_editais = data.get("editais", [])
    descartados = [e for e in all_editais if e.get("recomendacao", "").upper() == "DESCARTADO"]
    data["editais"] = [e for e in all_editais if e.get("recomendacao", "").upper() != "DESCARTADO"]
    data["_descartados_count"] = len(descartados)
    data["_descartados_motivos"] = _summarize_discard_reasons(descartados)
    if descartados:
        print(f"  Descartados: {len(descartados)} editais sem aderência ao perfil (excluídos do relatório)")
    print(f"  Relatório: {len(data['editais'])} editais incluídos")

    # CANONICAL SORT: establish a single, stable ordering for the entire report.
    # All sections reference editais by _display_idx instead of enumerate() position.
    # Order: PARTICIPAR first (by score desc), then AVALIAR, then NÃO RECOMENDADO.
    data["editais"] = sorted(
        data["editais"],
        key=lambda e: (
            REC_ORDER.get(_normalize_recommendation(_s(e.get("recomendacao", ""))), 9),
            -(_safe_int((e.get("risk_score") or {}).get("total", 0))),
        ),
    )
    for _i, _ed in enumerate(data["editais"], 1):
        _ed["_display_idx"] = _i

    gen_date = _today()
    styles = _build_styles()
    buffer = BytesIO()

    empresa = data.get("empresa", {})
    nome = _s(empresa.get("nome_fantasia") or empresa.get("razao_social", "Empresa"))
    cnpj = _s(empresa.get("cnpj", ""), restore_accents=False)

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=MARGIN + 10 * mm,
        title=f"Relatório B2G — {nome} — {gen_date}",
        author="Tiago Sasaki",
        creator="Report B2G Generator",
    )

    sec = _section_counter()
    elements: list = []
    elements.extend(_build_cover(data, styles, gen_date))

    # === CRÍTICA 10B: RESUMO DECISÓRIO (page 2 — immediate value delivery) ===
    elements.extend(_build_resumo_decisorio(data, styles))

    # === CAMADA 1: DECISÃO EXECUTIVA ===
    # Coverage warning (before any analysis if coverage < 70%)
    elements.extend(_build_coverage_warning(data, styles))
    # Sector divergence alert (CNAE vs actual contracts — before any edital analysis)
    elements.extend(_build_sector_divergence_alert(data, styles))
    # 1. Resumo Executivo (condensed)
    elements.extend(_build_executive_summary(data, styles, sec))
    # 2. Posicionamento Estratégico (thesis + signals)
    elements.extend(_build_strategic_positioning(data, styles, sec))
    # 3. Decisão em 30 Segundos (grouped summary table)
    elements.extend(_build_decision_table(data, styles, sec))

    # === CAMADA 2: INTELIGÊNCIA ESTRATÉGICA ===
    # 3. Inteligência Exclusiva — 4 differentials PNCP can't provide
    elements.extend(_build_exclusive_intelligence(data, styles, sec))
    # 4. Análise Detalhada (ONLY PARTICIPAR + AVALIAR — NÃO RECOMENDADO goes to Annex A)
    elements.extend(_build_detailed_analysis(data, styles, sec, _state))
    # 5. Matriz Estratégica + Regional (unified strategic view)
    elements.extend(_build_portfolio_section(data, styles, sec))
    elements.extend(_build_regional_analysis(data, styles, sec))
    # 6. Mapa Competitivo (condensed incumbency deep-dive)
    elements.extend(_build_competitive_section(data, styles, sec))
    # 7. Market intelligence
    elements.extend(_build_market_intelligence(data, styles, sec))
    # 8. Development plan
    elements.extend(_build_development_plan(data, styles, sec))
    # 9. Consolidated prioritization of PARTICIPAR editais
    elements.extend(_build_prioritization(data, styles, sec))
    # 10. Próximos Passos
    elements.extend(_build_next_steps(data, styles, sec))

    # === CAMADA 3: ANEXOS ===
    # Annex A: NÃO RECOMENDADO editais (condensed table)
    elements.extend(_build_annex_nao_recomendado(data, styles, sec))
    # Annex B: Company Profile + SICAF
    elements.extend(_build_annex_company(data, styles, sec))
    # Annex C: Data Sources + Querido Diário
    elements.extend(_build_annex_sources(data, styles, sec))

    # E10: Validate report completeness (blocking errors + warnings)
    validation_errors, validation_warnings = validate_report_completeness(data)
    if validation_errors:
        print(f"\n  BLOQUEIO: {len(validation_errors)} campo(s) critico(s) ausente(s)")
        for e in validation_errors[:10]:
            print(f"    - {e}")
        if len(validation_errors) > 10:
            print(f"    ... e mais {len(validation_errors) - 10}")
        print(f"\n  Para corrigir, execute:")
        print(f"    python scripts/collect-report-data.py --re-enrich <caminho-do-json>")
        raise ValueError(
            f"PDF bloqueado — {len(validation_errors)} campo(s) crítico(s) ausente(s) no JSON. "
            f"Execute --re-enrich para recalcular campos determinísticos sem re-coletar APIs.\n"
            + "\n".join(f"  - {e}" for e in validation_errors[:15])
        )
    if validation_warnings:
        print(f"  Validacao: {len(validation_warnings)} aviso(s) — seções opcionais podem estar incompletas")
        for w in validation_warnings:
            print(f"    - {w}")

    # F38: LayoutError graceful handling
    try:
        doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer,
                  canvasmaker=_NumberedCanvas)
    except Exception as exc:
        if "Flowable" in str(exc) or "too large" in str(exc).lower():
            print(f"WARNING: LayoutError detected. Retrying with reduced content...")
            # Truncate oversized flowables as safety net
            for flowable in elements:
                if hasattr(flowable, 'text') and len(getattr(flowable, 'text', '')) > 2000:
                    flowable.text = flowable.text[:2000] + "..."
            buffer2 = BytesIO()
            doc2 = SimpleDocTemplate(
                buffer2, pagesize=A4,
                leftMargin=MARGIN, rightMargin=MARGIN,
                topMargin=MARGIN, bottomMargin=MARGIN + 10 * mm,
            )
            doc2.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer,
                       canvasmaker=_NumberedCanvas)
            print("PDF generated with reduced content (LayoutError recovery)")
            buffer2.seek(0)
            return buffer2
        else:
            raise
    buffer.seek(0)
    return buffer



def _clean_excel_str(s: str) -> str:
    """Remove illegal XML characters that openpyxl rejects."""
    import re
    if not isinstance(s, str):
        return str(s) if s is not None else ""
    # Remove control chars except tab, newline, carriage return
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)


def generate_excel_companion(data: dict, output_path: str) -> None:
    """Generate Excel with all editais for filtering/sorting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Oportunidades"

    # Headers
    headers = ["#", "Recomendação", "Score", "Município", "UF", "Objeto",
               "Valor Estimado", "Modalidade", "Prazo (dias)", "Probabilidade (%)",
               "ROI Mín (R$)", "ROI Máx (R$)", "Distância (km)", "Cluster",
               "Justificativa", "Link PNCP"]

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Color fills for recommendations
    fills = {
        "PARTICIPAR": PatternFill(start_color="D5F5E3", fill_type="solid"),          # green
        "AVALIAR COM CAUTELA": PatternFill(start_color="FEF9E7", fill_type="solid"),  # amber
        "NÃO RECOMENDADO": PatternFill(start_color="FADBD8", fill_type="solid"),      # red
    }

    editais = data.get("editais", [])
    # Use the SAME canonical order as PDF (already sorted by recommendation type -> score desc)
    # Do NOT re-sort — _display_idx must match row position in both PDF and Excel
    editais_sorted = editais

    for excel_row, ed in enumerate(editais_sorted, 1):
        row = excel_row + 1
        rs = ed.get("risk_score", {})
        roi = ed.get("roi_potential", {})
        dist = ed.get("distancia", {})
        dist_km = dist.get("distancia_km") if isinstance(dist, dict) else None
        wp = ed.get("win_probability", {})
        prob = (wp.get("probability") or wp.get("probabilidade", 0)) if isinstance(wp, dict) else None
        rec = ed.get("recomendacao", "")

        values = [
            ed.get("_display_idx", excel_row),
            _clean_excel_str(rec or "N/A"),
            rs.get("total", 0),
            _clean_excel_str(ed.get("municipio", "")),
            _clean_excel_str(ed.get("uf", "")),
            _clean_excel_str(ed.get("objeto", "")),
            ed.get("valor_estimado", 0),
            _clean_excel_str(ed.get("modalidade", "")),
            ed.get("dias_restantes", ""),
            round(prob * 100, 1) if prob else "",
            roi.get("roi_min", "") if isinstance(roi, dict) else "",
            roi.get("roi_max", "") if isinstance(roi, dict) else "",
            round(dist_km, 0) if dist_km else "",
            _clean_excel_str(ed.get("_cluster_origin", "")),
            _clean_excel_str(ed.get("justificativa", "")),
            _clean_excel_str(ed.get("link", "")),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            # Apply recommendation color
            if rec in fills:
                cell.fill = fills[rec]

        # Make Link PNCP column (16) a clickable hyperlink
        link_val = ed.get("link", "")
        link_fixed = _fix_pncp_link(link_val)
        if link_fixed and link_fixed.startswith("http"):
            link_cell = ws.cell(row=row, column=16)
            link_cell.hyperlink = link_fixed
            link_cell.value = link_fixed
            link_cell.font = Font(color="0563C1", underline="single", size=10)

        # Format currency column (7 = Valor Estimado)
        ws.cell(row=row, column=7).number_format = '#,##0.00'

    # Column widths
    widths = [5, 22, 8, 20, 5, 60, 15, 25, 10, 12, 12, 12, 10, 25, 50, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(editais_sorted) + 1}"

    wb.save(output_path)

def main():
    parser = argparse.ArgumentParser(description="Generate B2G Report PDF from JSON data")
    parser.add_argument("--input", required=True, help="Path to JSON data file")
    parser.add_argument("--output", help="Output PDF path (auto-generated if omitted)")
    parser.add_argument("--save-json", action="store_true", help="Save enriched JSON (with backfilled recommendations) back to input file")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if args.output:
        output_path = Path(args.output)
    else:
        cnpj = data.get("empresa", {}).get("cnpj", "unknown").replace("/", "").replace(".", "").replace("-", "")
        nome = data.get("empresa", {}).get("nome_fantasia") or data.get("empresa", {}).get("razao_social", "")
        nome_slug = re.sub(r"[^a-z0-9]+", "-", nome.lower().strip()).strip("-")[:40] if nome else ""
        date_str = datetime.now().strftime("%Y-%m-%d")
        if nome_slug:
            output_path = input_path.parent / f"report-{cnpj}-{nome_slug}-{date_str}.pdf"
        else:
            output_path = input_path.parent / f"report-{cnpj}-{date_str}.pdf"

    # Deep copy before generate_report_b2g mutates data (filters descartados, encerrados)
    if args.save_json:
        save_data = copy.deepcopy(data)
        _backfill_recommendations(save_data)

    buffer = generate_report_b2g(data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "wb") as f:
        f.write(buffer.getvalue())

    # F37: Post-generation PDF validation
    if not output_path.exists():
        print(f"ERROR: PDF not generated at {output_path}")
        sys.exit(1)
    file_size = output_path.stat().st_size
    if file_size < 10_000:
        print(f"WARNING: PDF suspiciously small ({file_size} bytes). May be corrupt.")
    n_recommended = len([e for e in data.get("editais", []) if e.get("recomendacao") != "NÃO RECOMENDADO"])
    min_pages = max(5, n_recommended)
    print(f"PDF generated: {output_path} ({file_size:,} bytes)")
    print(f"Size: {file_size / 1024:.1f} KB")

    # Save enriched JSON back to input (preserves backfilled recommendations + justificativas)
    if args.save_json:
        with open(args.input, "w", encoding="utf-8") as f:
            json.dump(save_data, f, ensure_ascii=False, indent=2)
        print(f"  JSON atualizado: {args.input}")

    excel_path = str(output_path).replace(".pdf", ".xlsx")
    generate_excel_companion(data, excel_path)
    print(f"Excel generated: {excel_path}")


if __name__ == "__main__":
    main()


# ============================================================
# JSON INPUT SCHEMA (for agents generating the data file)
# ============================================================
"""
{
  "empresa": {
    "cnpj": "12.345.678/0001-90",
    "razao_social": "Empresa LTDA",
    "nome_fantasia": "Empresa",
    "cnae_principal": "4120400 - Construção de edifícios",
    "cnaes_secundarios": "4211101, 4213800",
    "porte": "EPP",
    "capital_social": 1500000.00,
    "cidade_sede": "Florianópolis",
    "uf_sede": "SC",
    "situacao_cadastral": "ATIVA",
    "email": "contato@empresa.com",
    "telefones": ["(48) 99999-9999"],
    "qsa": [
      {"nome": "João Silva", "qualificacao": "Sócio-Administrador"}
    ],
    "sancoes": {"ceis": false, "cnep": false, "cepim": false, "ceaf": false},
    "historico_contratos": [
      {"orgao": "Prefeitura X", "valor": 500000, "data": "2025-06-01"}
    ]
  },
  "setor": "Engenharia e Construção Civil",
  "keywords": ["construção", "obra", "reforma", "edificação"],
  "editais": [
    {
      "objeto": "Contratação de empresa para reforma do prédio...",
      "orgao": "Prefeitura Municipal de Florianópolis",
      "uf": "SC",
      "municipio": "Florianópolis",
      "valor_estimado": 1500000.00,
      "modalidade": "Pregão Eletrônico",
      "data_abertura": "2026-03-15",
      "data_encerramento": "2026-03-25",
      "dias_restantes": 15,
      "fonte": "PNCP",
      "link": "https://pncp.gov.br/...",
      "recomendacao": "PARTICIPAR",
      "justificativa": "Objeto altamente aderente ao perfil da empresa...",
      "analise": {
        "aderencia": "Alta - objeto 100% compatível com CNAE principal 4120400",
        "valor": "Dentro da faixa operacional (capital R$1.5M, contrato R$1.5M)",
        "geografica": "Mesmo município da sede - custo logístico mínimo",
        "prazo": "15 dias restantes - tempo adequado para preparação",
        "modalidade": "Pregão Eletrônico - disputa por menor preço",
        "competitividade": "Órgão tem histórico de 3-5 participantes por edital",
        "riscos": "Valor no limite do capital social - avaliar BDI com cuidado"
      },
      "perguntas_decisor": {
        "Vale a pena participar?": "Sim. Objeto altamente aderente...",
        "Quanto eu deveria ofertar?": "Baseado no histórico do órgão...",
        "Quem são os concorrentes prováveis?": "Empresas locais de porte similar...",
        "Quais documentos preciso preparar?": "CND, certidões negativas...",
        "Qual o risco de não conseguir executar?": "Baixo, considerando...",
        "Esse órgão paga em dia?": "Histórico indica pagamento em 30-45 dias...",
        "Existe restrição que me impeça?": "Nenhuma restrição identificada..."
      }
    }
  ],
  "resumo_executivo": {
    "texto": "Foram identificadas X oportunidades abertas...",
    "destaques": [
      "3 editais com alta aderência ao perfil da empresa",
      "Valor total em jogo: R$ X milhões"
    ]
  },
  "inteligencia_mercado": {
    "panorama": "O setor de engenharia em SC apresenta...",
    "tendencias": "- Pregão eletrônico domina (78% das modalidades)...",
    "vantagens": "- Localização estratégica em Florianópolis...",
    "recomendacao_geral": "Focar nos 3 editais de maior aderência..."
  },
  "querido_diario": [
    {
      "data": "2026-03-08",
      "territorio": "Florianópolis - SC",
      "excerpts": [{"text": "...trecho do diário oficial mencionando..."}]
    }
  ],
  "proximos_passos": [
    {"acao": "Preparar documentação para edital X", "prazo": "5 dias", "prioridade": "ALTA"},
    {"acao": "Agendar visita técnica ao local", "prazo": "3 dias", "prioridade": "MÉDIA"}
  ]
}
"""
