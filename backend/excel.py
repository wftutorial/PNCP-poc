"""
Gerador de planilhas Excel formatadas para licitações de uniformes.

Este módulo é responsável por criar arquivos Excel profissionalmente formatados
com as licitações filtradas, incluindo:
- Header verde (#2E7D32) com texto branco
- 11 colunas com larguras otimizadas
- Formatação de moeda (R$), datas e hyperlinks
- Linha de totais com fórmula SUM
- Aba de metadados com estatísticas da busca

Exemplo de uso:
    >>> from excel import create_excel
    >>> licitacoes = [{"codigoCompra": "123", "objetoCompra": "Uniformes", ...}]
    >>> buffer = create_excel(licitacoes)
    >>> with open("output.xlsx", "wb") as f:
    ...     f.write(buffer.getvalue())
"""

import re
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Regex matching illegal XML characters that openpyxl rejects
# Includes: \x00-\x08, \x0b-\x0c, \x0e-\x1f (control chars except tab, LF, CR)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def sanitize_for_excel(value: str | None) -> str:
    """
    Remove illegal XML/Excel control characters from strings.

    openpyxl raises IllegalCharacterError for control characters like:
    - \x00-\x08 (NUL to BS)
    - \x0b-\x0c (VT, FF)
    - \x0e-\x1f (SO to US, includes \x13 Device Control 3)

    This is commonly seen in PNCP data where em-dashes (–) are incorrectly
    encoded as \x13 control characters.

    Args:
        value: String to sanitize, or None

    Returns:
        Sanitized string with illegal characters replaced by space, or empty string if None

    Examples:
        >>> sanitize_for_excel("AME \x13 SUL")
        'AME   SUL'
        >>> sanitize_for_excel(None)
        ''
    """
    if value is None:
        return ""
    if not isinstance(value, str):
        return str(value)
    # Replace illegal control characters with a space (preserves readability)
    return ILLEGAL_CHARACTERS_RE.sub(' ', value)


def create_excel(licitacoes: list[dict], paywall_preview: bool = False, total_before_paywall: int | None = None, org_name: str | None = None) -> BytesIO:
    """
    Gera planilha Excel formatada com licitações filtradas.

    Args:
        licitacoes: Lista de dicionários com dados das licitações do PNCP

    Returns:
        BytesIO: Buffer com arquivo Excel pronto para download/salvamento

    Raises:
        ValueError: Se licitacoes não for uma lista
    """
    if not isinstance(licitacoes, list):
        raise ValueError("licitacoes deve ser uma lista")

    wb = Workbook()
    ws = wb.active
    ws.title = "Licitações Uniformes"

    # === ESTILOS ===
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2E7D32", end_color="2E7D32", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="top", wrap_text=True)
    currency_format = '[$R$-416] #.##0,00'
    date_format = "DD/MM/YYYY"
    datetime_format = "DD/MM/YYYY HH:MM"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # === HEADERS ===
    headers = [
        ("Código PNCP", 25),
        ("Objeto", 60),
        ("Órgão", 40),
        ("UF", 6),
        ("Município", 20),
        ("Valor Estimado", 18),
        ("Modalidade", 20),
        ("Publicação", 12),
        ("Início", 16),
        ("Situação", 15),
        ("Link", 15),
    ]

    for col, (header_name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Congelar header
    ws.freeze_panes = "A2"

    # === DADOS ===
    for row_idx, lic in enumerate(licitacoes, start=2):
        # A: Código PNCP
        ws.cell(row=row_idx, column=1, value=sanitize_for_excel(lic.get("codigoCompra")))

        # B: Objeto (sanitize to remove illegal control chars like \x13)
        ws.cell(row=row_idx, column=2, value=sanitize_for_excel(lic.get("objetoCompra")))

        # C: Órgão
        ws.cell(row=row_idx, column=3, value=sanitize_for_excel(lic.get("nomeOrgao")))

        # D: UF
        ws.cell(row=row_idx, column=4, value=sanitize_for_excel(lic.get("uf")))

        # E: Município
        ws.cell(row=row_idx, column=5, value=sanitize_for_excel(lic.get("municipio")))

        # F: Valor (formatado como moeda)
        valor_cell = ws.cell(row=row_idx, column=6, value=lic.get("valorTotalEstimado"))
        valor_cell.number_format = currency_format

        # G: Modalidade
        ws.cell(row=row_idx, column=7, value=sanitize_for_excel(lic.get("modalidadeNome")))

        # H: Data Publicação
        data_pub = parse_datetime(lic.get("dataPublicacaoPncp"))
        pub_cell = ws.cell(row=row_idx, column=8, value=data_pub)
        if data_pub:
            pub_cell.number_format = date_format

        # I: Data Abertura
        data_abertura = parse_datetime(lic.get("dataAberturaProposta"))
        abertura_cell = ws.cell(row=row_idx, column=9, value=data_abertura)
        if data_abertura:
            abertura_cell.number_format = datetime_format

        # J: Situação
        ws.cell(row=row_idx, column=10, value=sanitize_for_excel(lic.get("situacaoCompraNome")))

        # K: Link (hyperlink)
        # CRIT-FLT-008: linkSistemaOrigem (86% populated) > linkProcessoEletronico (0% — dead field) > fallback PNCP
        link = lic.get("linkSistemaOrigem") or lic.get("linkProcessoEletronico")

        # Fallback: construir URL do PNCP a partir do numeroControlePNCP
        # Formato numeroControlePNCP: {CNPJ}-{TIPO}-{SEQUENCIAL}/{ANO}
        # Formato URL PNCP: /editais/{CNPJ}/{ANO}/{SEQUENCIAL_SEM_ZEROS}
        if not link:
            numero_controle = lic.get("numeroControlePNCP", "")
            if numero_controle:
                try:
                    # Parse: "67366310000103-1-000189/2025" -> cnpj=67366310000103, ano=2025, seq=189
                    partes = numero_controle.split("/")
                    if len(partes) != 2:
                        raise ValueError("Formato inválido: esperado 'xxx/ano'")

                    ano = partes[1]
                    cnpj_tipo_seq = partes[0].split("-")

                    if len(cnpj_tipo_seq) < 3:
                        raise ValueError("Formato inválido: esperado 'cnpj-tipo-seq'")

                    cnpj = cnpj_tipo_seq[0]
                    sequencial = cnpj_tipo_seq[2].lstrip("0")

                    if cnpj and ano and sequencial:
                        link = f"https://pncp.gov.br/app/editais/{cnpj}/{ano}/{sequencial}"
                    else:
                        raise ValueError("Componentes vazios após parsing")

                except (IndexError, AttributeError, ValueError):
                    # Se parsing falhar, usar busca genérica
                    link = f"https://pncp.gov.br/app/editais?q={numero_controle}"

        link_cell = ws.cell(row=row_idx, column=11, value="Abrir")
        link_cell.hyperlink = link or "https://pncp.gov.br/app/editais"
        link_cell.font = Font(color="0563C1", underline="single")

        # Aplicar bordas e alinhamento em todas as células da linha
        for col in range(1, 12):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = cell_alignment

    # === LINHA DE TOTAIS ===
    if licitacoes:  # Só adiciona linha de totais se houver dados
        total_row = len(licitacoes) + 2
        ws.cell(row=total_row, column=5, value="TOTAL:").font = Font(bold=True)

        total_cell = ws.cell(
            row=total_row, column=6, value=f"=SUM(F2:F{total_row - 1})"
        )
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True)

    # === METADATA (aba separada) ===
    ws_meta = wb.create_sheet("Metadata")
    meta_row = 1
    # STORY-322 AC23: Include org name in metadata header
    if org_name:
        ws_meta[f"A{meta_row}"] = "Organização:"
        ws_meta[f"B{meta_row}"] = org_name
        ws_meta[f"B{meta_row}"].font = Font(bold=True, size=12)
        meta_row += 1
    ws_meta[f"A{meta_row}"] = "Gerado em:"
    ws_meta[f"B{meta_row}"] = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M:%S")
    meta_row += 1
    ws_meta[f"A{meta_row}"] = "Total de licitações:"
    ws_meta[f"B{meta_row}"] = len(licitacoes)
    meta_row += 1
    ws_meta[f"A{meta_row}"] = "Valor total estimado:"
    ws_meta[f"B{meta_row}"] = sum(lic.get("valorTotalEstimado", 0) or 0 for lic in licitacoes)
    ws_meta[f"B{meta_row}"].number_format = currency_format

    # STORY-320 AC4: Upsell sheet when paywall preview is active
    if paywall_preview and total_before_paywall and total_before_paywall > len(licitacoes):
        remaining = total_before_paywall - len(licitacoes)
        ws_upsell = wb.create_sheet("Desbloqueie Mais")
        ws_upsell["A1"] = f"Desbloqueie {remaining} resultados adicionais com SmartLic Pro"
        ws_upsell["A1"].font = Font(bold=True, size=14, color="1A237E")
        ws_upsell["A3"] = "Este arquivo contem uma preview com os primeiros 10 resultados."
        ws_upsell["A4"] = f"Com o SmartLic Pro, voce tera acesso a todos os {total_before_paywall} resultados."
        ws_upsell["A6"] = "Assine agora: https://smartlic.tech/planos"
        ws_upsell["A6"].font = Font(color="0563C1", underline="single")
        ws_upsell["A6"].hyperlink = "https://smartlic.tech/planos"
        ws_upsell.column_dimensions["A"].width = 80

    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def parse_datetime(value: str | None) -> datetime | None:
    """
    Parse datetime string do formato PNCP para objeto datetime.

    Tenta múltiplos formatos comuns:
    - ISO 8601 com timezone (2024-01-25T10:30:00Z)
    - ISO 8601 sem timezone (2024-01-25T10:30:00)
    - Data simples (2024-01-25)

    IMPORTANTE: Excel não suporta timezone-aware datetimes, então sempre
    retornamos naive datetime (tzinfo=None) mesmo para strings com timezone.

    Args:
        value: String de data/datetime do PNCP

    Returns:
        datetime object ou None se parsing falhar

    Examples:
        >>> parse_datetime("2024-01-25T10:30:00Z")
        datetime(2024, 1, 25, 10, 30, 0)
        >>> parse_datetime("2024-01-25")
        datetime(2024, 1, 25, 0, 0, 0)
        >>> parse_datetime(None)
        None
    """
    if not value:
        return None

    try:
        # Formato ISO com timezone (Z = UTC)
        # Excel não suporta timezone, então removemos tzinfo
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except (ValueError, AttributeError):
        pass

    try:
        # Formato sem timezone
        return datetime.strptime(value, "%Y-%m-%dT%H:%M:%S")
    except (ValueError, AttributeError):
        pass

    try:
        # Apenas data
        return datetime.strptime(value, "%Y-%m-%d")
    except (ValueError, AttributeError):
        return None
